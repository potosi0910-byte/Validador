from io import BytesIO
from datetime import datetime, timedelta
import unicodedata
import re
import os
 
from flask import Flask, render_template, request, send_file
import json
 #Dependencia para manejo de Excel: pip install openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from auditoria import validar_auditoria
 
app = Flask(__name__)
 
 
# ══════════════════════════════════════════════════════════════
# UTILIDADES GENERALES
# ══════════════════════════════════════════════════════════════
 
def normalizar_str(v):
    """Convierte cualquier valor a string limpio."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()
 
 
def quitar_tildes(texto):
    """Elimina diacríticos para comparaciones robustas."""
    nfkd = unicodedata.normalize('NFD', str(texto).upper())
    return ''.join(c for c in nfkd if unicodedata.category(c) != 'Mn')
#Importante: quitar_tildes se usa para matching de encabezados, no para mostrar datos al usuario.
 
def separar_tipo_num(valor_raw):
    """
    Separa 'CC 25283017'  → ('CC', '25283017')
           'TI  123456'   → ('TI', '123456')
           'CC25283017'   → ('CC', '25283017')
    """
    s = normalizar_str(valor_raw)
    m = re.match(r'^([A-Za-z]+)\s*(\d+)$', s)
    if m:
        return m.group(1).upper(), m.group(2)
    partes = s.split()
    if len(partes) >= 2:
        return partes[0].upper(), partes[1]
    return "", s
 
 
def encontrar_col(headers, palabras_clave):
    """
    Busca el índice de columna cuyo encabezado contenga todas las palabras clave.
    # Diferencias de encoding pueden causar tildes corruptas, por eso se usa quitar_tildes.
    Usa matching sin tildes + matching parcial por sufijo para tolerar encodings
    corruptos (p.ej. 'NÃšMERO' → se detecta por sufijo 'MERO').
    """
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_upper = quitar_tildes(normalizar_str(h))
        h_raw   = str(h).upper()
        coincide = True
        for kw in palabras_clave:
            kw_clean = quitar_tildes(kw.upper())
            # match directo sin tildes
            if kw_clean in h_upper:
                continue
            # match parcial: últimos 4 caracteres (útil para 'NUMERO' → 'MERO')
            if len(kw_clean) >= 4 and kw_clean[-4:] in h_raw:
                continue
            coincide = False
            break
        if coincide:
            return i
    return None


# ══════════════════════════════════════════════════════════════
# HELPERS PARA NOMBRES DE CAMPO RIPS (dos variantes en uso)
# ══════════════════════════════════════════════════════════════

def _tipo_doc(obj):
    """Lee tipoDocumento... probando ambas variantes de nombre RIPS."""
    return normalizar_str(
        obj.get("tipoDocumentoldentificacion") or
        obj.get("tipoDocumentoIdentificacion") or ""
#Importante: el tipo de documento se muestra al usuario sin normalizar, para respetar tildes y mayúsculas originales. La normalización se usa solo para matching entre RIPS y Excel, no para mostrar datos.
    )

def _num_doc_val(obj):
    """Lee numDocumento... probando ambas variantes de nombre RIPS."""
    return normalizar_str(
        obj.get("numDocumentoldentificacion") or
        obj.get("numDocumentoIdentificacion") or ""
    )

def _nit_obligado(obj):
    """Lee NIT del facturador — soporta ambas variantes del campo raíz."""
    return normalizar_str(
        obj.get("numDocumentoldObligado") or
        obj.get("numDocumentoIdObligado") or ""
    )


# ══════════════════════════════════════════════════════════════
# CARGA DE EXCEL DE AUTORIZACIONES (uno o varios archivos)
# ══════════════════════════════════════════════════════════════
 
def cargar_excel_autorizaciones(archivos_excel):
    """
    Lee uno o más Excel de autorizaciones EPS.
 
    Columnas requeridas en Excel:
      - 'TIPO ID AFILIADO'  → 'CC 25283017' → tipo_doc='CC', num_doc='25283017'
      - 'NÚMERO'            → número de autorización
 
    Retorna:
      registros_excel : dict { num_doc → list[{tipo_doc, numero_aut, archivo}] }
      set_aut_excel   : set  { numero_aut, ... }
      errores         : list[str]
    """
    registros_excel = {}
    set_aut_excel   = set()
    errores         = []
 
    for archivo in archivos_excel:
        if not archivo or archivo.filename == "":
            continue
        try:
            archivo.seek(0)
            wb = load_workbook(archivo, data_only=True, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
 
            col_tipo_id      = encontrar_col(headers, ["TIPO", "AFILIADO"])
            col_numero       = encontrar_col(headers, ["NUMERO"])
            col_codigo       = encontrar_col(headers, ["CODIGO"])
            # Nuevas columnas para hospitalización
            col_dias         = encontrar_col(headers, ["DIAS"])
            col_fecha_emision = encontrar_col(headers, ["FECHA", "EMISION"])

            if col_tipo_id is None:
#Error handling mejorado: si no se encuentra una columna requerida, se agrega un mensaje de error específico que incluye el nombre del archivo y los encabezados encontrados. Esto ayuda a identificar rápidamente problemas de formato en los archivos Excel cargados.
                errores.append(
                    f"[{archivo.filename}] No se encontró la columna 'TIPO ID AFILIADO'. "
                    f"Encabezados: {[h for h in headers if h]}"
                )
                continue
            if col_numero is None:
                errores.append(
                    f"[{archivo.filename}] No se encontró la columna 'NÚMERO'. "
                    f"Encabezados: {[h for h in headers if h]}"
                )
                continue

            filas_cargadas = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                tipo_id_raw = row[col_tipo_id]
                numero_raw  = row[col_numero]

                if tipo_id_raw is None:
                    continue

                tipo_doc, num_doc = separar_tipo_num(tipo_id_raw)
                numero_aut  = normalizar_str(numero_raw)
                codigo_eps  = normalizar_str(row[col_codigo])       if col_codigo       is not None else ""
                dias_aut    = normalizar_str(row[col_dias])          if col_dias         is not None else ""
                fecha_emis  = row[col_fecha_emision]                 if col_fecha_emision is not None else None

                if not num_doc:
                    continue

                registros_excel.setdefault(num_doc, []).append({
                    'tipo_doc':         tipo_doc,
                    'numero_aut':       numero_aut,
                    'codigo':           codigo_eps,
                    'archivo':          archivo.filename,
                    'dias_autorizados': dias_aut,    # Días autorizados de hospitalización
                    'fecha_emision':    fecha_emis,  # Fecha de emisión de la autorización
                })

                if numero_aut:
                    set_aut_excel.add(numero_aut)

                filas_cargadas += 1

            wb.close()

        except Exception as e:
            errores.append(f"[{archivo.filename}] Error leyendo Excel: {e}")

    return registros_excel, set_aut_excel, errores
 
 
# ══════════════════════════════════════════════════════════════
# EXTRACCIÓN DE MEDICAMENTOS INVÁLIDOS DEL RIPS (JSON)
# ══════════════════════════════════════════════════════════════
 
def extraer_medicamentos_invalidos(data, nombre_archivo=""):
    """
    Extrae SOLO medicamentos donde:
      - nomTecnologiaSalud sea vacío/None
      - o el código sea '000', vacío o None
    """
    resultados = []
    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
 
    def buscar_factura(d):
        for k in FACTURA_KEYS:
            if k in d:
                v = d.get(k)
                if isinstance(v, dict):
                    for kk in FACTURA_KEYS:
                        if kk in v:
                            return normalizar_str(v.get(kk))
                    return ""
                return normalizar_str(v)
        return ""
 
    def recorrer(nodo, factura_actual="", en_medicamentos=False):
        if isinstance(nodo, dict):
            f = buscar_factura(nodo)
            if f:
#RIPS puede tener facturas a nivel raíz o dentro de secciones, así que se actualiza factura_actual cada vez que se encuentra una nueva factura en el nodo actual. Esto asegura que los medicamentos extraídos se asocien con la factura correcta, incluso si hay múltiples facturas en el mismo JSON.                
                factura_actual = f
 
            if en_medicamentos and "vrServicio" in nodo and "consecutivo" in nodo:
                idrips  = f"1.2.{nodo.get('consecutivo')}"
                codigo  = normalizar_str(
                    nodo.get("codTecnologiaSalud")
                    or nodo.get("codMedicamento")
                    or nodo.get("codProcedimiento")
                )
                nombre  = normalizar_str(nodo.get("nomTecnologiaSalud"))
 
                if nombre == "" or nombre.lower() == "none" or codigo in {"", "000", "0"}:
                    resultados.append({
                        "archivo":            nombre_archivo,
                        "numeroFactura":      factura_actual,
                        "idrips":             idrips,
                        "codConsulta":        codigo if codigo else "000",
                        "nomTecnologiaSalud": nombre if nombre else "SIN NOMBRE",
                        "vrServicio":         nodo.get("vrServicio", "")
                    })
 
            for clave, valor in nodo.items():
                if clave == "medicamentos":
                    recorrer(valor, factura_actual, True)
                else:
                    recorrer(valor, factura_actual, en_medicamentos)
 
        elif isinstance(nodo, list):
            for item in nodo:
                recorrer(item, factura_actual, en_medicamentos)
 
    recorrer(data)
    return resultados
 
 
# ══════════════════════════════════════════════════════════════
# EXTRACCIÓN DE AUTORIZACIONES DEL RIPS (JSON)
# ══════════════════════════════════════════════════════════════

def extraer_autorizaciones_rips(data, nombre_archivo=""):
    """
    Por cada usuario del RIPS:
      - Captura num_doc y tipo_doc del PACIENTE (nivel usuarios[i]).
      - Detecta tipo_atencion:
          'hospitalario' → tiene registros en urgencias O hospitalizacion.
          'ambulatorio'  → sin urgencias ni hospitalizacion.
      - Recopila todos los pares (numAutorizacion, codigoServicio) de:
          consultas, procedimientos, medicamentos, otrosServicios.

    Retorna dict { num_doc → {
        tipo_doc, num_factura, archivo_rips, tipo_atencion,
        set_auths: set de numAutorizacion,
        codigos_por_auth: { auth → set(codigos) }
    }}
    Clave = num_doc (no auth) para poder cruzar desde Excel por paciente.
    """
    pacientes = {}
    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    # Incluye urgencias para capturar sus autorizaciones (necesario para regla solo-urgencias)
    SECCIONES    = [
        ("consultas",       "codConsulta"),
        ("procedimientos",  "codProcedimiento"),
        ("medicamentos",    "codTecnologiaSalud"),
        ("otrosServicios",  "codTecnologiaSalud"),
        ("urgencias",       "codDiagnosticoPrincipal"),  # Captura auths de urgencias
    ]

    def buscar_factura_top(d):
        for k in FACTURA_KEYS:
            if k in d and not isinstance(d[k], (dict, list)):
                return normalizar_str(d[k])
        return ""

    if not isinstance(data, dict):
        return pacientes

    num_factura = buscar_factura_top(data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue

        num_doc  = normalizar_str(usuario.get("numDocumentoIdentificacion"))
        tipo_doc = normalizar_str(usuario.get("tipoDocumentoIdentificacion"))
        servicios = usuario.get("servicios", {}) if isinstance(usuario.get("servicios"), dict) else {}

        # Tipo de atención: hospitalario si hay urgencias O hospitalizacion
        tiene_urg  = bool(isinstance(servicios.get("urgencias"), list)       and servicios["urgencias"])
        tiene_hosp = bool(isinstance(servicios.get("hospitalizacion"), list) and servicios["hospitalizacion"])
        tipo_atencion = "hospitalario" if (tiene_urg or tiene_hosp) else "ambulatorio"

        if num_doc not in pacientes:
            pacientes[num_doc] = {
                "tipo_doc":             tipo_doc,
                "num_factura":          num_factura,
                "archivo_rips":         nombre_archivo,
                "tipo_atencion":        tipo_atencion,
                "tiene_urg":            tiene_urg,
                "tiene_hosp":           tiene_hosp,
                "tiene_solo_urgencias": tiene_urg and not tiene_hosp,
                "set_auths":            set(),
                "codigos_por_auth":     {},
                # Datos de procedimientos para regla de urgencias (código > 870000 no requiere autorizacion)
                "procedimientos_pac":   [],
                # Fechas de hospitalización para cálculo de días de estancia
                "fecha_inicio_hosp":    None,
                "fecha_egreso_hosp":    None,
                "fecha_inicio_urg":     None,
                "fecha_egreso_urg":     None,
                # Internacion: autorizaciones y conteo de registros hospitalizacion
                "auths_hosp":           set(),
                "n_hosp_regs":          0,
                # Lista detallada de registros hospitalización {na, fecha_ini}
                "hosp_regs_lista":      [],
            }

        p = pacientes[num_doc]

        # ── Extraer fechas y autorizaciones de hospitalización ───────────────
        if tiene_hosp:
            hosp_regs = servicios.get("hospitalizacion", [])
            if isinstance(hosp_regs, list):
                for hreg in hosp_regs:
                    if not isinstance(hreg, dict):
                        continue
                    p["n_hosp_regs"] += 1
                    fi   = hreg.get("fechaInicioAtencion") or hreg.get("fechaInicio")
                    fe   = hreg.get("fechaEgreso")
                    na_h = normalizar_str(hreg.get("numAutorizacion", ""))
                    if fi and not p["fecha_inicio_hosp"]:
                        p["fecha_inicio_hosp"] = normalizar_str(fi)
                    if fe and not p["fecha_egreso_hosp"]:
                        p["fecha_egreso_hosp"] = normalizar_str(fe)
                    if na_h:
                        p["auths_hosp"].add(na_h)
                    p["hosp_regs_lista"].append({
                        "na":        na_h,
                        "fecha_ini": normalizar_str(fi or ""),
                    })

        # ── Extraer fechas de urgencias ──────────────────────────────
        if tiene_urg:
            urg_regs = servicios.get("urgencias", [])
            if isinstance(urg_regs, list):
                for ureg in urg_regs:
                    if not isinstance(ureg, dict):
                        continue
                    fi_u = ureg.get("fechaInicioAtencion") or ureg.get("fechaInicio")
                    fe_u = ureg.get("fechaEgreso")
                    if fi_u and not p["fecha_inicio_urg"]:
                        p["fecha_inicio_urg"] = normalizar_str(fi_u)
                    if fe_u:
                        p["fecha_egreso_urg"] = normalizar_str(fe_u)
 
        # ── Extraer procedimientos (todas las atenciones) ─────────────────
        proc_regs = servicios.get("procedimientos", [])
        if isinstance(proc_regs, list):
            for preg in proc_regs:
                if not isinstance(preg, dict):
#Para la regla de cirugía ambulatoria solo se consideran procedimientos con código válido, por eso se omiten registros sin codProcedimiento o con código vacío. Esto evita que registros incompletos o mal formateados afecten la clasificación de tipo de atención.
                    continue
                cod_p   = normalizar_str(preg.get("codProcedimiento", ""))
                na_p    = normalizar_str(preg.get("numAutorizacion", ""))
                fi_p    = normalizar_str(preg.get("fechaInicioAtencion") or preg.get("fechaInicio") or "")
                if cod_p:
                    p["procedimientos_pac"].append({"cod": cod_p, "num_aut": na_p, "fecha_inicio": fi_p})
 
        # ── Capturar registros de internación en otrosServicios ──────────────
        os_regs = servicios.get("otrosServicios", [])
        if isinstance(os_regs, list):
            for _os in os_regs:
                if not isinstance(_os, dict):
                    continue
                _nom = str(_os.get("nomTecnologiaSalud", "")).upper()
                if _nom.startswith("INTERNAC"):
                    p.setdefault("os_internacion", []).append({
                        "num_aut": normalizar_str(_os.get("numAutorizacion", "")),
                        "cod":     normalizar_str(_os.get("codTecnologiaSalud", "")),
                        "nom":     _os.get("nomTecnologiaSalud", ""),
                    })

        # ── Reclasificar a cirugía ambulatoria si aplica ─────────────────
        # Condición: tiene hospitalizacion SIN urgencias + al menos un codProcedimiento < 870000
        # Si hay urgencias el paciente es siempre hospitalario (ingreso de emergencia).
        if tiene_hosp and not tiene_urg and p["procedimientos_pac"]:
            for _pr in p["procedimientos_pac"]:
                try:
                    if int(_pr.get("cod", "")) < 870000:
                        p["tipo_atencion"] = "cirugia_ambulatoria"
                        break
                except (ValueError, TypeError):
                    pass

        # ── Recopilar autorizaciones de todas las secciones ──────────────────
        for sec_name, cod_field in SECCIONES:
            registros = servicios.get(sec_name, [])
            if not isinstance(registros, list):
                continue
            for reg in registros:
                if not isinstance(reg, dict):
                    continue
                na  = normalizar_str(reg.get("numAutorizacion"))
                cod = normalizar_str(
                    reg.get(cod_field)
                    or reg.get("codConsulta")
                    or reg.get("codProcedimiento")
                )
                if not na:
                    continue
                p["set_auths"].add(na)
                if na not in p["codigos_por_auth"]:
                    p["codigos_por_auth"][na] = set()
                if cod:
                    p["codigos_por_auth"][na].add(cod)

    return pacientes



# ══════════════════════════════════════════════════════════════
# CONTEO DE REGISTROS RIPS
# ══════════════════════════════════════════════════════════════
 
def contar_registros_rips(data):
    """
    Suma el total de registros en todas las secciones de servicios
    de todos los usuarios del RIPS JSON.
    """
    SECCIONES = ("consultas","procedimientos","medicamentos","urgencias",
                 "hospitalizacion","recienNacidos","otrosServicios")
    total = 0
    if not isinstance(data, dict):
        return total
    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        for sec in SECCIONES:
            lst = servicios.get(sec, [])
            if isinstance(lst, list):
                total += len(lst)
    return total
 
# ══════════════════════════════════════════════════════════════
# CÁLCULO DE DÍAS DE ESTANCIA HOSPITALARIA
# ══════════════════════════════════════════════════════════════

def calcular_dias_estancia(fecha_inicio_str, fecha_egreso_str):
    """
    Calcula los días de estancia facturables según las reglas hospitalarias:

      1. El primer día SIEMPRE se cuenta, sin importar la hora de ingreso.
      2. El último día (egreso) NO se cobra si la hora es exactamente 00:00.
      3. Si el egreso es posterior a las 00:00 (cualquier hora > medianoche),
         ese día SÍ se cuenta como día adicional.

    Ejemplos:
      Ingreso 10-sep, egreso 13-sep 14:38 → (13-10)=3 días base + 1 (14:38>00:00) = 4 días
      Ingreso 10-sep, egreso 13-sep 00:00 → 3 días (egreso exacto a medianoche, no suma)
      Ingreso 10-sep, egreso 10-sep 14:38 → 0 + 1 = 1 día (primer día siempre cuenta)

    Retorna el número de días facturables (int), o None si no se pueden parsear las fechas.
    """
    from datetime import time as dtime
    try:
        inicio = datetime.strptime(str(fecha_inicio_str).strip()[:16], "%Y-%m-%d %H:%M")
        egreso = datetime.strptime(str(fecha_egreso_str).strip()[:16], "%Y-%m-%d %H:%M")

        # Diferencia en días calendario (sin contar el día de egreso por defecto)
        dias = (egreso.date() - inicio.date()).days

        # Si la hora de egreso es mayor a 00:00:00, el día de egreso también se factura
        if egreso.time() > dtime(0, 0):
            dias += 1

        # El primer día siempre se cuenta (garantiza mínimo 1)
        return max(1, dias)
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════
# VALIDACIONES DE AUTORIZACIONES
# ══════════════════════════════════════════════════════════════

CODIGO_URGENCIAS = "890701"   # Código CUPS de urgencias en el Excel

# ── Catálogos estáticos Malla 2275/2023 ──────────────────────────────────────
TIPOS_MEDICAMENTO_VALIDOS   = {"01", "02", "03", "04"}          # TipoMedicamentoPOSVersion2
TIPOS_DOC_PRESCRIPTOR_MED   = {"CC", "CE", "CD", "PA", "SC",
                                "PE", "DE", "PT"}                # M16
CONCEPTOS_RECAUDO_MED       = {"01", "02", "03", "05"}          # M20 (04=Anticipos excluido)

# ── Catálogos adicionales Malla 2275/2023 ─────────────────────────────────────
TIPOS_DOC_USUARIO         = {"CC","TI","RC","CN","CE","PA","MS","AS","CD","SC","PE","DE","PT"}
TIPOS_DOC_PROFESIONAL     = {"CC","CE","CD","PA","SC","PE","DE","PT"}
# Catálogo Sexo: M=Masculino, F=Femenino, I=Indeterminado
CODIGOS_SEXO              = {"M","F","I"}
# Finalidades IVE (Interrupción Voluntaria del Embarazo)
_FINALIDAD_IVE            = {"34","35","36","49"}

def _cargar_sets_sexo_cie10():
    """Lee CIE10.json y construye sets de categorías femeninas y de pene."""
    import os as _os, json as _j
    ruta = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "CIE10.json")
    try:
        with open(ruta, encoding="utf-8") as _f:
            _datos = _j.load(_f)["datos"]
    except Exception:
        return frozenset(), frozenset()
    _KW_F = (
        "EMBARAZ","PARTO","PUERPERIO",
        "UTERO","UTERINO","UTERINA",
        "OVARIO","OOFORIT","SALPINGIT","TROMPA DE FALOPIO","LIGAMENTO ANCHO",
        "VAGINA","VULVA","BARTHOLIN",
        "PLACENTA","ABORTO",
        "MENSTR","MENOPAUS","FEMENIN",
        "MATERNA","MATERNO","OBSTETRIC",
        "ENDOMETRI","LEIOMIO",
        "CUELLO DEL UTERO","CUELLO UTERINO",
        "INFERTILIDAD FEMENINA","FECUNDACION ARTIFICIAL",
        "GINECOL","POSTPARTO",
    )
    _KW_PENE = (" PENE", "DEL PENE", "PREPUCIO", "FIMOSIS", "PARAFIMOSIS")
    _visto, _cats_f, _cats_pene = set(), set(), set()
    for _d in _datos:
        _c3 = _d.get("codigo_categoria_3", "")
        if _c3 in _visto:
            continue
        _visto.add(_c3)
        _desc = _d.get("descripcion_categoria_3", "").upper()
        if any(_k in _desc for _k in _KW_F):
            _cats_f.add(_c3)
        if any(_k in _desc for _k in _KW_PENE):
            _cats_pene.add(_c3)
    # N70-N99: genitales femeninas (rango completo)
    for _c3 in _visto:
        if _c3.startswith("N") and "N70" <= _c3 <= "N99":
            _cats_f.add(_c3)
    # Códigos adicionales confirmados femeninos
    for _c3 in ("A34","D06","D39","F53","Q97","R87","Y76",
                "Z32","Z33","Z34","Z35","Z36","Z37","Z39"):
        if _c3 in _visto:
            _cats_f.add(_c3)
    # D29: neoplasias benignas genitales masculinos (incluye pene D29.0)
    if "D29" in _visto:
        _cats_pene.add("D29")
    return frozenset(_cats_f), frozenset(_cats_pene)

_DX_F_CATS, _DX_M_PENE_CATS = _cargar_sets_sexo_cie10()

# Campos de diagnóstico CIE-10 a revisar en cada registro de servicio
_CAMPOS_DX = (
    "codDiagnosticoPrincipal","codDiagnosticoRelacionado1",
    "codDiagnosticoRelacionado2","codDiagnosticoRelacionado3",
    "codDiagnosticoEgreso1","codDiagnosticoEgreso2",
    "codDiagnosticoEgreso3","codDiagnosticoEgreso4",
)
VALORES_SINO              = {"SI","NO"}
# Catálogo conceptoRecaudo: 01=Copago, 02=Cuota moderadora, 03=Planes voluntarios, 04=Anticipo, 05=No aplica
CONCEPTOS_RECAUDO_CONSULTA = {"01","02","03","05"}      # C18: excluye solo 04=Anticipo
CONCEPTOS_RECAUDO_PROC    = {"01","02","03","05"}        # P17: admite copago
TIPOS_DOC_RN              = {"CN","RC","MS"}             # N02: recién nacidos
COND_EGRESO_MUERTO        = {"02"}                       # CondicionyDestinoUsuarioEgreso 02=PACIENTE MUERTO
TIPOS_OS_CON_PRESCRIPTOR  = {"01","02","03"}             # S: DM/SC/Honorarios

# ── Catálogos de dominio completos (tablas de referencia RIPS v2) ─────────────
# TipoNota: NA=Nota ajuste, NC=Nota crédito, ND=Nota débito, RS=RIPS sin Factura
TIPOS_NOTA_VALIDOS        = {"NA","NC","ND","RS"}

# RIPSTipoUsuarioVersion2
TIPOS_USUARIO_RIPS        = {"01","02","03","04","05","06","07","08","09","10","11","12","13"}

# ZonaVersion2: 01=Rural, 02=Urbano
ZONAS_VALIDAS             = {"01","02"}

# ModalidadAtencion (nota: código 05 no existe en catálogo oficial)
MODALIDADES_ATENCION      = {"01","02","03","04","06","07","08","09"}

# GrupoServicios: 01=Consulta externa, 02=Apoyo diagnóstico, 03=Internación, 04=Quirúrgico, 05=Atención inmediata
GRUPOS_SERVICIOS          = {"01","02","03","04","05"}

# RIPSFinalidadConsultaVersion2: códigos 11..44
FINALIDADES_CONSULTA      = {str(i).zfill(2) for i in range(11, 45)}

# RIPSCausaExternaVersion2: códigos 21..49
CAUSAS_EXTERNAS           = {str(i) for i in range(21, 50)}

# CondicionyDestinoUsuarioEgreso: 01..08
COND_EGRESO_VALIDAS       = {"01","02","03","04","05","06","07","08"}

# TipoOtrosServicios: 01..06
TIPOS_OS_VALIDOS          = {"01","02","03","04","05","06"}

def validar_autorizaciones(pacientes_rips, registros_excel, set_aut_excel):
    """
    Valida autorizaciones comparando RIPS JSON contra base de datos EPS (Excel).

    REGLAS (nueva lógica unificada):
    ─ Identificación  : solo procesa pacientes cuyo num_doc esté en el Excel.
    ─ Tipo documento  : alerta si tipo_doc RIPS != tipo_doc Excel.
    ─ Clasificación   :
        cirugia_ambulatoria  → tiene hospitalizacion + procedimientos con codProcedimiento < 870000
        hospitalario         → tiene urgencias y/o hospitalizacion (sin ser cirugia_amb)
        ambulatorio          → sin urgencias ni hospitalizacion
    ─ Ambulatorio / Cirugía ambulatoria:
        Por cada (numAutorizacion, codProcedimiento) en procedimientos del RIPS:
        verificar que el par exista en la base del paciente.
        Alerta: "Ambulatorio - La autorización o el código del procedimiento
                 no coinciden con la base de Nueva EPS según RIPS"
    ─ Hospitalario:
        Extraer fechaIngresoInicial y fechaEgresoFinal del episodio.
        Por cada procedimiento: buscar (auth + cod) en base dentro del rango de fechas.
        Alerta: "hospitalario - La autorización y/o el procedimiento no coinciden
                 con la base dentro del rango de fechas de la atención hospitalaria"
        Adicionalmente: auths de la base en ese rango que NO estén en el RIPS.
        Alerta: "Hospitalario - Existen autorizaciones no relacionadas en el RIPS Json."
    """

    alertas = {
        'tipo_doc_mismatch':       [],   # Tipo de documento difiere entre RIPS y EPS
        'amb_par_no_cruza':        [],   # Ambulatorio/CirAmb: par (auth+cod) no encontrado en base
        'hosp_proc_no_cruza':      [],   # Hospitalario: (auth+cod) no cruza dentro del rango de fechas
        'hosp_aut_no_relacionada': [],   # Hospitalario: auth en base dentro del rango pero sin RIPS
        # Reglas de internación
        'estancia_sin_aut':        [],   # Hospitalizacion sin auth válida en Excel
        'proc_qx_misma_aut_hosp':  [],   # Proc quirúrgico (<870000) usa misma auth que estancia
        'sin_num_aut_relacionado': [],   # Sin ningún numAutorizacion (ni estancia ni proc <870000)
        # Nuevas validaciones
        'amb_aut_emision_posterior': [],  # Ambulatorio: auth emitida DESPUÉS del servicio
        'hosp_cod_sin_aut':          [],  # Hosp: registro de hospitalización sin numAutorizacion
        'hosp_proc_cod_no_cruza':    [],  # Hosp+Procs: cod <870000 no coincide con auth en base
        'hosp_cups_duplicado':       [],  # Hosp: código CUPS repetido más de una vez
        'proc_sin_aut_amb':          [],  # Ambulatorio: procedimiento sin numAutorizacion
        'proc_aut_no_cruza_amb':     [],  # Ambulatorio: auth no corresponde al CUPS según base EPS
        'cups_noestandar_sin_aut':   [],  # Código no estándar (>6 chars o con letras) sin auth
        # Nuevas reglas: paciente no está en base EPS
        'proc_sin_aut_no_excel':     [],  # Hosp/ambul no-Excel: proc <870000 sin numAutorizacion
        'internacion_sin_aut_no_excel': [],  # Hosp no-Excel: otrosServicio INTERNAC sin numAutorizacion
        'internacion_aut_es_cedula': [],  # otrosServicio INTERNAC: numAutorizacion == cédula paciente
    }

    def parse_fecha(valor):
        """Intenta parsear fecha/datetime en varios formatos. Retorna datetime o None."""
        if valor is None:
            return None
        if isinstance(valor, __import__('datetime').datetime):
            return valor
        s = str(valor).strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(s[:len(fmt)], fmt)
            except ValueError:
                pass
        if len(s) >= 16:
            try:
                return datetime.strptime(s[:16], "%Y-%m-%d %H:%M")
            except ValueError:
                pass
        if len(s) >= 10:
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d")
            except ValueError:
                pass
        return None

    for num_doc, p in pacientes_rips.items():

        # ── 1. Identificación: el paciente debe existir en el Excel ──────
        if num_doc not in registros_excel:
            # Paciente NO está en la base EPS → aplicar reglas de autorización faltante
            # Solo si NO es un paciente de urgencias exclusivamente
            _tiene_solo_urg = p.get('tiene_solo_urgencias', False)
            if not _tiene_solo_urg:
                _archivo_rips = p.get('archivo_rips', '')
                _num_factura  = p.get('num_factura', '')
                # Regla: procedimientos con cod < 870000 deben tener numAutorizacion
                for _proc in p.get('procedimientos_pac', []):
                    try:
                        if int(_proc['cod']) < 870000 and not _proc['num_aut']:
                            alertas['proc_sin_aut_no_excel'].append({
                                'num_doc':      num_doc,
                                'cod_proc':     _proc['cod'],
                                'num_factura':  _num_factura,
                                'archivo_rips': _archivo_rips,
                                'mensaje': (
                                    f"Procedimiento {_proc['cod']} (<870000) sin número de "
                                    f"autorización — paciente no localizado en base EPS"
                                ),
                            })
                    except (ValueError, TypeError):
                        pass
                # Regla: otrosServicios de INTERNACION sin numAutorizacion o con cédula en auth
                for _os in p.get('os_internacion', []):
                    _na = _os['num_aut']
                    if not _na:
                        alertas['internacion_sin_aut_no_excel'].append({
                            'num_doc':      num_doc,
                            'cod':          _os['cod'],
                            'nom':          _os['nom'],
                            'num_factura':  _num_factura,
                            'archivo_rips': _archivo_rips,
                            'mensaje': (
                                f"Internación '{_os['nom'][:50]}' sin número de autorización "
                                f"— paciente no localizado en base EPS"
                            ),
                        })
                    elif _na == num_doc:
                        alertas['internacion_aut_es_cedula'].append({
                            'num_doc':      num_doc,
                            'cod':          _os['cod'],
                            'nom':          _os['nom'],
                            'num_aut':      _na,
                            'num_factura':  _num_factura,
                            'archivo_rips': _archivo_rips,
                            'mensaje': (
                                f"Internación '{_os['nom'][:50]}': numAutorizacion contiene "
                                f"la cédula del paciente ({_na}) — no es un número de autorización válido"
                            ),
                        })
            continue

        regs_pac          = registros_excel[num_doc]
        archivo_rips      = p.get('archivo_rips', '')
        num_factura       = p.get('num_factura', '')
        tipo_rips         = p.get('tipo_doc', '')
        tipo_atencion     = p.get('tipo_atencion', 'ambulatorio')
        tiene_urg         = p.get('tiene_urg', False)
        tiene_hosp        = p.get('tiene_hosp', False)
        procedimientos_pac = p.get('procedimientos_pac', [])
        set_auths         = p.get('set_auths', set())

        fecha_inicio_hosp = p.get('fecha_inicio_hosp')
        fecha_egreso_hosp = p.get('fecha_egreso_hosp')
        fecha_inicio_urg  = p.get('fecha_inicio_urg')
        fecha_egreso_urg  = p.get('fecha_egreso_urg')
        auths_hosp        = p.get('auths_hosp', set())
        n_hosp_regs       = p.get('n_hosp_regs', 0)
        hosp_regs_lista   = p.get('hosp_regs_lista', [])

        archivo_eps_def = regs_pac[0].get('archivo', '') if regs_pac else ''

        # ── 2. Tipo de documento ─────────────────────────────────────────
        tipo_eps = regs_pac[0]['tipo_doc'] if regs_pac else ''
        if tipo_eps and tipo_rips and tipo_eps != tipo_rips:
            ya = any(a['num_doc'] == num_doc for a in alertas['tipo_doc_mismatch'])
            if not ya:
                alertas['tipo_doc_mismatch'].append({
                    'num_doc':      num_doc,
                    'tipo_rips':    tipo_rips,
                    'tipo_eps':     tipo_eps,
                    'num_factura':  num_factura,
                    'archivo_rips': archivo_rips,
                    'archivo_eps':  archivo_eps_def,
                    'mensaje':      "El tipo de identificación no coincide con la base",
                })

        # ── 3. Construir índice de auths del Excel para este paciente ────
        # auths_excel_pac: {num_aut → {codigos: set, archivo, fecha_emision}}
        auths_excel_pac = {}
        for reg in regs_pac:
            na   = reg['numero_aut']
            cod  = reg.get('codigo', '')
            arc  = reg.get('archivo', '')
            fem  = reg.get('fecha_emision')
            if not na:
                continue
            if na not in auths_excel_pac:
                auths_excel_pac[na] = {
                    'codigos':       set(),
                    'archivo':       arc,
                    'fecha_emision': None,
                }
            if cod:
                auths_excel_pac[na]['codigos'].add(cod)
            if fem and not auths_excel_pac[na]['fecha_emision']:
                auths_excel_pac[na]['fecha_emision'] = fem

        # ════════════════════════════════════════════════════════════════
        # AMBULATORIO o CIRUGÍA AMBULATORIA
        # Verifica que cada par (numAutorizacion + codProcedimiento) del RIPS
        # exista en la base de datos del paciente.
        # ════════════════════════════════════════════════════════════════
        if tipo_atencion in ('ambulatorio', 'cirugia_ambulatoria'):
            vistos_par = set()
            for proc in procedimientos_pac:
                cod_proc = proc.get('cod', '')
                na_proc  = proc.get('num_aut', '')
                if not cod_proc or not na_proc:
                    continue
                # Buscar par (auth + cod) en la base del paciente
                par_valido = False
                if na_proc in auths_excel_pac:
                    codigos_base = auths_excel_pac[na_proc]['codigos']
                    # Sin códigos en base → dato incompleto, se acepta
                    if not codigos_base or cod_proc in codigos_base:
                        par_valido = True
                if not par_valido:
                    clave = (num_doc, cod_proc, na_proc)
                    if clave not in vistos_par:
                        vistos_par.add(clave)
                        alertas['amb_par_no_cruza'].append({
                            'cod_proc':     cod_proc,
                            'num_aut':      na_proc,
                            'num_doc':      num_doc,
                            'num_factura':  num_factura,
                            'archivo_rips': archivo_rips,
                            'archivo_eps':  (auths_excel_pac[na_proc]['archivo']
                                             if na_proc in auths_excel_pac
                                             else archivo_eps_def),
                            'mensaje': (
                                "La autorizacion o el codigo del procedimiento "
                                "no coinciden con la base de Nueva EPS segun RIPS"
                            ),
                        })

        # ════════════════════════════════════════════════════════════════
        # HOSPITALARIO
        # ════════════════════════════════════════════════════════════════
        elif tipo_atencion == 'hospitalario':

            # 5a. Determinar fechas de ingreso y egreso del episodio
            if tiene_urg and not tiene_hosp:
                # Solo urgencias: fechas desde urgencias
                fecha_ingreso_str = fecha_inicio_urg
                fecha_egreso_str  = fecha_egreso_urg
            elif tiene_urg and tiene_hosp:
                # Urgencias + hospitalización: ingreso desde urgencias (primer punto de atención),
                # egreso desde hospitalización (alta definitiva)
                fecha_ingreso_str = fecha_inicio_urg or fecha_inicio_hosp
                fecha_egreso_str  = fecha_egreso_hosp
            else:
                # Solo hospitalización sin urgencias previas
                fecha_ingreso_str = fecha_inicio_hosp
                fecha_egreso_str  = fecha_egreso_hosp

            fecha_ingreso_dt = parse_fecha(fecha_ingreso_str)
            fecha_egreso_dt  = parse_fecha(fecha_egreso_str)

            # 5b. Validación por procedimiento:
            # Para cada numAutorizacion del RIPS verificar que:
            #   - el auth exista en la base del paciente
            #   - la fechaEmision de la base esté en [fechaIngreso, fechaEgreso]
            # (Sin filtro por código: en hospitalario la auth cubre el episodio completo)
            if fecha_ingreso_dt and fecha_egreso_dt:
                vistos_hosp = set()
                for proc in procedimientos_pac:
                    cod_proc = proc.get('cod', '')
                    na_proc  = proc.get('num_aut', '')
                    if not cod_proc or not na_proc:
                        continue
                    par_en_rango = False
                    if na_proc in auths_excel_pac:
                        info   = auths_excel_pac[na_proc]
                        # Hospitalario: solo verifica que auth exista y fecha_emision esté en rango.
                        # No se filtra por código: la autorización cubre el episodio completo.
                        fem_dt = parse_fecha(info.get('fecha_emision'))
                        if fem_dt is None:
                            # Sin fecha en base → no se puede verificar → acepta
                            par_en_rango = True
                        elif fecha_ingreso_dt <= fem_dt <= fecha_egreso_dt:
                            par_en_rango = True
                    if not par_en_rango:
                        clave_h = (num_doc, cod_proc, na_proc)
                        if clave_h not in vistos_hosp:
                            vistos_hosp.add(clave_h)
                            alertas['hosp_proc_no_cruza'].append({
                                'cod_proc':     cod_proc,
                                'num_aut':      na_proc,
                                'num_doc':      num_doc,
                                'num_factura':  num_factura,
                                'archivo_rips': archivo_rips,
                                'archivo_eps':  (auths_excel_pac[na_proc]['archivo']
                                                 if na_proc in auths_excel_pac
                                                 else archivo_eps_def),
                                'seccion':      'procedimientos',
                                'fecha_ingreso': str(fecha_ingreso_str or ''),
                                'fecha_egreso':  str(fecha_egreso_str or ''),
                                'mensaje': (
                                    "La autorizacion y/o el procedimiento no coinciden "
                                    "con la base dentro del rango de fechas "
                                    "de la atencion hospitalaria"
                                ),
                            })

            # 5c. Auths en base dentro del rango que NO están en el RIPS
            if fecha_ingreso_dt and fecha_egreso_dt:
                vistos_no_rel = set()
                for reg in regs_pac:
                    na_base = reg.get('numero_aut', '')
                    if not na_base:
                        continue
                    fem_dt = parse_fecha(reg.get('fecha_emision'))
                    if fem_dt is None:
                        continue
                    if fecha_ingreso_dt <= fem_dt <= fecha_egreso_dt:
                        if na_base not in set_auths:
                            if na_base not in vistos_no_rel:
                                vistos_no_rel.add(na_base)
                                alertas['hosp_aut_no_relacionada'].append({
                                    'num_aut':      na_base,
                                    'num_doc':      num_doc,
                                    'num_factura':  num_factura,
                                    'archivo_rips': archivo_rips,
                                    'archivo_eps':  reg.get('archivo', archivo_eps_def),
                                    'fecha_emision': str(fem_dt)[:16],
                                    'fecha_ingreso': str(fecha_ingreso_str or ''),
                                    'fecha_egreso':  str(fecha_egreso_str or ''),
                                    'mensaje': (
                                        "Existen autorizaciones no "
                                        "relacionadas en el RIPS Json."
                                    ),
                                })

        # ════════════════════════════════════════════════════════════════
        # NUEVAS VALIDACIONES AMBULATORIAS
        # ════════════════════════════════════════════════════════════════
        if tipo_atencion in ('ambulatorio', 'cirugia_ambulatoria'):
            # Construir índice inverso: cod → set(auths) desde la base EPS
            cod_a_auths_base: dict = {}
            for na_b, info_b in auths_excel_pac.items():
                for cod_b in info_b.get('codigos', set()):
                    if cod_b not in cod_a_auths_base:
                        cod_a_auths_base[cod_b] = set()
                    cod_a_auths_base[cod_b].add(na_b)

            vistos_proc = set()
            for proc in procedimientos_pac:
                cod_p  = proc.get('cod', '')
                na_p   = proc.get('num_aut', '')
                fi_p   = proc.get('fecha_inicio', '')
                if not cod_p:
                    continue
                clave_p = (num_doc, cod_p, na_p)

                # ── proc_sin_aut_amb: procedimiento sin autorización ────
                if not na_p and clave_p not in vistos_proc:
                    # Determinar si el código es estándar (solo dígitos y ≤6 chars)
                    es_estandar = cod_p.isdigit() and len(cod_p) <= 6
                    if es_estandar:
                        alertas['proc_sin_aut_amb'].append({
                            'cod_proc':     cod_p,
                            'num_doc':      num_doc,
                            'num_factura':  num_factura,
                            'fecha_inicio': fi_p,
                            'seccion':      'procedimientos',
                            'archivo_rips': archivo_rips,
                        })
                    else:
                        # ── cups_noestandar_sin_aut: CUPS no estándar sin auth ─
                        alertas['cups_noestandar_sin_aut'].append({
                            'cod_proc':     cod_p,
                            'num_doc':      num_doc,
                            'num_factura':  num_factura,
                            'fecha_inicio': fi_p,
                            'seccion':      'procedimientos',
                            'archivo_rips': archivo_rips,
                        })
                    vistos_proc.add(clave_p)
                    continue

                if not na_p:
                    continue

                # ── amb_aut_emision_posterior: auth emitida después del servicio ──
                if na_p in auths_excel_pac and fi_p:
                    fem_info = auths_excel_pac[na_p].get('fecha_emision')
                    fem_dt   = parse_fecha(fem_info)
                    fi_dt    = parse_fecha(fi_p)
                    if fem_dt and fi_dt and fem_dt.date() > fi_dt.date():
                        clave_e = (num_doc, na_p, cod_p)
                        if clave_e not in vistos_proc:
                            vistos_proc.add(clave_e)
                            alertas['amb_aut_emision_posterior'].append({
                                'num_aut':      na_p,
                                'num_doc':      num_doc,
                                'num_factura':  num_factura,
                                'fecha_servicio': fi_p[:10],
                                'fecha_emision':  str(fem_dt)[:10],
                                'archivo_rips':   archivo_rips,
                                'archivo_eps':    auths_excel_pac[na_p].get('archivo', archivo_eps_def),
                            })

                # ── proc_aut_no_cruza_amb: auth no corresponde al CUPS según base ──
                if na_p and cod_p in cod_a_auths_base:
                    auths_correctas = cod_a_auths_base[cod_p]
                    if na_p not in auths_correctas:
                        clave_nc = (num_doc, cod_p, na_p)
                        if clave_nc not in vistos_proc:
                            vistos_proc.add(clave_nc)
                            alertas['proc_aut_no_cruza_amb'].append({
                                'cod_proc':     cod_p,
                                'num_aut':      na_p,
                                'num_doc':      num_doc,
                                'num_factura':  num_factura,
                                'fecha_inicio': fi_p,
                                'archivo_rips': archivo_rips,
                                'archivo_eps':  archivo_eps_def,
                            })

        # ════════════════════════════════════════════════════════════════
        # NUEVAS VALIDACIONES HOSPITALARIAS
        # ════════════════════════════════════════════════════════════════
        if n_hosp_regs > 0:
            # ── hosp_cod_sin_aut: registros hospitalización sin numAutorizacion ──
            for hreg in hosp_regs_lista:
                if not hreg.get('na'):
                    alertas['hosp_cod_sin_aut'].append({
                        'num_doc':      num_doc,
                        'num_factura':  num_factura,
                        'archivo_rips': archivo_rips,
                        'archivo_eps':  archivo_eps_def,
                        'mensaje':      "Registro de hospitalización sin numAutorizacion",
                    })

            # ── hosp_proc_cod_no_cruza: procs < 870000 sin coincidencia auth+cod en base ──
            vistos_hpc = set()
            for proc in procedimientos_pac:
                cod_p = proc.get('cod', '')
                na_p  = proc.get('num_aut', '')
                if not cod_p or not na_p:
                    continue
                try:
                    if int(cod_p) >= 870000:
                        continue
                except (ValueError, TypeError):
                    continue
                par_ok = False
                if na_p in auths_excel_pac:
                    codigos_base = auths_excel_pac[na_p].get('codigos', set())
                    if not codigos_base or cod_p in codigos_base:
                        par_ok = True
                if not par_ok:
                    clave_hpc = (num_doc, cod_p, na_p)
                    if clave_hpc not in vistos_hpc:
                        vistos_hpc.add(clave_hpc)
                        alertas['hosp_proc_cod_no_cruza'].append({
                            'cod_proc':     cod_p,
                            'num_aut':      na_p,
                            'num_doc':      num_doc,
                            'num_factura':  num_factura,
                            'archivo_rips': archivo_rips,
                            'archivo_eps':  (auths_excel_pac[na_p]['archivo']
                                             if na_p in auths_excel_pac else archivo_eps_def),
                            'mensaje': (
                                "El codProcedimiento (<870000) con su numAutorizacion "
                                "no coincide con la base de datos EPS"
                            ),
                        })

            # ── hosp_cups_duplicado: código CUPS repetido ────────────────────
            conteo_cups: dict = {}
            for proc in procedimientos_pac:
                cod_p = proc.get('cod', '')
                if not cod_p:
                    continue
                try:
                    if int(cod_p) >= 870000:
                        continue
                except (ValueError, TypeError):
                    continue
                conteo_cups[cod_p] = conteo_cups.get(cod_p, 0) + 1
            for cod_dup, cnt in conteo_cups.items():
                if cnt > 1:
                    alertas['hosp_cups_duplicado'].append({
                        'cod_proc':     cod_dup,
                        'repeticiones': cnt,
                        'num_doc':      num_doc,
                        'num_factura':  num_factura,
                        'archivo_rips': archivo_rips,
                        'mensaje':      f"Código {cod_dup} reportado {cnt} veces en hospitalización",
                    })

        # ════════════════════════════════════════════════════════════════
        # REGLAS DE INTERNACIÓN (aplica a todo paciente con hospitalizacion)
        # ════════════════════════════════════════════════════════════════
        if n_hosp_regs > 0:
            _internacion_reglas(
                num_doc, num_factura, archivo_rips, archivo_eps_def,
                auths_hosp, n_hosp_regs, auths_excel_pac,
                procedimientos_pac, alertas,
            )

    return alertas


def _internacion_reglas(
    num_doc, num_factura, archivo_rips, archivo_eps_def,
    auths_hosp, n_hosp_regs, auths_excel_pac,
    procedimientos_pac, alertas,
):
    """
    Reglas de internacion (hospitalizacion):
    1. Si n_hosp == 1 : la auth de estancia debe existir en el Excel.
       Si n_hosp > 1  : al menos una auth de estancia debe existir en el Excel.
       Si no          : alerta 'Estancia no cuenta con autorizaciones relacionadas en base de datos'
    2. Procedimientos con codProcedimiento < 870000 deben tener numAutorizacion
       DIFERENTE al numAutorizacion de la estancia.
       Si coincide    : alerta 'Procedimiento quirurgico detectado igual al numero de autorizacion de estancia'
    3. Si no hay numAutorizacion ni en estancia ni en procedimientos < 870000:
       alerta 'No se ha relacionado numeros de autorizacion'
    """
    # ── Regla 1: auth de estancia en Excel ──────────────────────────────
    auths_hosp_en_excel = {a for a in auths_hosp if a in auths_excel_pac}

    if n_hosp_regs == 1:
        if not auths_hosp or not auths_hosp_en_excel:
            alertas['estancia_sin_aut'].append({
                'num_doc':      num_doc,
                'num_factura':  num_factura,
                'archivo_rips': archivo_rips,
                'archivo_eps':  archivo_eps_def,
                'auths_hosp':   ', '.join(sorted(auths_hosp)) if auths_hosp else '(ninguna)',
                'mensaje':      "Estancia no cuenta con autorizaciones relacionadas en base de datos",
            })
    elif n_hosp_regs > 1:
        if not auths_hosp_en_excel:
            alertas['estancia_sin_aut'].append({
                'num_doc':      num_doc,
                'num_factura':  num_factura,
                'archivo_rips': archivo_rips,
                'archivo_eps':  archivo_eps_def,
                'auths_hosp':   ', '.join(sorted(auths_hosp)) if auths_hosp else '(ninguna)',
                'mensaje':      "Estancia no cuenta con autorizaciones relacionadas en base de datos",
            })

    # ── Regla 2: proc quirúrgico no debe compartir auth con la estancia ─
    vistos_qx = set()
    for proc in procedimientos_pac:
        cod = proc.get('cod', '')
        na  = proc.get('num_aut', '')
        if not na or not cod:
            continue
        try:
            if int(cod) < 870000 and na in auths_hosp:
                clave = (num_doc, cod, na)
                if clave not in vistos_qx:
                    vistos_qx.add(clave)
                    alertas['proc_qx_misma_aut_hosp'].append({
                        'cod_proc':     cod,
                        'num_aut':      na,
                        'num_doc':      num_doc,
                        'num_factura':  num_factura,
                        'archivo_rips': archivo_rips,
                        'archivo_eps':  (auths_excel_pac[na]['archivo']
                                         if na in auths_excel_pac else archivo_eps_def),
                        'mensaje': (
                            "Procedimiento quirurgico detectado igual al numero "
                            "de autorizacion de estancia"
                        ),
                    })
        except (ValueError, TypeError):
            pass

    # ── Regla 3: sin ninguna auth (ni estancia ni proc < 870000) ────────
    procs_lt870_con_aut = False
    for pr in procedimientos_pac:
        try:
            if pr.get('num_aut') and int(pr.get('cod', '0')) < 870000:
                procs_lt870_con_aut = True
                break
        except (ValueError, TypeError):
            pass

    if not auths_hosp and not procs_lt870_con_aut:
        alertas['sin_num_aut_relacionado'].append({
            'num_doc':      num_doc,
            'num_factura':  num_factura,
            'archivo_rips': archivo_rips,
            'archivo_eps':  archivo_eps_def,
            'mensaje':      "No se ha relacionado numeros de autorizacion",
        })



def validar_medicamentos_malla_2275(data, nombre_archivo=""):
    """
    Aplica las reglas del Bloque M (Medicamentos) de la Malla de Validación RIPS
    según Resolución 2275 de 2023 del Ministerio de Salud y Protección Social.

    Reglas implementadas (sin catálogos externos):
      M04  – fechaDispensAdmon: formato AAAA-MM-DD HH:MM, no futura
      M07  – tipoMedicamento: 2 chars, en TipoMedicamentoPOSVersion2
      M08  – codTecnologiaSalud: obligatorio
      M09  – nomTecnologiaSalud: obligatorio si magistral (M07=03)
      M10  – concentracionMedicamento: obligatorio si magistral
      M11  – unidadMedida: obligatorio si magistral
      M12  – formaFarmaceutica: obligatorio si magistral
      M13  – unidadMinDispensa: obligatorio
      M14  – cantidadMedicamento > 0
      M15  – diasTratamiento: obligatorio
      M16  – tipoDocumentoldentificacion prescriptor: en conjunto válido
      M17  – numDocumentoldentificacion prescriptor: longitud 4-20
      M18  – vrUnitMedicamento >= 0
      M19  – vrServicio >= 0
      M20/RVC092 – conceptoRecaudo: en {01,02,03,05}; "Anticipos" excluido
      M21/RVC060 – valorPagoModerador: >= 1 si CR∈{01,03}; = 0 si CR=05
      M22  – numFEVPagoModerador: null si conceptoRecaudo=05
      M23  – consecutivo: único y secuencial desde 1
      RVG13 – sin duplicados de codTecnologiaSalud por usuario (notificación)
      ARITH-01 – vrUnitMedicamento × cantidadMedicamento ≈ vrServicio

    Retorna lista de dicts con los errores/notificaciones encontrados.
    """
    errores = []
    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}

    if not isinstance(data, dict):
        return errores

    # Número de factura del archivo
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy = datetime.now()

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue

        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue

        meds = servicios.get("medicamentos", [])
        if not isinstance(meds, list) or not meds:
            continue

        ctx_base = {
            "archivo":     nombre_archivo,
            "num_factura": num_factura,
            "num_doc":     num_doc,
        }

        consecutivos_vistos = set()

        consec_esperado     = 1

        for med in meds:
            if not isinstance(med, dict):
                continue

            consec    = med.get("consecutivo")
            tipo_med  = normalizar_str(med.get("tipoMedicamento") or "")
            cod       = normalizar_str(
                med.get("codTecnologiaSalud")
                or med.get("codMedicamento")
                or ""
            )
            nom       = normalizar_str(med.get("nomTecnologiaSalud") or "")
            es_mag    = (tipo_med == "03")
#TOdo: revisar si es necesario normalizar otros campos (ej. codTecnologiaSalud) para evitar falsos positivos por espacios o mayúsculas. Por ejemplo, si el código de tecnología de salud tiene espacios adicionales o diferencias de mayúsculas, podría generar errores que no reflejan un problema real en los datos. Normalizar estos campos antes de las validaciones ayudaría a reducir este tipo de falsos positivos y hacer las reglas más robustas frente a inconsistencias menores en el formato de los datos.
            consec_s  = normalizar_str(consec)

            def _err(id_regla, severidad, campo, mensaje, valor_actual=""):
                errores.append({
                    **ctx_base,
                    "id_regla":     id_regla,
                    "severidad":    severidad,
                    "campo":        campo,
                    "consecutivo":  consec_s,
                    "mensaje":      mensaje,
                    "valor_actual": normalizar_str(valor_actual),
                })

            # ── M23: Consecutivo único y secuencial ──────────────────────
            if isinstance(consec, int):
                if consec in consecutivos_vistos:
                    _err("M23", "critica", "consecutivo",
                         f"Consecutivo {consec} repetido en medicamentos del mismo usuario.",
                         consec)
                else:
                    consecutivos_vistos.add(consec)
                if consec != consec_esperado:
                    _err("M23", "critica", "consecutivo",
                         f"Consecutivo {consec} no es secuencial (esperado: {consec_esperado}).",
                         consec)
                consec_esperado = consec + 1
            else:
                consec_esperado += 1

            # ── M04: fechaDispensAdmon ────────────────────────────────────
            fecha_raw = normalizar_str(med.get("fechaDispensAdmon") or "")
            if not fecha_raw:
                _err("M04", "critica", "fechaDispensAdmon",
                     "fechaDispensAdmon es obligatorio (RVC013).", "")
            elif len(fecha_raw) != 16:
                _err("M04", "critica", "fechaDispensAdmon",
                     f"fechaDispensAdmon debe tener formato AAAA-MM-DD HH:MM (16 caracteres). "
                     f"Longitud actual: {len(fecha_raw)}.", fecha_raw)
            else:
                try:
                    fecha_dt = datetime.strptime(fecha_raw, "%Y-%m-%d %H:%M")
                    if fecha_dt > fecha_hoy:
                        _err("M04", "critica", "fechaDispensAdmon",
                             "La fecha de dispensación/administración es mayor a la fecha actual (RVC013).",
                             fecha_raw)
                except ValueError:
                    _err("M04", "critica", "fechaDispensAdmon",
                         "fechaDispensAdmon no es una fecha/hora válida (AAAA-MM-DD HH:MM).",
                         fecha_raw)

            # ── M07: tipoMedicamento ──────────────────────────────────────
            if not tipo_med:
                _err("M07-DOMINIO", "critica", "tipoMedicamento",
                     "tipoMedicamento es obligatorio (catálogo TipoMedicamentoPOSVersion2).", "")
            elif len(tipo_med) != 2:
                _err("M07-DOMINIO", "critica", "tipoMedicamento",
                     f"tipoMedicamento debe tener exactamente 2 caracteres.", tipo_med)
            elif tipo_med not in TIPOS_MEDICAMENTO_VALIDOS:
                _err("M07-DOMINIO", "critica", "tipoMedicamento",
                     f"tipoMedicamento '{tipo_med}' no pertenece al catálogo "
                     f"TipoMedicamentoPOSVersion2 ({', '.join(sorted(TIPOS_MEDICAMENTO_VALIDOS))}).",
                     tipo_med)

            # ── M08: codTecnologiaSalud ───────────────────────────────────
            if not cod:
                _err("M08", "critica", "codTecnologiaSalud",
                     "codTecnologiaSalud es obligatorio.", "")

            # ── M09: nomTecnologiaSalud obligatorio para magistral ────────
            if es_mag and (not nom or nom.lower() == "none"):
                _err("M09-MAGISTRAL", "alta", "nomTecnologiaSalud",
                     "Para preparación magistral (tipoMedicamento=03), "
                     "nomTecnologiaSalud es obligatorio (RVC065).", nom)

            # ── M10: concentracionMedicamento obligatorio para magistral ──
            if es_mag:
                conc = med.get("concentracionMedicamento")
                if conc is None or normalizar_str(conc) in {"", "0", "0.0"}:
                    _err("M10-MAGISTRAL", "alta", "concentracionMedicamento",
                         "Para preparación magistral, concentracionMedicamento es obligatorio.",
                         normalizar_str(conc))

            # ── M11: unidadMedida obligatorio para magistral ──────────────
            if es_mag:
                um = normalizar_str(med.get("unidadMedida") or "")
                if not um:
                    _err("M11-MAGISTRAL", "alta", "unidadMedida",
                         "Para preparación magistral, unidadMedida es obligatorio.", um)

            # ── M12: formaFarmaceutica obligatorio para magistral ─────────
            if es_mag:
                ff = normalizar_str(med.get("formaFarmaceutica") or "")
                if not ff:
                    _err("M12-MAGISTRAL", "alta", "formaFarmaceutica",
                         "Para preparación magistral, formaFarmaceutica es obligatorio.", ff)

            # ── M13: unidadMinDispensa ────────────────────────────────────
            umd = normalizar_str(med.get("unidadMinDispensa") or "")
            if not umd:
                _err("M13-DOMINIO", "critica", "unidadMinDispensa",
                     "unidadMinDispensa es obligatorio (catálogo UPR).", "")

            # ── M14: cantidadMedicamento > 0 ─────────────────────────────
            cantidad_raw = med.get("cantidadMedicamento")
            cantidad_num = None
            if cantidad_raw is None:
                _err("M14-CANTIDAD", "critica", "cantidadMedicamento",
                     "cantidadMedicamento es obligatorio y debe ser > 0.", "")
            else:
                try:
                    cantidad_num = float(cantidad_raw)
                    if cantidad_num <= 0:
                        _err("M14-CANTIDAD", "critica", "cantidadMedicamento",
                             "cantidadMedicamento debe ser mayor a cero.", cantidad_raw)
                except (ValueError, TypeError):
                    _err("M14-CANTIDAD", "critica", "cantidadMedicamento",
                         "cantidadMedicamento debe ser un valor numérico mayor a cero.",
                         cantidad_raw)

            # ── M15: diasTratamiento ──────────────────────────────────────
            if med.get("diasTratamiento") is None:
                _err("M15", "critica", "diasTratamiento",
                     "diasTratamiento es obligatorio.", "")

            # ── M16: tipoDocumentoIdentificacion prescriptor ─────────────
            tdp = _tipo_doc(med)
            if not tdp:
                _err("M16", "critica", "tipoDocumentoIdentificacion",
                     "El tipo de documento del prescriptor (M16) es obligatorio.", "")
            elif tdp not in TIPOS_DOC_PRESCRIPTOR_MED:
                _err("M16", "critica", "tipoDocumentoIdentificacion",
                     f"Tipo de documento del prescriptor '{tdp}' no es válido. "
                     f"Permitidos: {', '.join(sorted(TIPOS_DOC_PRESCRIPTOR_MED))}.", tdp)

            # ── M17: numDocumentoIdentificacion prescriptor ──────────────
            ndp = _num_doc_val(med)
            if not ndp:
                _err("M17", "critica", "numDocumentoIdentificacion",
                     "El número de documento del prescriptor (M17) es obligatorio.", "")
            elif not (4 <= len(ndp) <= 20):
                _err("M17", "critica", "numDocumentoIdentificacion",
                     f"Número de documento del prescriptor debe tener entre 4 y 20 caracteres "
                     f"(tiene {len(ndp)}).", ndp)

            # ── M18: vrUnitMedicamento >= 0 ───────────────────────────────
            vru_raw = med.get("vrUnitMedicamento")
            vru_num = None
            if vru_raw is None:
                _err("M18", "critica", "vrUnitMedicamento",
                     "vrUnitMedicamento es obligatorio.", "")
            else:
                try:
                    vru_num = float(vru_raw)
                    if vru_num < 0:
                        _err("M18", "critica", "vrUnitMedicamento",
                             "vrUnitMedicamento no puede ser negativo.", vru_raw)
                except (ValueError, TypeError):
                    _err("M18", "critica", "vrUnitMedicamento",
                         "vrUnitMedicamento debe ser un valor numérico.", vru_raw)

            # ── M19: vrServicio >= 0 ──────────────────────────────────────
            vrs_raw = med.get("vrServicio")
            vrs_num = None
            if vrs_raw is None:
                _err("M19", "critica", "vrServicio",
                     "vrServicio es obligatorio en medicamentos.", "")
            else:
                try:
                    vrs_num = float(vrs_raw)
                    if vrs_num < 0:
                        _err("M19", "critica", "vrServicio",
                             "vrServicio no puede ser negativo.", vrs_raw)
                except (ValueError, TypeError):
                    _err("M19", "critica", "vrServicio",
                         "vrServicio debe ser un valor numérico.", vrs_raw)

            # ── M20/RVC092: conceptoRecaudo ───────────────────────────────
            cr = normalizar_str(med.get("conceptoRecaudo") or "")
            if not cr:
                _err("RVC092/M20", "critica", "conceptoRecaudo",
                     "conceptoRecaudo es obligatorio en medicamentos.", "")
            elif len(cr) != 2:
                _err("RVC092/M20", "critica", "conceptoRecaudo",
                     "conceptoRecaudo debe tener exactamente 2 caracteres.", cr)
            elif cr not in CONCEPTOS_RECAUDO_MED:
                _err("RVC092/M20", "critica", "conceptoRecaudo",
                     f"conceptoRecaudo '{cr}' no es válido. El concepto 'Anticipos' no aplica "
                     f"en RIPS (RVC092). Valores válidos: {', '.join(sorted(CONCEPTOS_RECAUDO_MED))}.", cr)

            # ── M21/RVC060-61: valorPagoModerador ────────────────────────
            vpm_raw = med.get("valorPagoModerador")
            if vpm_raw is None:
                _err("M21/RVC060", "critica", "valorPagoModerador",
                     "valorPagoModerador es obligatorio.", "")
            else:
                try:
                    vpm = float(vpm_raw)
                    if vpm < 0:
                        _err("M21/RVC060", "critica", "valorPagoModerador",
                             "valorPagoModerador no puede ser negativo.", vpm_raw)
                    if cr in {"01", "03"} and vpm < 1:
                        _err("M21/RVC060", "critica", "valorPagoModerador",
                             f"Con conceptoRecaudo='{cr}', valorPagoModerador debe ser >= 1 (RVC060).",
                             vpm_raw)
                    if cr == "05" and vpm != 0:
                        _err("M21/RVC061", "critica", "valorPagoModerador",
                             "Con conceptoRecaudo='05' (No aplica), valorPagoModerador debe ser 0 (RVC061).",
                             vpm_raw)
                except (ValueError, TypeError):
                    _err("M21/RVC060", "critica", "valorPagoModerador",
                         "valorPagoModerador debe ser un valor numérico.", vpm_raw)

            # ── M22: numFEVPagoModerador null si CR=05 ────────────────────
            if cr == "05":
                nfev = med.get("numFEVPagoModerador")
                if nfev is not None and normalizar_str(nfev) not in {"", "none", "null"}:
                    _err("M22", "alta", "numFEVPagoModerador",
                         "numFEVPagoModerador debe ser null cuando conceptoRecaudo='05' (No aplica).",
                         nfev)

            # ── ARITH-01: vrUnit × cantidad ≈ vrServicio ──────────────────
            if (vru_num is not None and vrs_num is not None and
                    cantidad_num is not None and vru_num > 0 and cantidad_num > 0):
                calculado  = vru_num * cantidad_num
                tolerancia = max(1.0, calculado * 0.01)
                if abs(calculado - vrs_num) > tolerancia:
                    _err("ARITH-01", "media", "vrServicio",
                         f"vrUnitMedicamento ({vru_num:,.0f}) × cantidadMedicamento ({cantidad_num:,.0f}) "
                         f"= {calculado:,.0f} no coincide con vrServicio ({vrs_num:,.0f}). "
                         f"Diferencia: {abs(calculado - vrs_num):,.0f}.",
                         normalizar_str(vrs_raw))


    return errores


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUES GENERALES (0, T)
# ══════════════════════════════════════════════════════════════

def validar_general_malla_2275(data, nombre_archivo=""):
    """
    Bloque 0 y T: RVG01, RVG03, RVG07, RVG12, T01, T02, T03-DOMINIO, T04-CONDICIONAL.
    """
    errores = []
    if not isinstance(data, dict):
        errores.append({
            "archivo": nombre_archivo, "num_factura": "", "num_doc": "",
            "consecutivo": "", "id_regla": "RVG01", "severidad": "critica",
            "campo": "raíz JSON",
            "mensaje": "El archivo no es un objeto JSON válido (raíz no es un objeto/dict).",
            "valor_actual": str(type(data)),
        })
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    def _e(id_regla, severidad, campo, mensaje, valor_actual="", num_doc="", consecutivo=""):
        errores.append({
            "archivo":      nombre_archivo,
            "num_factura":  num_factura,
            "num_doc":      num_doc,
            "consecutivo":  consecutivo,
            "id_regla":     id_regla,
            "severidad":    severidad,
            "campo":        campo,
            "mensaje":      mensaje,
            "valor_actual": normalizar_str(valor_actual),
        })

    # ── RVG01: Campos raíz obligatorios ──────────────────────
    # El campo NIT acepta dos variantes de nombre en uso
    tiene_nit = ("numDocumentoldObligado" in data or "numDocumentoIdObligado" in data)
    if not tiene_nit:
        _e("RVG01", "critica", "numDocumentoldObligado / numDocumentoIdObligado",
           "Campo obligatorio NIT del facturador ausente en la raíz del JSON (RVG01).")
    for campo_raiz in ("numFactura", "tipoNota", "numNota", "usuarios"):
        if campo_raiz not in data:
            _e("RVG01", "critica", campo_raiz,
               f"Campo obligatorio '{campo_raiz}' ausente en la raíz del JSON (RVG01).")

    # ── T01: NIT presente ─────────────────────────────────────
    nit = _nit_obligado(data)
    if not nit:
        _e("T01", "critica", "numDocumentoIdObligado",
           "NIT del facturador (numDocumentoIdObligado) es obligatorio y no puede ser vacío.")

    # ── T02: numFactura ───────────────────────────────────────
    factura_val = data.get("numFactura")
    if factura_val is not None and normalizar_str(factura_val) == "":
        _e("T02", "alta", "numFactura",
           "numFactura está presente pero vacío. Debe ser null si el RIPS no tiene FEV.", factura_val)

    # ── T03 / T04: tipoNota → numNota condicional ─────────────
    tipo_nota = data.get("tipoNota")
    num_nota  = data.get("numNota")
    if tipo_nota is not None:
        tipo_nota_s = normalizar_str(tipo_nota)
        if len(tipo_nota_s) != 2:
            _e("T03-DOMINIO", "critica", "tipoNota",
               "tipoNota debe tener exactamente 2 caracteres o ser null.", tipo_nota_s)
        elif tipo_nota_s not in TIPOS_NOTA_VALIDOS:
            _e("T03-DOMINIO", "critica", "tipoNota",
               f"tipoNota '{tipo_nota_s}' no pertenece al catálogo TipoNota. "
               f"Valores válidos: NA=Nota ajuste RIPS, NC=Nota crédito, ND=Nota débito, RS=RIPS sin Factura.",
               tipo_nota_s)
        if num_nota is None or normalizar_str(num_nota) == "":
            _e("T04-CONDICIONAL", "critica", "numNota",
               "numNota es obligatorio cuando tipoNota está informado.")
        else:
            nn_s = normalizar_str(num_nota)
            if len(nn_s) < 1 or len(nn_s) > 20:
                _e("T04-CONDICIONAL", "critica", "numNota",
                   f"numNota debe tener entre 1 y 20 caracteres (actual: {len(nn_s)}).", nn_s)

    # ── RVG03: Al menos un servicio en el RIPS ────────────────
    usuarios = data.get("usuarios", [])
    if not isinstance(usuarios, list) or len(usuarios) == 0:
        _e("RVG03", "critica", "usuarios",
           "El arreglo 'usuarios' está vacío o ausente. El RIPS debe tener al menos un usuario.")
        return errores

    SECCIONES = ("consultas", "procedimientos", "urgencias", "hospitalizacion",
                 "recienNacidos", "medicamentos", "otrosServicios")
    hay_servicios = any(
        isinstance(u, dict) and isinstance(u.get("servicios"), dict) and
        any(isinstance(u["servicios"].get(s), list) and u["servicios"][s] for s in SECCIONES)
        for u in usuarios
    )
    if not hay_servicios:
        _e("RVG03", "critica", "servicios",
           "No se encontraron servicios prestados en ninguna sección del RIPS (RVG03).")

    # ── RVG07: Cada usuario tiene al menos un servicio ────────
    for u in usuarios:
        if not isinstance(u, dict):
            continue
        num_doc_u = _num_doc_val(u)
        consec_u  = normalizar_str(u.get("consecutivo", ""))
        servicios = u.get("servicios", {})
        tiene_svc = isinstance(servicios, dict) and any(
            isinstance(servicios.get(s), list) and servicios[s] for s in SECCIONES
        )
        if not tiene_svc:
            _e("RVG07", "critica", "servicios",
               "El usuario no tiene ningún servicio registrado en el RIPS (RVG07).",
               "", num_doc_u, consec_u)

    # ── RVG12: Sin usuarios duplicados ───────────────────────
    usuarios_vistos = {}
    for u in usuarios:
        if not isinstance(u, dict):
            continue
        td = _tipo_doc(u)
        nd = _num_doc_val(u)
        clave = (td, nd)
        if clave in usuarios_vistos:
            _e("RVG12", "critica", "numDocumentoldentificacion",
               f"Usuario duplicado: {td}-{nd} aparece más de una vez en el RIPS (RVG12).",
               nd, nd)
        else:
            usuarios_vistos[clave] = True

    return errores


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUE U: USUARIOS
# ══════════════════════════════════════════════════════════════

def validar_usuarios_malla_2275(data, nombre_archivo=""):
    """
    Bloque U: U01-U11, RVC006, RVC007, RVC008, RVC009, SEX-IVE-01.
    """
    errores = []
    if not isinstance(data, dict):
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy       = datetime.now().date()
    consec_esperado = 1
    consecs_vistos  = set()

    for u in data.get("usuarios", []):
        if not isinstance(u, dict):
            consec_esperado += 1
            continue

        num_doc  = _num_doc_val(u)
        consec_u = u.get("consecutivo")
        consec_s = normalizar_str(consec_u)

        def _e(id_regla, severidad, campo, mensaje, valor_actual=""):
            errores.append({
                "archivo":      nombre_archivo,
                "num_factura":  num_factura,
                "num_doc":      num_doc,
                "consecutivo":  consec_s,
                "id_regla":     id_regla,
                "severidad":    severidad,
                "campo":        campo,
                "mensaje":      mensaje,
                "valor_actual": normalizar_str(valor_actual),
            })

        # ── U10: Consecutivo secuencial ───────────────────────
        if isinstance(consec_u, int):
            if consec_u in consecs_vistos:
                _e("U10-CONSECUTIVO", "critica", "consecutivo",
                   f"Consecutivo de usuario {consec_u} está repetido.", consec_u)
            else:
                consecs_vistos.add(consec_u)
            if consec_u != consec_esperado:
                _e("U10-CONSECUTIVO", "critica", "consecutivo",
                   f"Consecutivo {consec_u} no es secuencial (esperado: {consec_esperado}).", consec_u)
            consec_esperado = consec_u + 1
        else:
            consec_esperado += 1

        # ── U01: tipoDocumentoIdentificacion ──────────────────
        tipo_doc = _tipo_doc(u)
        if not tipo_doc:
            _e("U01-DOMINIO", "critica", "tipoDocumentoIdentificacion",
               "tipoDocumentoIdentificacion es obligatorio.")
        elif len(tipo_doc) != 2:
            _e("U01-DOMINIO", "critica", "tipoDocumentoIdentificacion",
               f"tipoDocumentoIdentificacion debe tener 2 caracteres.", tipo_doc)
        elif tipo_doc not in TIPOS_DOC_USUARIO:
            _e("U01-DOMINIO", "critica", "tipoDocumentoIdentificacion",
               f"tipoDocumentoIdentificacion '{tipo_doc}' no pertenece al catálogo TipoldPISIS.", tipo_doc)

        # ── U02: numDocumentoldentificacion longitud 4-20 ─────
        if not num_doc:
            _e("U02-LONGITUD", "critica", "numDocumentoldentificacion",
               "numDocumentoldentificacion es obligatorio.")
        elif len(num_doc) < 4 or len(num_doc) > 20:
            _e("U02-LONGITUD", "critica", "numDocumentoldentificacion",
               f"numDocumentoldentificacion debe tener entre 4 y 20 caracteres (actual: {len(num_doc)}).",
               num_doc)

        # ── U03: tipoUsuario 2 chars, dominio RIPSTipoUsuarioVersion2 ───
        tipo_usr = normalizar_str(u.get("tipoUsuario") or "")
        if not tipo_usr:
            _e("U03-DOMINIO", "critica", "tipoUsuario", "tipoUsuario es obligatorio.")
        elif len(tipo_usr) != 2:
            _e("U03-DOMINIO", "critica", "tipoUsuario",
               f"tipoUsuario debe tener exactamente 2 caracteres.", tipo_usr)
        elif tipo_usr not in TIPOS_USUARIO_RIPS:
            _e("U03-DOMINIO", "critica", "tipoUsuario",
               f"tipoUsuario '{tipo_usr}' no pertenece al catálogo RIPSTipoUsuarioVersion2 "
               f"(01=Contrib. cotizante, 02=Contrib. beneficiario, 03=Contrib. adicional, "
               f"04=Subsidiado, 05=No afiliado, 06-07=Especial/Excepción, 08=PPL, "
               f"09=ARL, 10=SOAT, 11=Planes voluntarios, 12=Particular, 13=Especial Ley 352).",
               tipo_usr)

        # ── U04 / RVC006: fechaNacimiento ─────────────────────
        fecha_nac_raw = normalizar_str(u.get("fechaNacimiento") or "")
        fecha_nac_dt  = None
        if not fecha_nac_raw:
            _e("U04-FORMATO", "critica", "fechaNacimiento",
               "fechaNacimiento es obligatorio (formato AAAA-MM-DD).")
        elif len(fecha_nac_raw) != 10:
            _e("U04-FORMATO", "critica", "fechaNacimiento",
               f"fechaNacimiento debe tener formato AAAA-MM-DD (10 caracteres).", fecha_nac_raw)
        else:
            try:
                fecha_nac_dt = datetime.strptime(fecha_nac_raw, "%Y-%m-%d").date()
                if fecha_nac_dt > fecha_hoy:
                    _e("RVC006", "critica", "fechaNacimiento",
                       "fechaNacimiento no puede ser mayor a la fecha actual (RVC006).", fecha_nac_raw)
            except ValueError:
                _e("U04-FORMATO", "critica", "fechaNacimiento",
                   "fechaNacimiento no es una fecha válida (AAAA-MM-DD).", fecha_nac_raw)

        # ── U05: codSexo ──────────────────────────────────────
        cod_sexo = normalizar_str(u.get("codSexo") or "")
        if not cod_sexo:
            _e("U05-DOMINIO", "critica", "codSexo", "codSexo es obligatorio.")
        elif len(cod_sexo) != 1:
            _e("U05-DOMINIO", "critica", "codSexo",
               f"codSexo debe tener exactamente 1 carácter.", cod_sexo)
        elif cod_sexo not in CODIGOS_SEXO:
            _e("U05-DOMINIO", "critica", "codSexo",
               f"codSexo '{cod_sexo}' no pertenece al catálogo Sexo. "
               "Valores válidos: M=Masculino, F=Femenino, I=Indeterminado.", cod_sexo)

        # ── U06: codPaisResidencia 3 chars ────────────────────
        cod_pais = normalizar_str(u.get("codPaisResidencia") or "")
        if not cod_pais:
            _e("U06-DOMINIO", "critica", "codPaisResidencia",
               "codPaisResidencia es obligatorio (3 caracteres ISO 3166-1).")
        elif len(cod_pais) != 3:
            _e("U06-DOMINIO", "critica", "codPaisResidencia",
               f"codPaisResidencia debe tener 3 caracteres. Actual: '{cod_pais}'.", cod_pais)

        # ── U07: codMunicipioResidencia obligatorio si Colombia ─
        cod_mun   = u.get("codMunicipioResidencia")
        cod_mun_s = normalizar_str(cod_mun or "")
        if cod_pais == "170":
            if cod_mun is None or cod_mun_s == "":
                _e("U07-CONDICIONAL", "critica", "codMunicipioResidencia",
                   "codMunicipioResidencia es obligatorio cuando codPaisResidencia='170' (Colombia).")
            elif len(cod_mun_s) != 5:
                _e("U07-CONDICIONAL", "critica", "codMunicipioResidencia",
                   f"codMunicipioResidencia debe tener 5 caracteres (DANE). Actual: '{cod_mun_s}'.",
                   cod_mun_s)

        # ── U08: codZonaTerritorialResidencia (opcional, dominio 01/02) ─
        cod_zona = u.get("codZonaTerritorialResidencia")
        if cod_zona is not None:
            cod_zona_s = normalizar_str(cod_zona)
            if cod_zona_s and len(cod_zona_s) != 2:
                _e("U08-DOMINIO", "alta", "codZonaTerritorialResidencia",
                   f"codZonaTerritorialResidencia debe tener 2 caracteres si se informa. Actual: '{cod_zona_s}'.",
                   cod_zona_s)
            elif cod_zona_s and cod_zona_s not in ZONAS_VALIDAS:
                _e("U08-DOMINIO", "alta", "codZonaTerritorialResidencia",
                   f"codZonaTerritorialResidencia '{cod_zona_s}' no pertenece al catálogo ZonaVersion2. "
                   "Valores válidos: 01=Rural, 02=Urbano.",
                   cod_zona_s)

        # ── U09: incapacidad SI/NO ────────────────────────────
        incapacidad = normalizar_str(u.get("incapacidad") or "")
        if not incapacidad:
            _e("U09-DOMINIO", "critica", "incapacidad", "incapacidad es obligatorio (SI/NO).")
        elif incapacidad not in VALORES_SINO:
            _e("U09-DOMINIO", "critica", "incapacidad",
               f"incapacidad '{incapacidad}' no pertenece al catálogo LstSiNo (SI/NO).", incapacidad)

        # ── RVC007: tipo de documento coherente con edad ──────
        if fecha_nac_dt and tipo_doc:
            edad = (fecha_hoy - fecha_nac_dt).days // 365
            if edad <= 3:
                permitidos_edad = {"CN", "RC", "PA", "CD", "SC", "PE", "DE", "PT", "MS"}
            elif edad <= 6:
                permitidos_edad = {"RC", "PA", "CD", "SC", "PE", "DE", "PT", "MS"}
            elif edad <= 17:
                permitidos_edad = {"TI", "CE", "PA", "CD", "SC", "PE", "DE", "PT", "MS"}
            elif edad <= 19:
                permitidos_edad = {"CC", "TI", "CE", "PA", "CD", "SC", "PE", "DE", "PT", "AS"}
            else:
                permitidos_edad = {"CC", "CE", "PA", "CD", "SC", "PE", "DE", "PT", "AS"}
            if tipo_doc not in permitidos_edad:
                _e("RVC007", "critica", "tipoDocumentoldentificacion",
                   f"Tipo de documento '{tipo_doc}' no es válido para la edad del usuario ({edad} años). "
                   f"Permitidos: {', '.join(sorted(permitidos_edad))} (RVC007).", tipo_doc)

        # ── RVC008 / RVC009: recién nacidos ───────────────────
        servicios = u.get("servicios", {})
        if isinstance(servicios, dict):
            rn_list = servicios.get("recienNacidos", [])
            if isinstance(rn_list, list) and rn_list:
                if fecha_nac_dt:
                    edad_madre = (fecha_hoy - fecha_nac_dt).days // 365
                    if not (9 <= edad_madre <= 60):
                        _e("RVC008", "media", "fechaNacimiento",
                           f"La edad de la madre es {edad_madre} años. Para recién nacidos se esperan 9-60 años (RVC008).",
                           fecha_nac_raw)
                if cod_sexo and cod_sexo != "F":
                    _e("RVC009", "media", "codSexo",
                       f"El usuario con recién nacidos tiene codSexo='{cod_sexo}'. "
                       "Se esperaba 'F' (Femenino) — la madre del recién nacido debe registrarse como sexo Femenino (RVC009).", cod_sexo)

            # ── SEX-IVE-01: finalidad IVE → sexo debe ser Femenino ────
            if cod_sexo and cod_sexo != "F":
                for seccion in ("consultas","procedimientos","urgencias","hospitalizacion","otrosServicios"):
                    for rec in (servicios.get(seccion) or []):
                        if isinstance(rec, dict):
                            fin = normalizar_str(rec.get("finalidadTecnologiaSalud") or "")
                            if fin in _FINALIDAD_IVE:
                                _e("SEX-IVE-01", "media", "codSexo",
                                   f"Se registra finalidad IVE (código {fin}) pero el usuario tiene codSexo='{cod_sexo}'. "
                                   "Se esperaba sexo 'F' (Femenino) para procedimientos de IVE.", cod_sexo)
                                break
                    else:
                        continue
                    break

            # ── SEX-DX: validar sexo según diagnósticos CIE-10 ───────
            if cod_sexo:
                todos_dx_u: set = set()
                for sec in ("consultas","procedimientos","urgencias","hospitalizacion","otrosServicios"):
                    for rec in (servicios.get(sec) or []):
                        if isinstance(rec, dict):
                            for campo in _CAMPOS_DX:
                                v = normalizar_str(rec.get(campo) or "")
                                if v:
                                    todos_dx_u.add(v)

                # Buscar primer dx femenino (O=obstétrico/aborto, o categoría en _DX_F_CATS)
                dx_f = next(
                    (dx for dx in todos_dx_u
                     if dx[0:1] == "O" or dx[:3] in _DX_F_CATS),
                    None,
                )
                if dx_f and cod_sexo != "F":
                    _e("SEX-DX-F", "media", "codSexo",
                       f"Diagnóstico CIE-10 '{dx_f}' corresponde a patología obstétrica/ginecológica "
                       f"pero el usuario tiene codSexo='{cod_sexo}'. Se esperaba 'F' (Femenino).", cod_sexo)

                # Buscar primer dx relacionado con el pene
                dx_m = next(
                    (dx for dx in todos_dx_u if dx[:3] in _DX_M_PENE_CATS),
                    None,
                )
                if dx_m and cod_sexo != "M":
                    _e("SEX-DX-M", "media", "codSexo",
                       f"Diagnóstico CIE-10 '{dx_m}' corresponde a patología del pene "
                       f"pero el usuario tiene codSexo='{cod_sexo}'. Se esperaba 'M' (Masculino).", cod_sexo)

    return errores


# ══════════════════════════════════════════════════════════════
# HELPER INTERNO: validar fecha de atención (formato + no futura)
# ══════════════════════════════════════════════════════════════

def _validar_fecha_atencion_campo(fecha_raw, campo, fecha_hoy, errores_list, ctx, regla):
    """Valida formato AAAA-MM-DD HH:MM y que no sea futura. Retorna datetime o None."""
    if not fecha_raw:
        errores_list.append({**ctx, "id_regla": regla, "severidad": "critica",
                             "campo": campo,
                             "mensaje": f"{campo} es obligatorio (formato AAAA-MM-DD HH:MM).",
                             "valor_actual": ""})
        return None
    if len(fecha_raw) != 16:
        errores_list.append({**ctx, "id_regla": regla, "severidad": "critica",
                             "campo": campo,
                             "mensaje": (f"{campo} debe tener formato AAAA-MM-DD HH:MM "
                                         f"(16 caracteres). Longitud actual: {len(fecha_raw)}."),
                             "valor_actual": fecha_raw})
        return None
    try:
        dt = datetime.strptime(fecha_raw, "%Y-%m-%d %H:%M")
        if dt > fecha_hoy:
            errores_list.append({**ctx, "id_regla": regla, "severidad": "critica",
                                 "campo": campo,
                                 "mensaje": f"{campo} es mayor a la fecha/hora actual (RVC013).",
                                 "valor_actual": fecha_raw})
        return dt
    except ValueError:
        errores_list.append({**ctx, "id_regla": regla, "severidad": "critica",
                             "campo": campo,
#segunda parte del mensaje corregida para no repetir "formato" dos veces
                             "mensaje": f"{campo} no es una fecha/hora válida (AAAA-MM-DD HH:MM).",
                             "valor_actual": fecha_raw})
        return None


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUE C: CONSULTAS
# ══════════════════════════════════════════════════════════════

def validar_consultas_malla_2275(data, nombre_archivo=""):
    """
    Bloque C (Consultas): C01-C21, RVC011, RVC013, RVC031, RVC060/61, RVC079, RVC086/87.
    """
    errores = []
    if not isinstance(data, dict):
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy = datetime.now()

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        fecha_nac_raw = normalizar_str(usuario.get("fechaNacimiento") or "")
        fecha_nac_dt  = None
        try:
            if len(fecha_nac_raw) == 10:
                fecha_nac_dt = datetime.strptime(fecha_nac_raw, "%Y-%m-%d")
        except ValueError:
            pass

        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        consultas = servicios.get("consultas", [])
        if not isinstance(consultas, list) or not consultas:
            continue

        consec_esperado = 1
        consecs_vistos  = set()

        for con in consultas:
            if not isinstance(con, dict):
                consec_esperado += 1
                continue

            consec   = con.get("consecutivo")
            consec_s = normalizar_str(consec)
            ctx      = {"archivo": nombre_archivo, "num_factura": num_factura,
                        "num_doc": num_doc, "consecutivo": consec_s}

            def _e(id_regla, severidad, campo, mensaje, valor_actual=""):
                errores.append({**ctx, "id_regla": id_regla, "severidad": severidad,
                                "campo": campo, "mensaje": mensaje,
                                "valor_actual": normalizar_str(valor_actual)})

            # ── C21: Consecutivo ──────────────────────────────
            if isinstance(consec, int):
                if consec in consecs_vistos:
                    _e("C21-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} repetido en consultas.", consec)
                else:
                    consecs_vistos.add(consec)
                if consec != consec_esperado:
                    _e("C21-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} no secuencial (esperado: {consec_esperado}).", consec)
                consec_esperado = consec + 1
            else:
                consec_esperado += 1

            # ── C01 / RVC011: codPrestador 12 chars ──────────
            cod_prest = normalizar_str(con.get("codPrestador") or "")
            if not cod_prest:
                _e("RVC011", "critica", "codPrestador",
                   "codPrestador es obligatorio (12 caracteres) (RVC011).")
            elif len(cod_prest) != 12:
                _e("RVC011", "critica", "codPrestador",
                   f"codPrestador debe tener 12 caracteres (RVC011). Actual: {len(cod_prest)}.",
                   cod_prest)

            # ── C02 / RVC013: fechaInicioAtencion ────────────
            fecha_raw    = normalizar_str(con.get("fechaInicioAtencion") or "")
            fecha_ini_dt = _validar_fecha_atencion_campo(
                fecha_raw, "fechaInicioAtencion", fecha_hoy, errores, ctx, "RVC013")
            if fecha_ini_dt and fecha_nac_dt and fecha_ini_dt < fecha_nac_dt:
                _e("RVC079", "critica", "fechaInicioAtencion",
                   "fechaInicioAtencion es menor a la fechaNacimiento del usuario (RVC079).", fecha_raw)

            # ── C04: codConsulta 6 chars ──────────────────────
            cod_con = normalizar_str(con.get("codConsulta") or "")
            if not cod_con:
                _e("C04-FORMATO", "critica", "codConsulta",
                   "codConsulta es obligatorio (6 caracteres CUPS).")
            elif len(cod_con) != 6:
                _e("C04-FORMATO", "critica", "codConsulta",
                   f"codConsulta debe tener 6 caracteres. Actual: {len(cod_con)}.", cod_con)

            # ── C05: modalidadGrupoServicioTecSal → catálogo ModalidadAtencion ─
            mod = normalizar_str(con.get("modalidadGrupoServicioTecSal") or "")
            if not mod:
                _e("C05-DOMINIO", "critica", "modalidadGrupoServicioTecSal",
                   "modalidadGrupoServicioTecSal es obligatorio (catálogo ModalidadAtencion).")
            elif len(mod) != 2:
                _e("C05-DOMINIO", "critica", "modalidadGrupoServicioTecSal",
                   f"modalidadGrupoServicioTecSal debe tener 2 caracteres. Actual: {len(mod)}.", mod)
            elif mod not in MODALIDADES_ATENCION:
                _e("C05-DOMINIO", "critica", "modalidadGrupoServicioTecSal",
                   f"modalidadGrupoServicioTecSal '{mod}' no pertenece al catálogo ModalidadAtencion. "
                   "Válidos: 01=Intramural, 02=Extramural unidad móvil, 03=Extramural domiciliaria, "
                   "04=Extramural jornada salud, 06=Telemedicina interactiva, 07=Telemedicina no interactiva, "
                   "08=Telemedicina telexperticia, 09=Telemedicina telemonitoreo.", mod)

            # ── C06: grupoServicios → catálogo GrupoServicios ─
            grupo = normalizar_str(con.get("grupoServicios") or "")
            if not grupo:
                _e("C06-DOMINIO", "critica", "grupoServicios",
                   "grupoServicios es obligatorio (catálogo GrupoServicios).")
            elif len(grupo) != 2:
                _e("C06-DOMINIO", "critica", "grupoServicios",
                   f"grupoServicios debe tener 2 caracteres. Actual: {len(grupo)}.", grupo)
            elif grupo not in GRUPOS_SERVICIOS:
                _e("C06-DOMINIO", "critica", "grupoServicios",
                   f"grupoServicios '{grupo}' no pertenece al catálogo GrupoServicios. "
                   "Válidos: 01=Consulta externa, 02=Apoyo diagnóstico, 03=Internación, "
                   "04=Quirúrgico, 05=Atención inmediata.", grupo)

            # ── C08: finalidadTecnologiaSalud → RIPSFinalidadConsultaVersion2 ─
            final = normalizar_str(con.get("finalidadTecnologiaSalud") or "")
            if not final:
                _e("C08-DOMINIO", "critica", "finalidadTecnologiaSalud",
                   "finalidadTecnologiaSalud es obligatorio (catálogo RIPSFinalidadConsultaVersion2, códigos 11-44).")
            elif len(final) != 2:
                _e("C08-DOMINIO", "critica", "finalidadTecnologiaSalud",
                   f"finalidadTecnologiaSalud debe tener 2 caracteres. Actual: {len(final)}.", final)
            elif final not in FINALIDADES_CONSULTA:
                _e("C08-DOMINIO", "critica", "finalidadTecnologiaSalud",
                   f"finalidadTecnologiaSalud '{final}' no pertenece al catálogo RIPSFinalidadConsultaVersion2 "
                   "(códigos 11 al 44).", final)

            # ── C09: causaMotivoAtencion → RIPSCausaExternaVersion2 ───────────
            causa = normalizar_str(con.get("causaMotivoAtencion") or "")
            if not causa:
                _e("C09-DOMINIO", "critica", "causaMotivoAtencion",
                   "causaMotivoAtencion es obligatorio (catálogo RIPSCausaExternaVersion2, códigos 21-49).")
            elif len(causa) != 2:
                _e("C09-DOMINIO", "critica", "causaMotivoAtencion",
                   f"causaMotivoAtencion debe tener 2 caracteres. Actual: {len(causa)}.", causa)
            elif causa not in CAUSAS_EXTERNAS:
                _e("C09-DOMINIO", "critica", "causaMotivoAtencion",
                   f"causaMotivoAtencion '{causa}' no pertenece al catálogo RIPSCausaExternaVersion2 "
                   "(códigos 21=Accidente trabajo ... 49=IVE voluntad semana 24).", causa)

            # ── C10 / RVC031: diagnóstico principal ──────────
            diag_p = normalizar_str(con.get("codDiagnosticoPrincipal") or "")
            if not diag_p:
                _e("C10-OBLIGATORIO", "critica", "codDiagnosticoPrincipal",
                   "codDiagnosticoPrincipal es obligatorio en consultas.")
            elif len(diag_p) >= 3 and diag_p[0] in ("V", "W", "X", "Y"):
                _e("RVC031", "media", "codDiagnosticoPrincipal",
                   f"codDiagnosticoPrincipal '{diag_p}' pertenece al rango de causas externas "
                   "CIE10 V01-Y98. No válido como diagnóstico principal (RVC031).", diag_p)

            # ── C11-C13 / RVC086/87: diagnósticos relacionados ─
            diags_rel = [
                normalizar_str(con.get("codDiagnosticoRelacionado1") or ""),
                normalizar_str(con.get("codDiagnosticoRelacionado2") or ""),
                normalizar_str(con.get("codDiagnosticoRelacionado3") or ""),
            ]
            diags_notnull = [d for d in diags_rel if d]
            for dr in diags_notnull:
                if dr == diag_p:
                    _e("RVC086", "media", "codDiagnosticoRelacionado",
                       f"Diagnóstico relacionado '{dr}' es igual al diagnóstico principal (RVC086).", dr)
            if len(diags_notnull) != len(set(diags_notnull)):
                _e("RVC087", "media", "codDiagnosticoRelacionado",
                   "Existen diagnósticos relacionados repetidos entre sí (RVC087).")

            # ── C14: tipoDiagnosticoPrincipal 2 chars ─────────
            tipo_diag = normalizar_str(con.get("tipoDiagnosticoPrincipal") or "")
            if not tipo_diag:
                _e("C14-DOMINIO", "critica", "tipoDiagnosticoPrincipal",
                   "tipoDiagnosticoPrincipal es obligatorio (2 caracteres).")
            elif len(tipo_diag) != 2:
                _e("C14-DOMINIO", "critica", "tipoDiagnosticoPrincipal",
                   f"tipoDiagnosticoPrincipal debe tener 2 caracteres. Actual: {len(tipo_diag)}.",
                   tipo_diag)

            # ── C15: tipoDoc profesional ──────────────────────
            tipo_doc_prof = _tipo_doc(con)
            if not tipo_doc_prof:
                _e("C15-DOMINIO", "critica", "tipoDocumentoIdentificacion (profesional)",
                   "tipoDocumentoIdentificacion del profesional es obligatorio.")
            elif tipo_doc_prof not in TIPOS_DOC_PROFESIONAL:
                _e("C15-DOMINIO", "critica", "tipoDocumentoIdentificacion (profesional)",
                   f"tipoDocumentoIdentificacion '{tipo_doc_prof}' no es válido para profesionales.",
                   tipo_doc_prof)

            # ── C16: numDoc profesional ───────────────────────
            num_doc_prof = _num_doc_val(con)
            if not num_doc_prof:
                _e("C16-OBLIGATORIO", "critica", "numDocumentoIdentificacion (profesional)",
                   "numDocumentoIdentificacion del profesional es obligatorio.")
            elif len(num_doc_prof) < 4 or len(num_doc_prof) > 20:
                _e("C16-OBLIGATORIO", "critica", "numDocumentoIdentificacion (profesional)",
                   f"numDocumentoIdentificacion del profesional debe tener 4-20 caracteres "
                   f"(actual: {len(num_doc_prof)}).", num_doc_prof)

            # ── C17: vrServicio >= 0 ──────────────────────────
            vrs_raw = con.get("vrServicio")
            if vrs_raw is not None:
                try:
                    if float(str(vrs_raw).strip()) < 0:
                        _e("C17-RANGO", "critica", "vrServicio",
                           "vrServicio no puede ser negativo.", vrs_raw)
                except (ValueError, TypeError):
                    _e("C17-RANGO", "critica", "vrServicio",
                       "vrServicio debe ser un valor numérico.", vrs_raw)

            # ── C18: conceptoRecaudo ──────────────────────────
            cr = normalizar_str(con.get("conceptoRecaudo") or "")
            if not cr:
                _e("C18-DOMINIO", "critica", "conceptoRecaudo",
                   "conceptoRecaudo es obligatorio.")
            elif cr not in CONCEPTOS_RECAUDO_CONSULTA:
                _e("C18-DOMINIO", "critica", "conceptoRecaudo",
                   f"conceptoRecaudo '{cr}' no válido para consultas. "
                   "Permitidos: 01=Copago, 02=Cuota moderadora, 03=Planes voluntarios, 05=No aplica.", cr)

            # ── C19 / RVC060-61: valorPagoModerador ──────────
            vpm_raw = con.get("valorPagoModerador")
            if vpm_raw is not None and cr:
                try:
                    vpm = float(str(vpm_raw).strip())
                    if cr in {"01", "03"} and vpm < 1:
                        _e("C19/RVC060", "critica", "valorPagoModerador",
                           f"Con conceptoRecaudo='{cr}', valorPagoModerador debe ser >= 1 (RVC060).",
                           vpm_raw)
                    if cr == "05" and vpm != 0:
                        _e("C19/RVC061", "critica", "valorPagoModerador",
                           "Con conceptoRecaudo='05' (No aplica), valorPagoModerador debe ser 0 (RVC061).",
                           vpm_raw)
                except (ValueError, TypeError):
                    _e("C19/RVC060", "critica", "valorPagoModerador",
                       "valorPagoModerador debe ser un valor numérico.", vpm_raw)

            # ── C20: numFEVPagoModerador null si CR=05 ────────
            if cr == "05":
                nfev = con.get("numFEVPagoModerador")
                if nfev is not None and normalizar_str(nfev) not in {"", "none", "null"}:
                    _e("C20-CONDICIONAL", "alta", "numFEVPagoModerador",
                       "numFEVPagoModerador debe ser null cuando conceptoRecaudo='05'.", nfev)

    return errores


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUE P: PROCEDIMIENTOS
# ══════════════════════════════════════════════════════════════

def validar_procedimientos_malla_2275(data, nombre_archivo=""):
    """
    Bloque P (Procedimientos): P01-P20 — análogo a consultas + idMIPRES, copago.
    """
    errores = []
    if not isinstance(data, dict):
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy = datetime.now()

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        fecha_nac_raw = normalizar_str(usuario.get("fechaNacimiento") or "")
        fecha_nac_dt  = None
        try:
            if len(fecha_nac_raw) == 10:
                fecha_nac_dt = datetime.strptime(fecha_nac_raw, "%Y-%m-%d")
        except ValueError:
            pass

        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        procedimientos = servicios.get("procedimientos", [])
        if not isinstance(procedimientos, list) or not procedimientos:
            continue

        consec_esperado = 1
        consecs_vistos  = set()

        for proc in procedimientos:
            if not isinstance(proc, dict):
                consec_esperado += 1
                continue

            consec   = proc.get("consecutivo")
            consec_s = normalizar_str(consec)
            ctx      = {"archivo": nombre_archivo, "num_factura": num_factura,
                        "num_doc": num_doc, "consecutivo": consec_s}

            def _e(id_regla, severidad, campo, mensaje, valor_actual=""):
                errores.append({**ctx, "id_regla": id_regla, "severidad": severidad,
                                "campo": campo, "mensaje": mensaje,
                                "valor_actual": normalizar_str(valor_actual)})

            # ── P20: Consecutivo ──────────────────────────────
            if isinstance(consec, int):
                if consec in consecs_vistos:
                    _e("P20-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} repetido en procedimientos.", consec)
                else:
                    consecs_vistos.add(consec)
                if consec != consec_esperado:
                    _e("P20-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} no secuencial (esperado: {consec_esperado}).", consec)
                consec_esperado = consec + 1
            else:
                consec_esperado += 1

            # ── P01 / RVC011: codPrestador 12 chars ──────────
            cod_prest = normalizar_str(proc.get("codPrestador") or "")
            if not cod_prest:
                _e("RVC011-P", "critica", "codPrestador",
                   "codPrestador es obligatorio en procedimientos (12 caracteres, RVC011).")
            elif len(cod_prest) != 12:
                _e("RVC011-P", "critica", "codPrestador",
                   f"codPrestador debe tener 12 caracteres (RVC011). Actual: {len(cod_prest)}.",
                   cod_prest)

            # ── P02 / RVC013: fechaInicioAtencion ────────────
            fecha_raw    = normalizar_str(proc.get("fechaInicioAtencion") or "")
            fecha_ini_dt = _validar_fecha_atencion_campo(
                fecha_raw, "fechaInicioAtencion", fecha_hoy, errores, ctx, "RVC013-P")
            if fecha_ini_dt and fecha_nac_dt and fecha_ini_dt < fecha_nac_dt:
                _e("RVC079-P", "critica", "fechaInicioAtencion",
                   "fechaInicioAtencion es menor a la fechaNacimiento del usuario.", fecha_raw)

            # ── P03: idMIPRES longitud ────────────────────────
            id_mipres = proc.get("idMIPRES")
            if id_mipres is not None:
                id_mipres_s = normalizar_str(id_mipres)
                if id_mipres_s and (len(id_mipres_s) < 1 or len(id_mipres_s) > 15):
                    _e("P03-CONDICIONAL", "alta", "idMIPRES",
                       f"idMIPRES debe tener entre 1 y 15 caracteres si se informa.", id_mipres_s)

            # ── P05: codProcedimiento 6 chars ─────────────────
            cod_proc = normalizar_str(proc.get("codProcedimiento") or "")
            if not cod_proc:
                _e("P05-FORMATO", "critica", "codProcedimiento",
                   "codProcedimiento es obligatorio (6 caracteres CUPS).")
            elif len(cod_proc) != 6:
                _e("P05-FORMATO", "critica", "codProcedimiento",
                   f"codProcedimiento debe tener 6 caracteres. Actual: {len(cod_proc)}.", cod_proc)

            # ── P08: grupoServicios → catálogo GrupoServicios ─
            grupo = normalizar_str(proc.get("grupoServicios") or "")
            if not grupo:
                _e("P08-DOMINIO", "critica", "grupoServicios",
                   "grupoServicios es obligatorio en procedimientos (catálogo GrupoServicios).")
            elif len(grupo) != 2:
                _e("P08-DOMINIO", "critica", "grupoServicios",
                   f"grupoServicios debe tener 2 caracteres. Actual: {len(grupo)}.", grupo)
            elif grupo not in GRUPOS_SERVICIOS:
                _e("P08-DOMINIO", "critica", "grupoServicios",
                   f"grupoServicios '{grupo}' no pertenece al catálogo GrupoServicios. "
                   "Válidos: 01=Consulta externa, 02=Apoyo diagnóstico, 03=Internación, "
                   "04=Quirúrgico, 05=Atención inmediata.", grupo)

            # ── P10: finalidadTecnologiaSalud → RIPSFinalidadConsultaVersion2 ─
            final = normalizar_str(proc.get("finalidadTecnologiaSalud") or "")
            if not final:
                _e("P10-DOMINIO", "critica", "finalidadTecnologiaSalud",
                   "finalidadTecnologiaSalud es obligatorio en procedimientos (catálogo RIPSFinalidadConsultaVersion2).")
            elif len(final) != 2:
                _e("P10-DOMINIO", "critica", "finalidadTecnologiaSalud",
                   f"finalidadTecnologiaSalud debe tener 2 caracteres. Actual: {len(final)}.", final)
            elif final not in FINALIDADES_CONSULTA:
                _e("P10-DOMINIO", "critica", "finalidadTecnologiaSalud",
                   f"finalidadTecnologiaSalud '{final}' no pertenece al catálogo RIPSFinalidadConsultaVersion2 "
                   "(códigos 11 al 44).", final)

            # ── P13 / RVC031: diagnóstico principal ──────────
            diag_p = normalizar_str(proc.get("codDiagnosticoPrincipal") or "")
            if not diag_p:
                _e("P13-OBLIGATORIO", "critica", "codDiagnosticoPrincipal",
                   "codDiagnosticoPrincipal es obligatorio en procedimientos.")
            elif len(diag_p) >= 3 and diag_p[0] in ("V", "W", "X", "Y"):
                _e("RVC031-P", "media", "codDiagnosticoPrincipal",
                   f"codDiagnosticoPrincipal '{diag_p}' pertenece a causas externas V01-Y98 (RVC031).",
                   diag_p)

            # ── P14: tipoDoc profesional ──────────────────────
            tipo_doc_prof = _tipo_doc(proc)
            if not tipo_doc_prof:
                _e("P14-DOMINIO", "critica", "tipoDocumentoIdentificacion (profesional)",
                   "tipoDocumentoIdentificacion del profesional es obligatorio en procedimientos.")
            elif tipo_doc_prof not in TIPOS_DOC_PROFESIONAL:
                _e("P14-DOMINIO", "critica", "tipoDocumentoIdentificacion (profesional)",
                   f"tipoDocumentoIdentificacion '{tipo_doc_prof}' no válido para profesionales.",
                   tipo_doc_prof)

            # ── P16: vrServicio >= 0 ──────────────────────────
            vrs_raw = proc.get("vrServicio")
            if vrs_raw is not None:
                try:
                    vrs_num_p = float(str(vrs_raw).strip())
                    if vrs_num_p < 0:
                        _e("P16-RANGO", "critica", "vrServicio",
                           "vrServicio no puede ser negativo.", vrs_raw)
                except (ValueError, TypeError):
                    vrs_num_p = None
                    _e("P16-RANGO", "critica", "vrServicio",
                       "vrServicio debe ser numérico.", vrs_raw)
            else:
                vrs_num_p = None

            # ── TERAPIA-MULTIPLE: facturación múltiple de terapia física/respiratoria ──
            # 931001 = Terapia Física  |  939403 = Terapia Respiratoria
            # Si el vrServicio supera $50.000 se infiere que se agruparon sesiones;
            # cada sesión debe facturarse de forma individual.
            TERAPIAS_INDIVIDUALES = {"931001": "Terapia Física", "939403": "Terapia Respiratoria"}
            if cod_proc in TERAPIAS_INDIVIDUALES and vrs_num_p is not None and vrs_num_p > 50000:
                nombre_terapia = TERAPIAS_INDIVIDUALES[cod_proc]
                _e("TERAPIA-MULTIPLE", "alta", "vrServicio",
                   f"Existen terapias físicas o respiratorias facturadas de forma múltiple "
                   f"(código {cod_proc} – {nombre_terapia}, vrServicio={vrs_num_p:,.0f}). "
                   "Cada sesión de terapia debe ser facturada individualmente.",
                   vrs_raw)

            # ── P17: conceptoRecaudo (admite copago=02) ───────
            cr = normalizar_str(proc.get("conceptoRecaudo") or "")
            if not cr:
                _e("P17-COPAGO", "critica", "conceptoRecaudo",
                   "conceptoRecaudo es obligatorio en procedimientos.")
            elif cr not in CONCEPTOS_RECAUDO_PROC:
                _e("P17-COPAGO", "critica", "conceptoRecaudo",
                   f"conceptoRecaudo '{cr}' no válido para procedimientos (01,02,03,05).", cr)

            # ── P18 / RVC060-61: valorPagoModerador ──────────
            vpm_raw = proc.get("valorPagoModerador")
            if vpm_raw is not None and cr:
                try:
                    vpm = float(str(vpm_raw).strip())
                    if cr in {"01", "03"} and vpm < 1:
                        _e("P18/RVC060", "critica", "valorPagoModerador",
                           f"Con conceptoRecaudo='{cr}', valorPagoModerador debe ser >= 1.", vpm_raw)
                    if cr == "05" and vpm != 0:
                        _e("P18/RVC061", "critica", "valorPagoModerador",
                           "Con conceptoRecaudo='05', valorPagoModerador debe ser 0.", vpm_raw)
                except (ValueError, TypeError):
                    _e("P18/RVC060", "critica", "valorPagoModerador",
                       "valorPagoModerador debe ser numérico.", vpm_raw)

            # ── numFEVPagoModerador null si CR=05 ─────────────
            if cr == "05":
                nfev = proc.get("numFEVPagoModerador")
                if nfev is not None and normalizar_str(nfev) not in {"", "none", "null"}:
                    _e("P-FEV-CONDICIONAL", "alta", "numFEVPagoModerador",
                       "numFEVPagoModerador debe ser null cuando conceptoRecaudo='05'.", nfev)

    return errores


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUE R: URGENCIAS
# ══════════════════════════════════════════════════════════════

def validar_urgencias_malla_2275(data, nombre_archivo=""):
    """
    Bloque R (Urgencias): R01-R12, RVC038, RVC040, RVC042, RVC043.
    """
    errores = []
    if not isinstance(data, dict):
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy = datetime.now()

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = normalizar_str(
            _num_doc_val(usuario)
        )
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        urgencias = servicios.get("urgencias", [])
        if not isinstance(urgencias, list) or not urgencias:
            continue

        consec_esperado = 1
        consecs_vistos  = set()

        for urg in urgencias:
            if not isinstance(urg, dict):
                consec_esperado += 1
                continue

            consec   = urg.get("consecutivo")
            consec_s = normalizar_str(consec)
            ctx      = {"archivo": nombre_archivo, "num_factura": num_factura,
                        "num_doc": num_doc, "consecutivo": consec_s}

            def _e(id_regla, severidad, campo, mensaje, valor_actual=""):
                errores.append({**ctx, "id_regla": id_regla, "severidad": severidad,
                                "campo": campo, "mensaje": mensaje,
                                "valor_actual": normalizar_str(valor_actual)})

            # ── R12: Consecutivo ──────────────────────────────
            if isinstance(consec, int):
                if consec in consecs_vistos:
                    _e("R12-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} repetido en urgencias.", consec)
                else:
                    consecs_vistos.add(consec)
                if consec != consec_esperado:
                    _e("R12-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} no secuencial (esperado: {consec_esperado}).", consec)
                consec_esperado = consec + 1
            else:
                consec_esperado += 1

            # ── R01 / RVC011: codPrestador 12 chars ──────────
            cod_prest = normalizar_str(urg.get("codPrestador") or "")
            if not cod_prest:
                _e("RVC011-R", "critica", "codPrestador",
                   "codPrestador es obligatorio en urgencias (12 caracteres, RVC011).")
            elif len(cod_prest) != 12:
                _e("RVC011-R", "critica", "codPrestador",
                   f"codPrestador debe tener 12 caracteres (RVC011). Actual: {len(cod_prest)}.",
                   cod_prest)

            # ── R02 / RVC038: fechaInicioAtencion ────────────
            f_ini_raw = normalizar_str(urg.get("fechaInicioAtencion") or "")
            f_ini_dt  = _validar_fecha_atencion_campo(
                f_ini_raw, "fechaInicioAtencion", fecha_hoy, errores, ctx, "RVC038")

            # ── R11 / RVC043: fechaEgreso ─────────────────────
            f_egr_raw = normalizar_str(urg.get("fechaEgreso") or "")
            f_egr_dt  = None
            if not f_egr_raw:
                _e("RVC043", "critica", "fechaEgreso",
                   "fechaEgreso es obligatorio en urgencias.")
            elif len(f_egr_raw) != 16:
                _e("RVC043", "critica", "fechaEgreso",
                   f"fechaEgreso debe tener formato AAAA-MM-DD HH:MM. "
                   f"Longitud actual: {len(f_egr_raw)}.", f_egr_raw)
            else:
                try:
                    f_egr_dt = datetime.strptime(f_egr_raw, "%Y-%m-%d %H:%M")
                    if f_egr_dt > fecha_hoy:
                        _e("RVC043", "critica", "fechaEgreso",
                           "fechaEgreso es mayor a la fecha/hora actual (RVC043).", f_egr_raw)
                except ValueError:
                    _e("RVC043", "critica", "fechaEgreso",
                       "fechaEgreso no es una fecha/hora válida (AAAA-MM-DD HH:MM).", f_egr_raw)

            # Ingreso <= egreso
            if f_ini_dt and f_egr_dt and f_ini_dt > f_egr_dt:
                _e("RVC038", "critica", "fechaInicioAtencion",
                   "fechaInicioAtencion es mayor a fechaEgreso (RVC038).", f_ini_raw)

            # ── RVC040: estancia <= 48h (notificación) ────────
            if f_ini_dt and f_egr_dt:
                horas = (f_egr_dt - f_ini_dt).total_seconds() / 3600
                if horas > 48:
                    _e("RVC040", "media", "fechaEgreso",
                       f"Estancia en urgencias de {horas:.1f} horas supera las 48 horas (RVC040).",
                       f_egr_raw)

            # ── R03: causaMotivoAtencion → RIPSCausaExternaVersion2 ──────────
            causa = normalizar_str(urg.get("causaMotivoAtencion") or "")
            if not causa:
                _e("R03-DOMINIO", "critica", "causaMotivoAtencion",
                   "causaMotivoAtencion es obligatorio en urgencias (catálogo RIPSCausaExternaVersion2, códigos 21-49).")
            elif len(causa) != 2:
                _e("R03-DOMINIO", "critica", "causaMotivoAtencion",
                   f"causaMotivoAtencion debe tener 2 caracteres. Actual: {len(causa)}.", causa)
            elif causa not in CAUSAS_EXTERNAS:
                _e("R03-DOMINIO", "critica", "causaMotivoAtencion",
                   f"causaMotivoAtencion '{causa}' no pertenece al catálogo RIPSCausaExternaVersion2 "
                   "(códigos 21=Accidente trabajo ... 49=IVE voluntad semana 24).", causa)

            # ── R04: codDiagnosticoPrincipal ──────────────────
            diag_p = normalizar_str(urg.get("codDiagnosticoPrincipal") or "")
            if not diag_p:
                _e("R04-OBLIGATORIO", "critica", "codDiagnosticoPrincipal",
                   "codDiagnosticoPrincipal es obligatorio en urgencias.")

            # ── R09: condicionDestinoUsuarioEgreso → catálogo CondicionyDestinoUsuarioEgreso ─
            cond_egr = normalizar_str(urg.get("condicionDestinoUsuarioEgreso") or "")
            if not cond_egr:
                _e("R09-OBLIGATORIO", "critica", "condicionDestinoUsuarioEgreso",
                   "condicionDestinoUsuarioEgreso es obligatorio en urgencias.")
            elif cond_egr not in COND_EGRESO_VALIDAS:
                _e("R09-DOMINIO", "critica", "condicionDestinoUsuarioEgreso",
                   f"condicionDestinoUsuarioEgreso '{cond_egr}' no pertenece al catálogo. "
                   "Válidos: 01=A domicilio, 02=Muerto, 03=Derivado otro servicio, 04=Referido otra institución, "
                   "05=Contrareferido, 06=Hospitalización domiciliaria, 07=Servicio social, 08=Continúa en servicio.",
                   cond_egr)

            # ── R10 / RVC042: causa muerte si condición = muerto ─
            if cond_egr in COND_EGRESO_MUERTO:
                cod_muerte = normalizar_str(urg.get("codDiagnosticoCausaMuerte") or "")
                if not cod_muerte:
                    _e("RVC042", "critica", "codDiagnosticoCausaMuerte",
                       "codDiagnosticoCausaMuerte es obligatorio cuando "
                       "condicionDestinoUsuarioEgreso indica paciente muerto (RVC042).")

    return errores


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUE H: HOSPITALIZACIÓN
# ══════════════════════════════════════════════════════════════

def validar_hospitalizacion_malla_2275(data, nombre_archivo=""):
    """
    Bloque H (Hospitalización): H01-H15, RVC011, RVC013, RVC041, RVC042.
    """
    errores = []
    if not isinstance(data, dict):
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy = datetime.now()

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = normalizar_str(
            _num_doc_val(usuario)
        )
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        hospitalizacion = servicios.get("hospitalizacion", [])
        if not isinstance(hospitalizacion, list) or not hospitalizacion:
            continue

        consec_esperado = 1
        consecs_vistos  = set()

        for hosp in hospitalizacion:
            if not isinstance(hosp, dict):
                consec_esperado += 1
                continue

            consec   = hosp.get("consecutivo")
            consec_s = normalizar_str(consec)
            ctx      = {"archivo": nombre_archivo, "num_factura": num_factura,
                        "num_doc": num_doc, "consecutivo": consec_s}

            def _e(id_regla, severidad, campo, mensaje, valor_actual=""):
                errores.append({**ctx, "id_regla": id_regla, "severidad": severidad,
                                "campo": campo, "mensaje": mensaje,
                                "valor_actual": normalizar_str(valor_actual)})

            # ── H15: Consecutivo ──────────────────────────────
            if isinstance(consec, int):
                if consec in consecs_vistos:
                    _e("H15-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} repetido en hospitalización.", consec)
                else:
                    consecs_vistos.add(consec)
                if consec != consec_esperado:
                    _e("H15-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} no secuencial (esperado: {consec_esperado}).", consec)
                consec_esperado = consec + 1
            else:
                consec_esperado += 1

            # ── H01 / RVC011: codPrestador 12 chars ──────────
            cod_prest = normalizar_str(hosp.get("codPrestador") or "")
            if not cod_prest:
                _e("RVC011-H", "critica", "codPrestador",
                   "codPrestador es obligatorio en hospitalización (12 caracteres, RVC011).")
            elif len(cod_prest) != 12:
                _e("RVC011-H", "critica", "codPrestador",
                   f"codPrestador debe tener 12 caracteres (RVC011). Actual: {len(cod_prest)}.",
                   cod_prest)

            # ── H03 / RVC013: fechaInicioAtencion ────────────
            f_ini_raw = normalizar_str(hosp.get("fechaInicioAtencion") or "")
            f_ini_dt  = _validar_fecha_atencion_campo(
                f_ini_raw, "fechaInicioAtencion", fecha_hoy, errores, ctx, "RVC013-H")

            # ── H14: fechaEgreso ──────────────────────────────
            f_egr_raw = normalizar_str(hosp.get("fechaEgreso") or "")
            f_egr_dt  = None
            if not f_egr_raw:
                _e("H14-OBLIGATORIO", "critica", "fechaEgreso",
                   "fechaEgreso es obligatorio en hospitalización.")
            elif len(f_egr_raw) != 16:
                _e("H14-OBLIGATORIO", "critica", "fechaEgreso",
                   f"fechaEgreso debe tener formato AAAA-MM-DD HH:MM. "
                   f"Longitud: {len(f_egr_raw)}.", f_egr_raw)
            else:
                try:
                    f_egr_dt = datetime.strptime(f_egr_raw, "%Y-%m-%d %H:%M")
                    if f_egr_dt > fecha_hoy:
                        _e("H14-OBLIGATORIO", "critica", "fechaEgreso",
                           "fechaEgreso es mayor a la fecha/hora actual.", f_egr_raw)
                except ValueError:
                    _e("H14-OBLIGATORIO", "critica", "fechaEgreso",
                       "fechaEgreso no es una fecha/hora válida.", f_egr_raw)

            # Inicio <= egreso
            if f_ini_dt and f_egr_dt and f_ini_dt > f_egr_dt:
                _e("H-CRONOLOGIA", "critica", "fechaInicioAtencion",
                   "fechaInicioAtencion es mayor a fechaEgreso en hospitalización.", f_ini_raw)

            # ── RVC041: estancia < 6 horas (notificación) ─────
            if f_ini_dt and f_egr_dt:
                horas = (f_egr_dt - f_ini_dt).total_seconds() / 3600
                if horas < 6:
                    _e("RVC041", "media", "fechaEgreso",
                       f"Estancia en hospitalización de {horas:.1f} horas es menor a 6 horas (RVC041).",
                       f_egr_raw)

            # ── H09: condicionDestinoUsuarioEgreso → catálogo CondicionyDestinoUsuarioEgreso ─
            cond_egr = normalizar_str(hosp.get("condicionDestinoUsuarioEgreso") or "")
            if not cond_egr:
                _e("H09-OBLIGATORIO", "critica", "condicionDestinoUsuarioEgreso",
                   "condicionDestinoUsuarioEgreso es obligatorio en hospitalización.")
            elif cond_egr not in COND_EGRESO_VALIDAS:
                _e("H09-DOMINIO", "critica", "condicionDestinoUsuarioEgreso",
                   f"condicionDestinoUsuarioEgreso '{cond_egr}' no pertenece al catálogo. "
                   "Válidos: 01=A domicilio, 02=Muerto, 03=Derivado otro servicio, 04=Referido otra institución, "
                   "05=Contrareferido, 06=Hospitalización domiciliaria, 07=Servicio social, 08=Continúa en servicio.",
                   cond_egr)

            # ── H10 / RVC042: causa muerte ────────────────────
            if cond_egr in COND_EGRESO_MUERTO:
                cod_muerte = normalizar_str(hosp.get("codDiagnosticoCausaMuerte") or "")
                if not cod_muerte:
                    _e("RVC042-H", "critica", "codDiagnosticoCausaMuerte",
                       "codDiagnosticoCausaMuerte es obligatorio cuando el paciente fallece "
                       "en hospitalización (RVC042).")

    return errores


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUE N: RECIÉN NACIDOS
# ══════════════════════════════════════════════════════════════

def validar_recien_nacidos_malla_2275(data, nombre_archivo=""):
    """
    Bloque N (Recién Nacidos): N01-N13, RVC045, RVC046, RVC057, RVC058.
    """
    errores = []
    if not isinstance(data, dict):
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy = datetime.now()

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = normalizar_str(
            _num_doc_val(usuario)
        )
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        recien_nacidos = servicios.get("recienNacidos", [])
        if not isinstance(recien_nacidos, list) or not recien_nacidos:
            continue

        consec_esperado = 1
        consecs_vistos  = set()

        for rn in recien_nacidos:
            if not isinstance(rn, dict):
                consec_esperado += 1
                continue

            consec   = rn.get("consecutivo")
            consec_s = normalizar_str(consec)
            ctx      = {"archivo": nombre_archivo, "num_factura": num_factura,
                        "num_doc": num_doc, "consecutivo": consec_s}

            def _e(id_regla, severidad, campo, mensaje, valor_actual=""):
                errores.append({**ctx, "id_regla": id_regla, "severidad": severidad,
                                "campo": campo, "mensaje": mensaje,
                                "valor_actual": normalizar_str(valor_actual)})

            # ── N13: Consecutivo ──────────────────────────────
            if isinstance(consec, int):
                if consec in consecs_vistos:
                    _e("N13-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} repetido en recién nacidos.", consec)
                else:
                    consecs_vistos.add(consec)
                if consec != consec_esperado:
                    _e("N13-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} no secuencial (esperado: {consec_esperado}).", consec)
                consec_esperado = consec + 1
            else:
                consec_esperado += 1

            # ── N01 / RVC011: codPrestador 12 chars ──────────
            cod_prest = normalizar_str(rn.get("codPrestador") or "")
            if not cod_prest:
                _e("RVC011-N", "critica", "codPrestador",
                   "codPrestador es obligatorio en recién nacidos (12 caracteres, RVC011).")
            elif len(cod_prest) != 12:
                _e("RVC011-N", "critica", "codPrestador",
                   f"codPrestador debe tener 12 caracteres (RVC011). Actual: {len(cod_prest)}.",
                   cod_prest)

            # ── N02: tipoDocumentoIdentificacion CN/RC/MS ─────
            tipo_doc_rn = _tipo_doc(rn)
            if not tipo_doc_rn:
                _e("N02-DOMINIO", "critica", "tipoDocumentoIdentificacion",
                   "tipoDocumentoIdentificacion del recién nacido es obligatorio (CN/RC/MS).")
            elif tipo_doc_rn not in TIPOS_DOC_RN:
                _e("N02-DOMINIO", "critica", "tipoDocumentoIdentificacion",
                   f"tipoDocumentoIdentificacion '{tipo_doc_rn}' no válido para recién nacidos "
                   "(solo CN, RC o MS).", tipo_doc_rn)

            # ── N04 / RVC045: fechaNacimiento con hora ────────
            fecha_nac_rn_raw = normalizar_str(rn.get("fechaNacimiento") or "")
            fecha_nac_rn_dt  = None
            if not fecha_nac_rn_raw:
                _e("RVC045", "critica", "fechaNacimiento",
                   "fechaNacimiento del recién nacido es obligatorio (formato AAAA-MM-DD HH:MM).")
            elif len(fecha_nac_rn_raw) != 16:
                _e("RVC045", "critica", "fechaNacimiento",
                   f"fechaNacimiento del recién nacido debe tener formato AAAA-MM-DD HH:MM "
                   f"(16 caracteres). Actual: {len(fecha_nac_rn_raw)}.", fecha_nac_rn_raw)
            else:
                try:
                    fecha_nac_rn_dt = datetime.strptime(fecha_nac_rn_raw, "%Y-%m-%d %H:%M")
                    if fecha_nac_rn_dt > fecha_hoy:
                        _e("RVC045", "critica", "fechaNacimiento",
                           "fechaNacimiento del recién nacido es mayor a la fecha/hora actual.",
                           fecha_nac_rn_raw)
                except ValueError:
                    _e("RVC045", "critica", "fechaNacimiento",
                       "fechaNacimiento del recién nacido no es una fecha/hora válida.",
                       fecha_nac_rn_raw)

            # ── N05 / RVC057: edadGestacional 20-46 ──────────
            edad_gest = rn.get("edadGestacional")
            if edad_gest is not None:
                try:
                    eg = int(str(edad_gest).strip())
                    if not (20 <= eg <= 46):
                        _e("RVC057", "media", "edadGestacional",
                           f"edadGestacional {eg} no está entre 20 y 46 semanas (RVC057).",
                           edad_gest)
                except (ValueError, TypeError):
                    _e("RVC057", "media", "edadGestacional",
                       "edadGestacional debe ser un número entero (semanas).", edad_gest)

            # ── N08 / RVC058: peso 500-5000 g ─────────────────
            peso = rn.get("peso")
            if peso is not None:
                try:
                    p = int(str(peso).strip())
                    if not (500 <= p <= 5000):
                        _e("RVC058", "media", "peso",
                           f"Peso del recién nacido {p}g no está entre 500 y 5000 gramos (RVC058).",
                           peso)
                except (ValueError, TypeError):
                    _e("RVC058", "media", "peso",
                       "peso debe ser un valor numérico en gramos.", peso)

            # ── N10: condicionDestinoUsuarioEgreso → catálogo CondicionyDestinoUsuarioEgreso ─
            cond_egr = normalizar_str(rn.get("condicionDestinoUsuarioEgreso") or "")
            if not cond_egr:
                _e("N10-OBLIGATORIO", "critica", "condicionDestinoUsuarioEgreso",
                   "condicionDestinoUsuarioEgreso es obligatorio en recién nacidos.")
            elif cond_egr not in COND_EGRESO_VALIDAS:
                _e("N10-DOMINIO", "critica", "condicionDestinoUsuarioEgreso",
                   f"condicionDestinoUsuarioEgreso '{cond_egr}' no pertenece al catálogo. "
                   "Válidos: 01=A domicilio, 02=Muerto, 03=Derivado otro servicio, 04=Referido otra institución, "
                   "05=Contrareferido, 06=Hospitalización domiciliaria, 07=Servicio social, 08=Continúa en servicio.",
                   cond_egr)

            # ── N11 / RVC042: causa muerte ────────────────────
            if cond_egr in COND_EGRESO_MUERTO:
                cod_muerte = normalizar_str(rn.get("codDiagnosticoCausaMuerte") or "")
                if not cod_muerte:
                    _e("RVC042-N", "critica", "codDiagnosticoCausaMuerte",
                       "codDiagnosticoCausaMuerte es obligatorio cuando el recién nacido "
                       "fallece (RVC042).")

            # ── N12 / RVC046: fechaEgreso >= fechaNacimiento ──
            f_egr_raw = normalizar_str(rn.get("fechaEgreso") or "")
            if not f_egr_raw:
                _e("RVC046", "critica", "fechaEgreso",
                   "fechaEgreso del recién nacido es obligatorio.")
            else:
                try:
                    f_egr_dt = datetime.strptime(f_egr_raw[:16], "%Y-%m-%d %H:%M")
                    if fecha_nac_rn_dt and f_egr_dt < fecha_nac_rn_dt:
                        _e("RVC046", "critica", "fechaEgreso",
                           f"fechaEgreso ({f_egr_raw}) es menor a fechaNacimiento del recién nacido "
                           f"({fecha_nac_rn_raw}) (RVC046).", f_egr_raw)
                except ValueError:
                    _e("RVC046", "critica", "fechaEgreso",
                       "fechaEgreso del recién nacido no es una fecha/hora válida.", f_egr_raw)

    # ── Tamizaje Neonatal Metabólico completo ─────────────────────────────────
    TAMIZAJE_NEONATAL  = {"904509", "904902", "906958", "903301", "908316", "908854", "908355"}
    ESTANCIA_UCI_INT   = {"108A01", "105M01"}
    LIMITE_NEONATAL_H  = 720   # solo evaluar pacientes con < 30 días de vida

    # Factura del archivo (reutilizado en todos los sub-bloques)
    num_factura_tn = ""
    for k in ("numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"):
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura_tn = normalizar_str(v)
            break

    # ── Recolectar códigos de procedimientos y otrosServicios del JSON ────────
    def _recolectar_procedimientos(data_obj):
        """Retorna set de codProcedimiento de todos los usuarios."""
        codigos = set()
        for u in data_obj.get("usuarios", []):
            if not isinstance(u, dict):
                continue
            svc = u.get("servicios", {})
            if not isinstance(svc, dict):
                continue
            for proc in svc.get("procedimientos", []):
                if isinstance(proc, dict):
                    cod = normalizar_str(proc.get("codProcedimiento") or "")
                    if cod:
                        codigos.add(cod)
        return codigos

    def _recolectar_otros_servicios(data_obj, fecha_limite_dt=None):
        """
        Retorna set de codTecnologiaSalud de otrosServicios.
        Si fecha_limite_dt se indica, solo incluye registros con
        fechaSuministroTecnologia / fechaInicioAtencion <= fecha_limite_dt.
        """
        codigos = set()
        for u in data_obj.get("usuarios", []):
            if not isinstance(u, dict):
                continue
#if not isinstance(u, dict):
            svc = u.get("servicios", {})
            if not isinstance(svc, dict):
                continue
            for os_rec in svc.get("otrosServicios", []):
                if not isinstance(os_rec, dict):
                    continue
                cod = normalizar_str(os_rec.get("codTecnologiaSalud") or "")
                if not cod:
                    continue
                if fecha_limite_dt is not None:
                    f_raw = normalizar_str(
                        os_rec.get("fechaSuministroTecnologia") or
                        os_rec.get("fechaInicioAtencion") or ""
                    )
                    if f_raw:
                        try:
                            f_dt = datetime.strptime(f_raw[:16], "%Y-%m-%d %H:%M")
                            if f_dt <= fecha_limite_dt:
                                codigos.add(cod)
                        except ValueError:
                            pass
                    # si no tiene fecha, no se incluye cuando hay límite
                else:
                    codigos.add(cod)
        return codigos

    # ── CASO 1: existe arreglo recienNacidos → validar los 7 códigos ──────────
    hay_rn = any(
        isinstance(u, dict) and
        isinstance(u.get("servicios", {}).get("recienNacidos"), list) and
        u["servicios"]["recienNacidos"]
        for u in data.get("usuarios", [])
    )
    if hay_rn:
        codigos_en_json = _recolectar_procedimientos(data)
        faltantes = sorted(TAMIZAJE_NEONATAL - codigos_en_json)
        if faltantes:
            errores.append({
                "archivo":      nombre_archivo,
                "num_factura":  num_factura_tn,
                "num_doc":      "",
                "consecutivo":  "",
                "id_regla":     "TAMIZAJE-N",
                "severidad":    "alta",
                "campo":        "codProcedimiento",
                "mensaje":      "Se detecta que no está completo el Tamizaje Neonatal Metabólico completo, "
                                f"faltan los siguientes códigos: {', '.join(faltantes)}.",
                "valor_actual": ", ".join(faltantes),
            })

    # ── CASO 2: NO existe recienNacidos → evaluar por fechaNacimiento ─────────
    else:
        codigos_proc_json = _recolectar_procedimientos(data)
        tiene_tamizaje    = bool(TAMIZAJE_NEONATAL & codigos_proc_json)

        for usuario in data.get("usuarios", []):
            if not isinstance(usuario, dict):
                continue

            fecha_nac_raw = normalizar_str(usuario.get("fechaNacimiento") or "")
            if len(fecha_nac_raw) != 10:
                continue
            try:
                fecha_nac_dt = datetime.strptime(fecha_nac_raw, "%Y-%m-%d")
            except ValueError:
                continue

            num_doc_u = _num_doc_val(usuario)
            svc = usuario.get("servicios", {})
            if not isinstance(svc, dict):
                continue

            # Recolectar todas las fechaInicioAtencion del usuario
            fechas_ini = []
            for seccion in ("consultas", "procedimientos", "urgencias", "hospitalizacion"):
                for reg in svc.get(seccion, []):
                    if not isinstance(reg, dict):
                        continue
                    f_raw = normalizar_str(reg.get("fechaInicioAtencion") or "")
                    if len(f_raw) >= 16:
                        try:
                            fechas_ini.append(
                                datetime.strptime(f_raw[:16], "%Y-%m-%d %H:%M")
                            )
                        except ValueError:
                            pass

            if not fechas_ini:
                continue

            fecha_ini_min = min(fechas_ini)
            edad_horas    = (fecha_ini_min - fecha_nac_dt).total_seconds() / 3600

            # Solo aplica a neonatos (< 30 días de vida al inicio de la atención)
            if not (0 <= edad_horas < LIMITE_NEONATAL_H):
                continue

            if tiene_tamizaje:
                continue   # ya tiene los códigos, no alertar

            if edad_horas < 72:
                # ── Caso 2A: neonato < 72 horas → debería realizarse en esta atención
                errores.append({
                    "archivo":      nombre_archivo,
                    "num_factura":  num_factura_tn,
                    "num_doc":      num_doc_u,
                    "consecutivo":  "",
                    "id_regla":     "TAMIZAJE-N",
                    "severidad":    "alta",
                    "campo":        "codProcedimiento",
                    "mensaje":      f"Usuario con fecha de nacimiento '{fecha_nac_raw}', "
                                    "no se evidencia facturación de los códigos Tamizaje Neonatal "
                                    "durante la atención, validar su realización.",
                    "valor_actual": fecha_nac_raw,
                })
            else:
                # ── Caso 2B: neonato ≥ 72 horas → verificar si hubo estancia UCI/intermedio
                #    en las primeras 72h de vida que justifique la no realización
                fecha_limite_72h  = fecha_nac_dt + timedelta(hours=72)
                codigos_estancia  = _recolectar_otros_servicios(data, fecha_limite_72h)
                tiene_estancia_72 = bool(ESTANCIA_UCI_INT & codigos_estancia)

                if not tiene_estancia_72:
                    errores.append({
                        "archivo":      nombre_archivo,
                        "num_factura":  num_factura_tn,
                        "num_doc":      num_doc_u,
                        "consecutivo":  "",
                        "id_regla":     "TAMIZAJE-N",
                        "severidad":    "alta",
                        "campo":        "codProcedimiento",
                        "mensaje":      "No se evidencia realización del Tamizaje Neonatal durante las 72 horas "
                                        "de vida y no se evidencia facturación de estancia en UCI o intermedio "
                                        "que justifique la no realización.",
                        "valor_actual": fecha_nac_raw,
                    })

    return errores


# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 2275/2023 – BLOQUE S: OTROS SERVICIOS
# ══════════════════════════════════════════════════════════════

def validar_otros_servicios_malla_2275(data, nombre_archivo=""):
    """
    Bloque S (Otros Servicios): S01-S16, RVC013, RVC050, S08 honorarios cantidad=1.
    """
    errores = []
    if not isinstance(data, dict):
        return errores

    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break

    fecha_hoy = datetime.now()

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = normalizar_str(
            _num_doc_val(usuario)
        )
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        otros = servicios.get("otrosServicios", [])
        if not isinstance(otros, list) or not otros:
            continue

        consec_esperado = 1
        consecs_vistos  = set()

        for svc in otros:
            if not isinstance(svc, dict):
                consec_esperado += 1
                continue

            consec   = svc.get("consecutivo")
            consec_s = normalizar_str(consec)
            ctx      = {"archivo": nombre_archivo, "num_factura": num_factura,
                        "num_doc": num_doc, "consecutivo": consec_s}

            def _e(id_regla, severidad, campo, mensaje, valor_actual=""):
                errores.append({**ctx, "id_regla": id_regla, "severidad": severidad,
                                "campo": campo, "mensaje": mensaje,
                                "valor_actual": normalizar_str(valor_actual)})

            # ── S-CONSECUTIVO ─────────────────────────────────
            if isinstance(consec, int):
                if consec in consecs_vistos:
                    _e("S-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} repetido en otrosServicios.", consec)
                else:
                    consecs_vistos.add(consec)
                if consec != consec_esperado:
                    _e("S-CONSECUTIVO", "critica", "consecutivo",
                       f"Consecutivo {consec} no secuencial (esperado: {consec_esperado}).", consec)
                consec_esperado = consec + 1
            else:
                consec_esperado += 1

            # ── S02 / RVC013: fecha de suministro (dos variantes de campo) ─
            fecha_raw = normalizar_str(
                svc.get("fechaInicioAtencion") or svc.get("fechaSuministroTecnologia") or ""
            )
            nombre_campo_fecha = ("fechaInicioAtencion" if svc.get("fechaInicioAtencion")
                                  else "fechaSuministroTecnologia")
            if fecha_raw:
                _validar_fecha_atencion_campo(
                    fecha_raw, nombre_campo_fecha, fecha_hoy, errores, ctx, "RVC013-S")

            # ── S05: tipoOS → catálogo TipoOtrosServicios ─────
            tipo_os = normalizar_str(svc.get("tipoOS") or "")
            if not tipo_os:
                _e("S05-DOMINIO", "critica", "tipoOS",
                   "tipoOS es obligatorio en otrosServicios (catálogo TipoOtrosServicios).")
            elif len(tipo_os) != 2:
                _e("S05-DOMINIO", "critica", "tipoOS",
                   f"tipoOS debe tener 2 caracteres. Actual: {len(tipo_os)}.", tipo_os)
            elif tipo_os not in TIPOS_OS_VALIDOS:
                _e("S05-DOMINIO", "critica", "tipoOS",
                   f"tipoOS '{tipo_os}' no pertenece al catálogo TipoOtrosServicios. "
                   "Válidos: 01=Dispositivos médicos e insumos, 02=Traslados, 03=Estancias, "
                   "04=Servicios complementarios, 05=Honorarios, 06=Servicios salud comunidades indígenas.",
                   tipo_os)

            # ── S06: codTecnologiaSalud obligatorio ───────────
            cod_tec = normalizar_str(svc.get("codTecnologiaSalud") or "")
            if not cod_tec:
                _e("S06-OBLIGATORIO", "critica", "codTecnologiaSalud",
                   "codTecnologiaSalud es obligatorio en otrosServicios.")

            # ── S08: cantidadOS solo para códigos de 6 chars sin letras (honorarios CUPS) ──
            # Solo se valida cuando el código tiene exactamente 6 caracteres y todos son dígitos.
            # Para cualquier otro tipo (DM*, traslados con letras, estancias, etc.) no se genera alerta.
            cantidad_os = svc.get("cantidadOS")
            es_honorario_cups = len(cod_tec) == 6 and cod_tec.isdigit()

            if es_honorario_cups:
                try:
                    if int(str(cantidad_os).strip()) != 1:
                        _e("S08-HONORARIOS", "alta", "cantidadOS",
                           f"Para honorarios (código '{cod_tec}') cantidadOS debe ser 1. "
                           "Solo se permite reportar 1 honorario por procedimiento y por profesional.",
                           cantidad_os)
                except (ValueError, TypeError, AttributeError):
                    _e("S08-HONORARIOS", "alta", "cantidadOS",
                       f"Para honorarios (código '{cod_tec}') cantidadOS debe ser un valor numérico entero igual a 1. "
                       "Solo se permite reportar 1 honorario por procedimiento y por profesional.",
                       cantidad_os)

            # ── RVC050: S09-S10 obligatorios para DM/SC/HO ───
            if tipo_os in TIPOS_OS_CON_PRESCRIPTOR:
                tipo_doc_ord = _tipo_doc(svc)
                num_doc_ord  = _num_doc_val(svc)
                if not tipo_doc_ord:
                    _e("RVC050", "critica", "tipoDocumentoIdentificacion (ordenante)",
                       f"tipoDocumentoIdentificacion del ordenante es obligatorio para "
                       f"tipoOS='{tipo_os}' (RVC050).")
                if not num_doc_ord:
                    _e("RVC050", "critica", "numDocumentoIdentificacion (ordenante)",
                       f"numDocumentoIdentificacion del ordenante es obligatorio para "
                       f"tipoOS='{tipo_os}' (RVC050).")

            # ── vrServicio >= 0 ───────────────────────────────
            vrs_raw = svc.get("vrServicio")
            if vrs_raw is not None:
                try:
                    if float(str(vrs_raw).strip()) < 0:
                        _e("S-VRSERVICIO", "critica", "vrServicio",
                           "vrServicio no puede ser negativo.", vrs_raw)
                except (ValueError, TypeError):
                    _e("S-VRSERVICIO", "critica", "vrServicio",
                       "vrServicio debe ser un valor numérico.", vrs_raw)

            # ── conceptoRecaudo (si presente) ─────────────────
            cr = normalizar_str(svc.get("conceptoRecaudo") or "")
            if cr and cr not in CONCEPTOS_RECAUDO_PROC:
                _e("S-CONCEPTO-RECAUDO", "critica", "conceptoRecaudo",
                   f"conceptoRecaudo '{cr}' no válido para otrosServicios (01,02,03,05).", cr)

    return errores


# ══════════════════════════════════════════════════════════════
# GENERACIÓN DE EXCEL DE REPORTE
# ══════════════════════════════════════════════════════════════
 
def construir_excel(registros, alertas=None, validaciones_malla=None,
                    validaciones_general=None, validaciones_auditoria=None,
                    validaciones_pertinencia=None, validaciones_malla_0948=None):
    wb  = Workbook()
 
    # ── Hoja 1: Medicamentos inválidos ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "Med_Invalidos"
    headers1  = ["Archivo", "Número Factura", "ID RIPS", "Código", "Nombre Tecnología", "Valor Servicio"]
    ws1.append(headers1)
 
    header_fill = PatternFill("solid", fgColor="0B3B76")
    for col in range(1, len(headers1) + 1):
        c = ws1.cell(row=1, column=col)
        c.font      = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.fill      = header_fill
 
    for r in registros:
        vr = r.get("vrServicio", "")
        try:
            vr_num = float(vr) if str(vr).strip() != "" else None
        except Exception:
            vr_num = None
        ws1.append([
            r.get("archivo", ""),
            r.get("numeroFactura", ""),
            r.get("idrips", ""),
            r.get("codConsulta", ""),
            r.get("nomTecnologiaSalud", ""),
            vr_num
        ])
 
    for row in range(2, ws1.max_row + 1):
        c = ws1.cell(row=row, column=6)
        if c.value is not None:
            c.number_format = "#,##0"
 
    for i in range(1, len(headers1) + 1):
        max_len = max(
            (len(str(ws1.cell(r, i).value or "")) for r in range(1, ws1.max_row + 1)),
            default=10
        )
        ws1.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 60)
 
    # ── Hoja 2: Alertas de autorizaciones ────────────────────────────────
    if alertas:
        ws2 = wb.create_sheet("Alertas_Autorizaciones")
        headers2 = ["Tipo de Alerta", "Detalle"]
        ws2.append(headers2)
        for col in range(1, 3):
            c = ws2.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill      = header_fill
 
        headers2 = ["Tipo de Alerta", "N° Autorización", "N° Doc Afiliado",
                    "Factura RIPS", "Archivo RIPS", "Archivo EPS"]
        ws2.delete_rows(1)
        ws2.append(headers2)
        for col in range(1, len(headers2) + 1):
            c = ws2.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill      = header_fill
 
        for item in alertas.get('tipo_doc_mismatch', []):
            ws2.append([
                "Tipo de documento no coincide con la EPS",
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', '')
            ])
 
        for item in alertas.get('amb_par_no_cruza', []):
            ws2.append([
                item.get("mensaje", "Ambulatorio - Par auth/cod no cruza"),
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])
 
        for item in alertas.get('hosp_proc_no_cruza', []):
            ws2.append([
                item.get("mensaje", "Hospitalario - proc no cruza en rango fechas"),
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])
 
        for item in alertas.get('hosp_aut_no_relacionada', []):
            ws2.append([
                item.get("mensaje", "Hospitalario - auth no relacionada en RIPS"),
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        # ── Nuevas alertas de urgencias ───────────────────────────────────────
        for item in alertas.get('urgencias_sin_aut', []):
            ws2.append([
                "Paciente de urgencias sin autorización asociada en el RIPS.",
                '',
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('procedimiento_sin_aut', []):
            ws2.append([
                item.get('mensaje', ''),
                item.get('cod_proc', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        # ── Nuevas alertas de hospitalización ────────────────────────────────
        for item in alertas.get('hosp_dias_excedidos', []):
            ws2.append([
                f"Días de estancia facturables ({item.get('dias_facturables','')}) superan "
                f"los días autorizados ({item.get('dias_autorizados','')}).",
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('hosp_aut_fuera_plazo', []):
            ws2.append([
                f"Fecha de emisión de autorización ({item.get('fecha_emision','')}) supera "
                f"24h después del egreso ({item.get('fecha_egreso','')}). "
                f"Diferencia: {item.get('horas_diff','')}h.",
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('estancia_sin_aut', []):
            ws2.append([
                item.get('mensaje', 'Estancia hospitalaria sin autorización en base de datos'),
                item.get('auths_hosp', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('proc_qx_misma_aut_hosp', []):
            ws2.append([
                item.get('mensaje', 'Proc. quirúrgico con auth igual a la de la estancia'),
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('sin_num_aut_relacionado', []):
            ws2.append([
                item.get('mensaje', 'Sin números de autorización relacionados'),
                '',
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('amb_aut_emision_posterior', []):
            ws2.append([
                f"Auth emitida POSTERIOR al servicio (Serv: {item.get('fecha_servicio','')} / Emis: {item.get('fecha_emision','')})",
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                '',
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('hosp_cod_sin_aut', []):
            ws2.append([
                item.get('mensaje', 'Hospitalización: código sin autorización asociada'),
                '',
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('hosp_proc_cod_no_cruza', []):
            ws2.append([
                item.get('mensaje', 'Hosp + Procedimientos: auth y código no coinciden'),
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('hosp_cups_duplicado', []):
            ws2.append([
                f"{item.get('mensaje', 'Hosp: código CUPS duplicado')} (x{item.get('repeticiones', '')})",
                item.get('cod_proc', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                '',
            ])

        for item in alertas.get('proc_sin_aut_amb', []):
            ws2.append([
                f"Proc. ambulatorio sin auth obligatoria ({item.get('seccion', '')})",
                item.get('cod_proc', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                '',
            ])

        for item in alertas.get('proc_aut_no_cruza_amb', []):
            ws2.append([
                'Auth no corresponde al CUPS del RIPS (Ambulatorio)',
                item.get('num_aut', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                item.get('archivo_eps', ''),
            ])

        for item in alertas.get('cups_noestandar_sin_aut', []):
            ws2.append([
                f"CUPS no estándar sin auth ({item.get('seccion', '')})",
                item.get('cod_proc', ''),
                item.get('num_doc', ''),
                item.get('num_factura', ''),
                item.get('archivo_rips', ''),
                '',
            ])

        for i in range(1, len(headers2) + 1):
            max_len = max(
                (len(str(ws2.cell(r, i).value or "")) for r in range(1, ws2.max_row + 1)),
                default=10
            )
            ws2.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 60)

    # ── Hoja 3: Malla 2275/2023 – Medicamentos ───────────────────────────────
    if validaciones_malla:
        ws3 = wb.create_sheet("Malla_Medicamentos")
        headers3 = [
            "Archivo", "Factura", "N° Doc Paciente", "Consecutivo Med",
            "ID Regla", "Severidad", "Campo", "Mensaje", "Valor Actual"
        ]
        ws3.append(headers3)

        fill_critica = PatternFill("solid", fgColor="C00000")
        fill_alta    = PatternFill("solid", fgColor="E26B0A")
        fill_media   = PatternFill("solid", fgColor="F0AD00")
        fill_head    = PatternFill("solid", fgColor="0B3B76")

        for col in range(1, len(headers3) + 1):
            c = ws3.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill      = fill_head

        for v in validaciones_malla:
            fila = [
                v.get("archivo",      ""),
                v.get("num_factura",  ""),
                v.get("num_doc",      ""),
                v.get("consecutivo",  ""),
                v.get("id_regla",     ""),
                v.get("severidad",    ""),
                v.get("campo",        ""),
                v.get("mensaje",      ""),
                v.get("valor_actual", ""),
            ]
            ws3.append(fila)
            sev = v.get("severidad", "")
            row_fill = (fill_critica if sev == "critica"
                        else fill_alta  if sev == "alta"
                        else fill_media)
            for col in range(1, len(headers3) + 1):
                cell = ws3.cell(row=ws3.max_row, column=col)
                if sev in {"critica", "alta", "media"}:
                    cell.fill = row_fill
                    cell.font = Font(color="FFFFFF" if sev == "critica" else "000000")

        for i in range(1, len(headers3) + 1):
            max_len = max(
                (len(str(ws3.cell(r, i).value or "")) for r in range(1, ws3.max_row + 1)),
                default=10
            )
            ws3.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 80)

    # ── Hoja 4: Malla 2275/2023 – General (todos los demás bloques) ─────────
    if validaciones_general:
        ws4 = wb.create_sheet("Malla_General")
        headers4 = [
            "Archivo", "Factura", "N° Doc Paciente", "Consecutivo",
            "ID Regla", "Severidad", "Bloque", "Campo", "Mensaje", "Valor Actual"
        ]
        ws4.append(headers4)

        fill_critica = PatternFill("solid", fgColor="C00000")
        fill_alta    = PatternFill("solid", fgColor="E26B0A")
        fill_media   = PatternFill("solid", fgColor="F0AD00")
        fill_head4   = PatternFill("solid", fgColor="0B3B76")

        for col in range(1, len(headers4) + 1):
            c = ws4.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill      = fill_head4

        for v in validaciones_general:
            # Deducir bloque a partir del prefijo del id_regla
            id_r   = v.get("id_regla", "")
            bloque = ("General/T"  if any(id_r.startswith(p) for p in ("RVG","T0","T01","T02","T03","T04","RVC001","RVC002","RVC003"))
                      else "U"     if any(id_r.startswith(p) for p in ("U0","U1","RVC006","RVC007","RVC008","RVC009","SEX-IVE","SEX-DX"))
                      else "C"     if any(id_r.startswith(p) for p in ("C0","C1","C2","RVC011","RVC013","RVC031","RVC060","RVC061","RVC079","RVC086","RVC087"))
                      else "P"     if any(id_r.startswith(p) for p in ("P0","P1","P2","RVC011-P","RVC013-P","RVC031-P","RVC079-P"))
                      else "R"     if any(id_r.startswith(p) for p in ("R0","R1","RVC038","RVC040","RVC042","RVC043"))
                      else "H"     if any(id_r.startswith(p) for p in ("H0","H1","RVC011-H","RVC013-H","RVC041","RVC042-H"))
                      else "N"     if any(id_r.startswith(p) for p in ("N0","N1","RVC045","RVC046","RVC057","RVC058","RVC042-N","RVC011-N"))
                      else "S"     if any(id_r.startswith(p) for p in ("S0","S-","RVC050","RVC013-S"))
                      else "")
            fila = [
                v.get("archivo",      ""),
                v.get("num_factura",  ""),
                v.get("num_doc",      ""),
                v.get("consecutivo",  ""),
                id_r,
                v.get("severidad",    ""),
                bloque,
                v.get("campo",        ""),
                v.get("mensaje",      ""),
                v.get("valor_actual", ""),
            ]
            ws4.append(fila)
            sev      = v.get("severidad", "")
            row_fill = (fill_critica if sev == "critica"
                        else fill_alta  if sev == "alta"
                        else fill_media if sev == "media"
                        else None)
            for col in range(1, len(headers4) + 1):
                cell = ws4.cell(row=ws4.max_row, column=col)
                if row_fill:
                    cell.fill = row_fill
                    cell.font = Font(color="FFFFFF" if sev == "critica" else "000000")

        for i in range(1, len(headers4) + 1):
            max_len = max(
                (len(str(ws4.cell(r, i).value or "")) for r in range(1, ws4.max_row + 1)),
                default=10
            )
            ws4.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 80)

    # ── Hoja: Malla 0948/2026 (nuevas reglas en transición) ──────────────
    if validaciones_malla_0948:
        ws0948 = wb.create_sheet("Malla_0948_2026")
        headers0948 = [
            "Archivo", "Factura", "N° Doc Paciente", "Consecutivo",
            "ID Regla", "Severidad", "Campo", "Mensaje", "Valor Actual"
        ]
        ws0948.append(headers0948)

        fill_critica = PatternFill("solid", fgColor="C00000")
        fill_alta    = PatternFill("solid", fgColor="E26B0A")
        fill_media   = PatternFill("solid", fgColor="F0AD00")
        fill_baja    = PatternFill("solid", fgColor="FFF2CC")
        fill_head0948 = PatternFill("solid", fgColor="6A1B9A")

        for col in range(1, len(headers0948) + 1):
            c = ws0948.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill      = fill_head0948

        for v in validaciones_malla_0948:
            fila = [
                v.get("archivo",      ""),
                v.get("num_factura",  ""),
                v.get("num_doc",      ""),
                v.get("consecutivo",  ""),
                v.get("id_regla",     ""),
                v.get("severidad",    ""),
                v.get("campo",        ""),
                v.get("mensaje",      ""),
                v.get("valor_actual", ""),
            ]
            ws0948.append(fila)
            sev = v.get("severidad", "")
            row_fill = (fill_critica if sev == "critica"
                        else fill_alta  if sev == "alta"
                        else fill_media if sev == "media"
                        else fill_baja  if sev == "baja"
                        else None)
            for col in range(1, len(headers0948) + 1):
                cell = ws0948.cell(row=ws0948.max_row, column=col)
                if row_fill:
                    cell.fill = row_fill
                    cell.font = Font(color="FFFFFF" if sev == "critica" else "000000")

        for i in range(1, len(headers0948) + 1):
            max_len = max(
                (len(str(ws0948.cell(r, i).value or "")) for r in range(1, ws0948.max_row + 1)),
                default=10
            )
            ws0948.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 80)

    # ── Hoja 5: Auditoría clínica ────────────────────────────────────────
    if validaciones_auditoria:
        ws5 = wb.create_sheet("Auditoria_Clinica")
        headers5 = [
            "Archivo", "Factura", "N° Doc Paciente",
            "ID Regla", "Severidad", "Campo", "Mensaje", "Valor Actual"
        ]
        ws5.append(headers5)

        fill_critica5 = PatternFill("solid", fgColor="C00000")
        fill_alta5    = PatternFill("solid", fgColor="E26B0A")
        fill_media5   = PatternFill("solid", fgColor="F0AD00")
        fill_head5    = PatternFill("solid", fgColor="1F4E79")

        for col in range(1, len(headers5) + 1):
            c = ws5.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill      = fill_head5

        for v in validaciones_auditoria:
            sev  = v.get("severidad", "")
            fila = [
                v.get("archivo",      ""),
                v.get("num_factura",  ""),
                v.get("num_doc",      ""),
                v.get("id_regla",     ""),
                sev,
                v.get("campo",        ""),
                v.get("mensaje",      ""),
                v.get("valor_actual", ""),
            ]
            ws5.append(fila)
            row_fill = (fill_critica5 if sev == "critica"
                        else fill_alta5  if sev == "alta"
                        else fill_media5 if sev == "media"
                        else None)
            if row_fill:
                for col in range(1, len(headers5) + 1):
                    cell = ws5.cell(row=ws5.max_row, column=col)
                    cell.fill = row_fill
                    cell.font = Font(color="FFFFFF" if sev == "critica" else "000000")

        for i in range(1, len(headers5) + 1):
            max_len = max(
                (len(str(ws5.cell(r, i).value or "")) for r in range(1, ws5.max_row + 1)),
                default=10
            )
            ws5.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 80)

    # ── Hoja 6: Pertinencia clínica — amarillo ───────────────────────────
    if validaciones_pertinencia:
        ws6 = wb.create_sheet("Pertinencia_Clinica")
        headers6 = [
            "Archivo", "Factura", "N° Doc Paciente",
            "ID Regla", "Grupo Diagnóstico", "Diagnósticos CIE-10",
            "Observación", "CUPS Esperados"
        ]
        ws6.append(headers6)

        fill_head6 = PatternFill("solid", fgColor="7B5E00")
        fill_pert  = PatternFill("solid", fgColor="FFF2CC")

        for col in range(1, len(headers6) + 1):
            c = ws6.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill      = fill_head6

        for v in validaciones_pertinencia:
            # Separar grupo y cups_esperados del mensaje
            # mensaje = "Grupo: diag [cie_desc] sin procedimientos. Se esperan: cups_desc."
            msg = v.get("mensaje", "")
            cups_esperados = ""
            grupo = ""
            if "Se esperan:" in msg:
                partes = msg.split("Se esperan:")
                grupo  = partes[0].strip().rstrip(".")
                cups_esperados = partes[1].strip().rstrip(".")
            else:
                grupo = msg

            fila = [
                v.get("archivo",      ""),
                v.get("num_factura",  ""),
                v.get("num_doc",      ""),
                v.get("id_regla",     ""),
                grupo,
                v.get("valor_actual", ""),
                v.get("mensaje",      ""),
                cups_esperados,
            ]
            ws6.append(fila)
            for col in range(1, len(headers6) + 1):
                cell = ws6.cell(row=ws6.max_row, column=col)
                cell.fill = fill_pert
                cell.font = Font(color="7B5E00", bold=(col == 4))

        for i in range(1, len(headers6) + 1):
            max_len = max(
                (len(str(ws6.cell(r, i).value or "")) for r in range(1, ws6.max_row + 1)),
                default=10
            )
            ws6.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 80)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
 
 
# ══════════════════════════════════════════════════════════════
# CACHÉ DEL ÚLTIMO RESULTADO (permite exportar sin re-procesar)
# ══════════════════════════════════════════════════════════════
_ultimo_resultado = {}


# ══════════════════════════════════════════════════════════════
# RUTAS FLASK
# ══════════════════════════════════════════════════════════════
 
@app.route('/', methods=['GET', 'POST'])
def index():
    global _ultimo_resultado
    registros               = []
    alertas                 = None
    validaciones_auditoria  = []
    error                   = None
    stats                   = {}
 
    if request.method == 'POST':
        action          = request.form.get("action", "view")
        archivos_json   = request.files.getlist('json_files')
        archivos_excel  = request.files.getlist('excel_files')

        # ── Exportar desde caché ─────────────────────────────────────────
        # El frontend envía action=excel sin archivos (el input Excel se
        # pierde tras el DOM-update del AJAX). El servidor usa el último
        # resultado procesado para construir el Excel.
        if action == "excel":
            if _ultimo_resultado.get('stats'):
                output = construir_excel(
                    _ultimo_resultado['registros'],
                    _ultimo_resultado['alertas'],
                    _ultimo_resultado['validaciones_malla'],
                    _ultimo_resultado['validaciones_general'],
                    _ultimo_resultado.get('validaciones_auditoria'),
                )
                nombre = f"Alertas_Malla_Validadora_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=nombre,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                error = "No hay resultados procesados. Primero procese los archivos RIPS."
                return render_template('index.html', registros=None, alertas=None,
                                       error=error, stats={})

        if not archivos_json or all(a.filename == "" for a in archivos_json):
            error = "Por favor seleccione uno o más archivos JSON (RIPS)."
            return render_template('index.html', registros=None, alertas=None,
                                   error=error, stats={})
 
        auts_rips_global   = {}
        registros_excel    = {}
        set_aut_excel      = set()
        errores_acum       = []
        validaciones_malla    = []
        validaciones_general  = []
        validaciones_auditoria = []

        # ── Cargar Excel de autorizaciones (opcional) ─────────────────────
        hay_excel = archivos_excel and any(a.filename != "" for a in archivos_excel)
        if hay_excel:
            registros_excel, set_aut_excel, errores_excel = cargar_excel_autorizaciones(archivos_excel)
            errores_acum.extend(errores_excel)

        # ── Procesar JSONs RIPS ───────────────────────────────────────────
        archivos_procesados  = 0
        total_rips           = 0
        pacientes_rips_global = {}   # { num_doc → datos del paciente }
        for archivo in archivos_json:
            if not archivo or archivo.filename == "":
                continue
            try:
                data = json.load(archivo)
                regs = extraer_medicamentos_invalidos(data, archivo.filename)
                registros.extend(regs)
                pacientes = extraer_autorizaciones_rips(data, archivo.filename)
                pacientes_rips_global.update(pacientes)
                total_rips += contar_registros_rips(data)
                validaciones_malla.extend(
                    validar_medicamentos_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_general_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_usuarios_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_consultas_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_procedimientos_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_urgencias_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_hospitalizacion_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_recien_nacidos_malla_2275(data, archivo.filename)
                )
                validaciones_general.extend(
                    validar_otros_servicios_malla_2275(data, archivo.filename)
                )
                validaciones_auditoria.extend(
                    validar_auditoria(data, archivo.filename)
                )
                archivos_procesados += 1
            except Exception as e:
                errores_acum.append(f"Error en {archivo.filename}: {e}")

        # ── Ejecutar validaciones de autorizaciones ───────────────────────
        if hay_excel and (registros_excel or set_aut_excel):
            alertas = validar_autorizaciones(pacientes_rips_global, registros_excel, set_aut_excel)

        # Estadísticas resumen
        total_auths_rips = sum(len(p.get('set_auths', set())) for p in pacientes_rips_global.values())

        def _top_reglas(lista, n=5):
            cnt = {}
            for v in lista:
                r = v.get('id_regla', '')
                cnt[r] = cnt.get(r, 0) + 1
            return sorted(cnt.items(), key=lambda x: -x[1])[:n]

        stats = {
            'archivos_json':        archivos_procesados,
            'total_rips':           total_rips,
            'alerta_volumen':       total_rips > 800,
            'med_invalidos':        len(registros),
            'auts_rips':            total_auths_rips,
            'auts_excel':           len(set_aut_excel),
            'tipo_mismatch':     len(alertas['tipo_doc_mismatch'])       if alertas else 0,
            'amb_par_nc':        len(alertas['amb_par_no_cruza'])         if alertas else 0,
            'hosp_proc_nc':      len(alertas['hosp_proc_no_cruza'])       if alertas else 0,
            'hosp_aut_no_rel':   len(alertas['hosp_aut_no_relacionada'])  if alertas else 0,
            'estancia_sin_aut':  len(alertas['estancia_sin_aut'])         if alertas else 0,
            'proc_qx_aut_hosp':  len(alertas['proc_qx_misma_aut_hosp'])  if alertas else 0,
            'sin_aut_rel':       len(alertas['sin_num_aut_relacionado'])  if alertas else 0,
            'malla_total':          len(validaciones_malla),
            'malla_criticas':       sum(1 for v in validaciones_malla if v.get('severidad') == 'critica'),
            'malla_notificaciones': sum(1 for v in validaciones_malla if v.get('severidad') in {'media', 'alta'}),
            'malla_top_reglas':     _top_reglas(validaciones_malla),
            'general_total':        len(validaciones_general),
            'general_criticas':     sum(1 for v in validaciones_general if v.get('severidad') == 'critica'),
            'general_notificaciones': sum(1 for v in validaciones_general if v.get('severidad') in {'media', 'alta'}),
            'general_top_reglas':   _top_reglas(validaciones_general),
            'auditoria_total':          len(validaciones_auditoria),
            'auditoria_criticas':       sum(1 for v in validaciones_auditoria if v.get('severidad') == 'critica'),
            'auditoria_notificaciones': sum(1 for v in validaciones_auditoria if v.get('severidad') in {'media', 'alta'}),
            'auditoria_top_reglas':     _top_reglas(validaciones_auditoria),
        }
 
        if errores_acum:
            error = " | ".join(errores_acum)

        # ── Guardar resultado en caché para exportar sin re-procesar ─────
        _ultimo_resultado = {
            'registros':              registros,
            'alertas':                alertas,
            'validaciones_malla':     validaciones_malla,
            'validaciones_general':   validaciones_general,
            'validaciones_auditoria': validaciones_auditoria,
            'stats':                  stats,
        }
 
    return render_template(
        'index.html',
        registros=registros if registros else None,
        alertas=alertas,
        validaciones_auditoria=validaciones_auditoria if validaciones_auditoria else None,
        error=error,
        stats=stats
    )
 
 
# ══════════════════════════════════════════════════════════════
# VALIDACIONES MALLA 0948/2026 – complementarias a 2275
# ══════════════════════════════════════════════════════════════
#
# Fuente: docs_normativos/anexo_tecnico1_resolucion_948_2026.txt
# ("Documento Técnico 1 – Especificaciones técnicas de los campos de datos
# y las reglas de validación del RIPS como soporte de la FEV en salud",
# anexo a la Resolución 000948 de 2026), transcripción limpia hecha por el
# usuario a partir del PDF oficial. Reemplaza por completo la versión
# anterior basada en el texto OCR degradado.
#
# Esta malla implementa, por cada bloque de servicios:
#   1) Presencia/formato de los CAMPOS NUEVOS (numeral 3.1.1 / numeral 4).
#   2) Chequeos INFORMATIVOS de tamaño para los campos existentes que
#      recortan su tamaño en 0948 (numeral 3.2.1 / numeral 4) — nunca se
#      escala a "critica" porque 2275 sigue vigente en paralelo.
#   3) Las reglas nuevas concretas RVC094-RVC098 (numeral 3.1.2 / numeral 6),
#      implementadas literalmente cuando el dato es derivable del RIPS.
#   4) Las escaladas de Notificación→Rechazo listadas en 3.2.2/3.2.3: si el
#      chequeo homólogo ya existe en *_malla_2275, se reporta una versión
#      0948 con severidad más alta; si no existe, se implementa de cero acá
#      con severidad informativa ("media"/"baja").
#   5) M02 (numAutorizacion en medicamentos) NO se exige en 0948 porque el
#      campo fue eliminado (numeral 3.3); validar_medicamentos_malla_2275
#      no se modifica y sigue exigiéndolo bajo 2275.
#
# RVC/RVG ids del numeral 6 que dependían de catálogos de referencia
# oficiales (Cups2026enjson.json y las tablas en "tablas de referencia/")
# — ya implementados con esos catálogos:
#   - RVC096 (C04,P05,S06 contra tabla "CUPSRips"): se carga
#     Cups2026enjson.json de forma perezosa (_cups_codigos_validos_0948) y
#     se valida codConsulta/codProcedimiento/codTecnologiaSalud (este
#     último solo para tipoOS de traslados/estancias/honorarios). RVC023
#     (P05, escalada N→R) se mantiene aparte por ser un chequeo distinto
#     (CUPS=parto ⇒ debe existir recienNacidos), ver validar_procedimientos_malla_0948.
#   - RVC098 (fecha de servicio vs. fecha de fallecimiento + 24h): se usa
#     fechaEgreso del registro (urgencias/hospitalizacion/recienNacidos)
#     con condicionDestinoUsuarioEgreso='02' (PACIENTE MUERTO) como proxy
#     de fecha de fallecimiento (decisión explícita del usuario, ya que el
#     RIPS no trae un campo dedicado y RUAF-ND no está disponible en este
#     pipeline). Implementado en validar_usuarios_malla_0948.
#   - RVC084 (C08 finalidadTecnologiaSalud ⇒ C18 conceptoRecaudo="05"): se
#     usa el listado de finalidades de promoción/prevención/materno-
#     perinatal tomado de TablaReferencia_RIPSFinalidadConsultaVersion2
#     (códigos 11,12,14,19,20,21,22,23,24,25,27). Implementado en
#     validar_consultas_malla_0948.
#   - RVG14/RVG15 (nacimientos múltiples ⇔ procedimiento de parto múltiple;
#     urgencias con observación ⇔ consulta de urgencia asociada):
#     implementados en validar_usuarios_malla_0948 usando los códigos CUPS
#     de parto múltiple (735930/735931) y de consulta de urgencias
#     (890701-890793) obtenidos del catálogo CUPS oficial.
#
# ── Cronograma de exigibilidad (Resolución 000948 de 2026, art. 23 y
#    parágrafo transitorio) ──────────────────────────────────────────
#   - Vigencia general: desde la fecha de expedición (14-may-2026); deroga
#     2275/2023, 558/2024 y 1884/2024.
#   - Desde el 01-jun-2026: las reglas que hoy son de "notificación" pasan a
#     ser de "rechazo" (esto es lo que _severidad_0948 aproxima).
#   - Desde el 01-jul-2026: son exigibles los ajustes de versión de software
#     de carácter estructural (i.e. los campos nuevos dejan de ser opcionales
#     "de transición").
#   - RVC095 (U12/registroSIRAS): Notificación durante los primeros 3 meses
#     desde la implementación de la resolución (14-may-2026 + 3 meses =
#     ~14-ago-2026), luego Rechazo.
# Estas fechas se usan para escalar automáticamente la severidad de los
# hallazgos 0948 con el paso del tiempo, en vez de dejarlos fijos.

from datetime import date as _date

# ── Catálogo oficial CUPS (RVC096/RVC023) ─────────────────────────
# Cargado perezosamente (una sola vez) desde Cups2026enjson.json en la raíz
# del repo. Si el archivo no existe o no puede leerse, se cachea un
# conjunto vacío y los chequeos que dependen de él simplemente se omiten
# (no se rompe el resto de la validación por falta del catálogo).
_CUPS_CODIGOS_VALIDOS_0948 = None


def _cups_codigos_validos_0948():
    global _CUPS_CODIGOS_VALIDOS_0948
    if _CUPS_CODIGOS_VALIDOS_0948 is not None:
        return _CUPS_CODIGOS_VALIDOS_0948
    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Cups2026enjson.json")
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            data = json.load(f)
        codigos = {
            normalizar_str(d.get("Codigo"))
            for d in data
            if isinstance(d, dict) and d.get("Tabla") == "CUPS"
            and normalizar_str(d.get("Habilitado")).upper() == "SI"
        }
        _CUPS_CODIGOS_VALIDOS_0948 = codigos
    except Exception:
        _CUPS_CODIGOS_VALIDOS_0948 = set()
    return _CUPS_CODIGOS_VALIDOS_0948


# Códigos CUPS de "consulta de urgencias" (tabla CUPS, Nombre inicia con
# "CONSULTA DE URGENCIAS"), usados por RVG15.
CUPS_CONSULTA_URGENCIAS_0948 = {
    "890701", "890702", "890703", "890704", "890735",
    "890750", "890763", "890780", "890781", "890783", "890793",
}

# Códigos CUPS de "parto múltiple" (asistencia del parto espontáneo/
# intervenido gemelar o múltiple), usados por RVG14.
CUPS_PARTO_MULTIPLE_0948 = {"735930", "735931"}

# Finalidades (C08, tabla RIPSFinalidadConsultaVersion2) de promoción y
# mantenimiento / prevención / materno-perinatal que exigen
# conceptoRecaudo (C18) = "05" (No aplica pago moderador), según RVC084.
FINALIDADES_NO_APLICA_PAGO_MODERADOR_0948 = {
    "11", "12", "14", "19", "20", "21", "22", "23", "24", "25", "27",
}

# tipoOS (S05) para los que S06 debe corresponder a un código CUPS válido:
# 02=Traslados, 03=Estancias, 05=Honorarios. Se excluyen 01 (dispositivos
# médicos e insumos: código IDM/UDI o propio, no CUPS), 04 (servicios
# complementarios: código de tabla MIPRES, no CUPS) y 06 (servicios de
# salud a comunidades indígenas, cuya codificación no está documentada en
# el anexo técnico consultado, así que se omite en vez de adivinar).
TIPOS_OS_VALIDABLES_CONTRA_CUPS_0948 = {"02", "03", "05"}

FECHA_EXPEDICION_0948 = _date(2026, 5, 14)
FECHA_RECHAZO_0948 = _date(2026, 6, 1)
FECHA_ESTRUCTURAL_0948 = _date(2026, 7, 1)
FECHA_RVC095_RECHAZO = _date(2026, 8, 14)  # 3 meses desde la expedición


def _severidad_0948(base_media="media", base_baja="baja", hoy=None):
    """
    Severidad para un hallazgo 0948 según el cronograma del parágrafo
    transitorio del art. 23: antes del 01-jul-2026 se mantienen como
    informativas; desde esa fecha, las reglas de estructura (campos nuevos)
    se tratan como "alta" en vez de "media"/"baja". Nunca llega a "critica"
    para campos nuevos porque la resolución no define rechazo automático
    por su ausencia (sí lo define para las reglas de escalada N→R, que se
    manejan aparte en _severidad_escalada_0948).
    """
    hoy = hoy or _date.today()
    if hoy >= FECHA_ESTRUCTURAL_0948:
        return "alta"
    return base_media


def _severidad_escalada_0948(hoy=None):
    """
    Severidad para las reglas que el numeral 3.2.2/3.2.3 marca como
    "cambia de Notificación a Rechazo" en 0948. Antes del 01-jun-2026 se
    reportan como "alta" (más fuerte que una notificación normal, para
    anticipar el cambio); desde el 01-jun-2026 se reportan como "critica"
    porque pasan a ser causal de rechazo.
    """
    hoy = hoy or _date.today()
    return "critica" if hoy >= FECHA_RECHAZO_0948 else "alta"


def _campo_nuevo_0948(errores, ctx, campo, valor, min_len=None, max_len=None,
                       opcional_transicion=False, id_regla=None):
    """
    Chequeo genérico para un campo NUEVO introducido en la Resolución
    0948/2026 (numeral 3.1.1 / numeral 4).

    IMPORTANTE: la sola AUSENCIA de un campo nuevo NO se reporta como
    hallazgo. Los RIPS generados bajo 2275 legítimamente no traen estos
    campos (codigoVIDA, diagnósticos CIE11, etc.) y marcar cada registro
    como alerta produce miles de falsos positivos sin valor (un campo
    nuevo por definición no existe aún en datos históricos). Solo se
    valida el FORMATO cuando el campo sí viene informado.
    """
    id_regla = id_regla or f"0948-{campo}"
    hoy = _date.today()
    val = normalizar_str(valor)
    if not val or val.lower() in {"none", "null"}:
        return
    if min_len is not None and len(val) < min_len:
        errores.append({
            **ctx, "id_regla": id_regla,
            "severidad": _severidad_0948(base_media="baja", hoy=hoy),
            "campo": campo,
            "mensaje": f"[0948] Campo '{campo}' tiene longitud {len(val)}, "
                       f"menor al mínimo esperado ({min_len}) según 0948.",
            "valor_actual": val,
        })
    if max_len is not None and len(val) > max_len:
        errores.append({
            **ctx, "id_regla": id_regla,
            "severidad": _severidad_0948(base_media="baja", hoy=hoy),
            "campo": campo,
            "mensaje": f"[0948] Campo '{campo}' tiene longitud {len(val)}, "
                       f"mayor al máximo esperado ({max_len}) según 0948.",
            "valor_actual": val,
        })


def _tamano_cambio_0948(errores, ctx, campo, valor, tamano_exacto, id_regla, permite_null=True):
    """
    Chequeo INFORMATIVO (numeral 3.2.1) para campos existentes en 2275 cuyo
    tamaño se acorta a un valor exacto en 0948 (p.ej. C10 de "4-25" a "4";
    C11 de "0-25" a "0,4"). Como 2275 sigue vigente en paralelo, esto NUNCA
    se reporta como "critica": es una advertencia de que, bajo 0948, ese
    valor sería rechazado por tamaño.
    """
    val = normalizar_str(valor)
    if not val or val.lower() in {"none", "null"}:
        if not permite_null:
            errores.append({
                **ctx, "id_regla": id_regla, "severidad": "baja",
                "campo": campo,
                "mensaje": f"[0948] Campo '{campo}' no informado; en 0948 su tamaño exacto "
                           f"exigido es {tamano_exacto} (no admite null).",
                "valor_actual": "",
            })
        return
    if len(val) != tamano_exacto:
        errores.append({
            **ctx, "id_regla": id_regla, "severidad": "baja",
            "campo": campo,
            "mensaje": f"[0948] Campo '{campo}' tiene tamaño {len(val)}; bajo la Resolución 0948 "
                       f"el tamaño exacto exigido es {tamano_exacto} (2275 admite un rango mayor; "
                       "este hallazgo es informativo de cara a la migración).",
            "valor_actual": val,
        })


def _escalada_0948(errores, ctx, id_regla, campo, mensaje, valor_actual=""):
    """
    Reporta un hallazgo para una regla que el numeral 3.2.2/3.2.3 del texto
    fuente marca explícitamente como "cambia de Notificación a Rechazo" (o
    ajuste equivalente) en 0948. Severidad creciente según _severidad_escalada_0948.
    """
    errores.append({
        **ctx, "id_regla": id_regla,
        "severidad": _severidad_escalada_0948(),
        "campo": campo,
        "mensaje": f"[0948-escalada N→R] {mensaje}",
        "valor_actual": normalizar_str(valor_actual),
    })


def _ctx_base_0948(nombre_archivo, data):
    FACTURA_KEYS = {"numFactura", "numeroFactura", "nroFactura", "noFactura", "factura"}
    num_factura = ""
    for k in FACTURA_KEYS:
        v = data.get(k)
        if v and not isinstance(v, (dict, list)):
            num_factura = normalizar_str(v)
            break
    return num_factura


def _parse_fecha_0948(valor):
    """Parsea una fecha RIPS ('AAAA-MM-DD HH:MM[:SS]' o 'AAAA-MM-DD') a
    datetime, siguiendo el mismo patrón usado en el resto de la malla 0948
    (ver validar_urgencias_malla_0948/validar_hospitalizacion_malla_0948).
    Devuelve None si no se puede parsear."""
    raw = normalizar_str(valor or "")
    if not raw:
        return None
    try:
        if len(raw) >= 16:
            return datetime.strptime(raw[:16], "%Y-%m-%d %H:%M")
        elif len(raw) == 10:
            return datetime.strptime(raw, "%Y-%m-%d")
    except ValueError:
        return None
    return None


def _condicion_indica_muerto_0948(cond_egr):
    """CondicionyDestinoUsuarioEgreso: '02' = PACIENTE MUERTO (tabla oficial
    TablaReferencia_CondicionyDestinoUsuarioEgreso). Antes se comparaba
    erróneamente contra '21'; corregido según la tabla de referencia oficial."""
    return normalizar_str(cond_egr) == "02"


def _condicion_indica_derivado_0948(cond_egr):
    """Código '03' = PACIENTE DERIVADO A OTRO SERVICIO (tabla oficial
    CondicionyDestinoUsuarioEgreso). Se excluye '04' (REFERIDO A OTRA
    INSTITUCION), que es un concepto distinto no referenciado por
    RVC053/RVC062 en el anexo técnico."""
    return normalizar_str(cond_egr) in {"03"}


# ────────────────────────────────────────────────────────────────
# BLOQUE C: CONSULTAS
# ────────────────────────────────────────────────────────────────

def validar_consultas_malla_0948(data, nombre_archivo=""):
    """
    Consultas bajo 0948:
    - Campos nuevos C22-C30 (codigoVIDA + diagnósticos CIE11).
    - Tamaño exacto de C10 (4) y C11-C13 (0,4) — informativo.
    - RVC086/RVC087 escaladas a Rechazo (diagnóstico relacionado repetido /
      igual al principal).
    - RVC096 (C04 codConsulta contra catálogo oficial CUPS).
    - RVC084 (C08 finalidadTecnologiaSalud ⇒ C18 conceptoRecaudo='05').
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        consultas = servicios.get("consultas", [])
        if not isinstance(consultas, list):
            continue
        for con in consultas:
            if not isinstance(con, dict):
                continue
            ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
                   "num_doc": num_doc, "consecutivo": normalizar_str(con.get("consecutivo"))}

            # ── Campos nuevos ──────────────────────────────
            _campo_nuevo_0948(errores, ctx, "codigoVIDA", con.get("codigoVIDA"),
                               1, 256, opcional_transicion=True, id_regla="C22-0948")
            _campo_nuevo_0948(errores, ctx, "codDiagnosticoPrincipalCIE11",
                               con.get("codDiagnosticoPrincipalCIE11"), 4, 256, id_regla="C23-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoPrincipalCIE11",
                               con.get("nomCodDiagnosticoPrincipalCIE11"), 4, 1000, id_regla="C24-0948")
            for n, letra in ((1, "C25"), (2, "C27"), (3, "C29")):
                _campo_nuevo_0948(errores, ctx, f"codDiagnosticoRelacionado{n}CIE11",
                                   con.get(f"codDiagnosticoRelacionado{n}CIE11"),
                                   opcional_transicion=True, id_regla=f"{letra}-0948")
            for n, letra in ((1, "C26"), (2, "C28"), (3, "C30")):
                _campo_nuevo_0948(errores, ctx, f"nomCodDiagnosticoRelacionado{n}CIE11",
                                   con.get(f"nomCodDiagnosticoRelacionado{n}CIE11"),
                                   opcional_transicion=True, id_regla=f"{letra}-0948")

            # ── Tamaño (3.2.1): C10 exacto 4, C11-C13 "0,4" ──
            _tamano_cambio_0948(errores, ctx, "codDiagnosticoPrincipal",
                                 con.get("codDiagnosticoPrincipal"), 4, "C10-0948", permite_null=False)
            for campo in ("codDiagnosticoRelacionado1", "codDiagnosticoRelacionado2",
                          "codDiagnosticoRelacionado3"):
                _tamano_cambio_0948(errores, ctx, campo, con.get(campo), 4, "C11C12C13-0948")

            # ── RVC086/RVC087 escaladas a Rechazo ────────────
            diag_p = normalizar_str(con.get("codDiagnosticoPrincipal") or "")
            diags_rel = [normalizar_str(con.get(f"codDiagnosticoRelacionado{n}") or "") for n in (1, 2, 3)]
            diags_notnull = [d for d in diags_rel if d]
            for dr in diags_notnull:
                if dr == diag_p:
                    _escalada_0948(errores, ctx, "RVC086", "codDiagnosticoRelacionado",
                                    f"Diagnóstico relacionado '{dr}' es igual al diagnóstico principal "
                                    "(RVC086 pasa de Notificación a Rechazo en 0948).", dr)
            if len(diags_notnull) != len(set(diags_notnull)):
                _escalada_0948(errores, ctx, "RVC087", "codDiagnosticoRelacionado",
                                "Existen diagnósticos relacionados repetidos entre sí "
                                "(RVC087 pasa de Notificación a Rechazo en 0948).")

            # RVC096 (nueva 0948): codConsulta debe existir en el catálogo CUPS.
            cups_validos = _cups_codigos_validos_0948()
            cod_consulta = normalizar_str(con.get("codConsulta") or "")
            if cups_validos and cod_consulta and cod_consulta not in cups_validos:
                _escalada_0948(errores, ctx, "RVC096", "codConsulta",
                                f"codConsulta '{cod_consulta}' no existe en el catálogo oficial CUPS "
                                "(RVC096, nueva en 0948).", cod_consulta)

            # RVC084 (nueva/escalada 0948): finalidad de promoción/prevención/
            # materno-perinatal ⇒ conceptoRecaudo debe ser "05" (No aplica pago moderador).
            finalidad = normalizar_str(con.get("finalidadTecnologiaSalud") or "")
            concepto_recaudo = normalizar_str(con.get("conceptoRecaudo") or "")
            if finalidad in FINALIDADES_NO_APLICA_PAGO_MODERADOR_0948 and concepto_recaudo != "05":
                _escalada_0948(errores, ctx, "RVC084", "conceptoRecaudo",
                                f"finalidadTecnologiaSalud='{finalidad}' exige conceptoRecaudo='05' "
                                f"(No aplica pago moderador); valor actual='{concepto_recaudo}' "
                                "(RVC084 pasa de Notificación a Rechazo en 0948).", concepto_recaudo)
    return errores


# ────────────────────────────────────────────────────────────────
# BLOQUE P: PROCEDIMIENTOS
# ────────────────────────────────────────────────────────────────

def validar_procedimientos_malla_0948(data, nombre_archivo=""):
    """
    Procedimientos bajo 0948:
    - Campos nuevos P21-P27 (CIE11 + codigoVIDA).
    - Tamaño "0,4" para P14 (codDiagnosticoRelacionado) y P15 (codComplicacion) — informativo.
    - RVC086 escalada (diagnóstico relacionado == principal), aplicable a P14.
    - RVC096 (P05 codProcedimiento contra catálogo oficial CUPS).
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        procs = servicios.get("procedimientos", [])
        if not isinstance(procs, list):
            continue
        for p in procs:
            if not isinstance(p, dict):
                continue
            ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
                   "num_doc": num_doc, "consecutivo": normalizar_str(p.get("consecutivo"))}

            _campo_nuevo_0948(errores, ctx, "codDiagnosticoPrincipalCIE11",
                               p.get("codDiagnosticoPrincipalCIE11"), 4, 256, id_regla="P21-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoPrincipalCIE11",
                               p.get("nomCodDiagnosticoPrincipalCIE11"), 1, 1000, id_regla="P22-0948")
            _campo_nuevo_0948(errores, ctx, "codDiagnosticoRelacionadoCIE11",
                               p.get("codDiagnosticoRelacionadoCIE11"),
                               opcional_transicion=True, id_regla="P23-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoRelacionadoCIE11",
                               p.get("nomCodDiagnosticoRelacionadoCIE11"),
                               opcional_transicion=True, id_regla="P24-0948")
            _campo_nuevo_0948(errores, ctx, "codComplicacionCIE11",
                               p.get("codComplicacionCIE11"), opcional_transicion=True, id_regla="P25-0948")
            _campo_nuevo_0948(errores, ctx, "nomComplicacionCIE11",
                               p.get("nomComplicacionCIE11"), opcional_transicion=True, id_regla="P26-0948")
            _campo_nuevo_0948(errores, ctx, "codigoVIDA", p.get("codigoVIDA"),
                               1, 256, opcional_transicion=True, id_regla="P27-0948")

            _tamano_cambio_0948(errores, ctx, "codDiagnosticoRelacionado",
                                 p.get("codDiagnosticoRelacionado"), 4, "P14-0948")
            _tamano_cambio_0948(errores, ctx, "codComplicacion",
                                 p.get("codComplicacion"), 4, "P15-0948")

            diag_p = normalizar_str(p.get("codDiagnosticoPrincipal") or "")
            diag_rel = normalizar_str(p.get("codDiagnosticoRelacionado") or "")
            if diag_rel and diag_rel == diag_p:
                _escalada_0948(errores, ctx, "RVC086", "codDiagnosticoRelacionado",
                                f"codDiagnosticoRelacionado '{diag_rel}' es igual al diagnóstico principal "
                                "(RVC086 pasa de Notificación a Rechazo en 0948).", diag_rel)

            # RVC096 (nueva 0948): codProcedimiento debe existir en el catálogo CUPS.
            cups_validos = _cups_codigos_validos_0948()
            cod_proc = normalizar_str(p.get("codProcedimiento") or "")
            if cups_validos and cod_proc and cod_proc not in cups_validos:
                _escalada_0948(errores, ctx, "RVC096", "codProcedimiento",
                                f"codProcedimiento '{cod_proc}' no existe en el catálogo oficial CUPS "
                                "(RVC096, nueva en 0948).", cod_proc)
    return errores


# ────────────────────────────────────────────────────────────────
# BLOQUE R: URGENCIAS
# ────────────────────────────────────────────────────────────────

def validar_urgencias_malla_0948(data, nombre_archivo=""):
    """
    Urgencias bajo 0948:
    - Campos nuevos R13-R25 (CIE11 + codigoVIDA).
    - Tamaño exacto R04/R05 (4), R06-R08/R10 (0,4) — informativo.
    - RVC088/RVC089 escaladas a Rechazo (relacionados de egreso == principal
      de egreso / repetidos entre sí).
    - RVC053/RVC062 escaladas a Rechazo (condición=muerto ⇒ sin servicios
      posteriores a 24h de fallecimiento; condición=derivado ⇒ deben existir
      los servicios de destino). Implementadas con lo disponible en el
      propio bloque de urgencias del RIPS (fechaEgreso como proxy de fecha
      de fallecimiento cuando condicionDestinoUsuarioEgreso='21').
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        urg = servicios.get("urgencias", [])
        if not isinstance(urg, list):
            continue

        # Fecha de fallecimiento derivada (proxy): fechaEgreso del primer
        # registro de urgencias con condicionDestinoUsuarioEgreso='21'.
        fecha_falle_dt = None
        for r in urg:
            if isinstance(r, dict) and _condicion_indica_muerto_0948(r.get("condicionDestinoUsuarioEgreso")):
                feg = normalizar_str(r.get("fechaEgreso") or "")
                try:
                    if len(feg) >= 16:
                        fecha_falle_dt = datetime.strptime(feg[:16], "%Y-%m-%d %H:%M")
                    elif len(feg) == 10:
                        fecha_falle_dt = datetime.strptime(feg, "%Y-%m-%d")
                except ValueError:
                    pass
                break

        for r in urg:
            if not isinstance(r, dict):
                continue
            ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
                   "num_doc": num_doc, "consecutivo": normalizar_str(r.get("consecutivo"))}

            _campo_nuevo_0948(errores, ctx, "codDiagnosticoPrincipalCIE11",
                               r.get("codDiagnosticoPrincipalCIE11"), 4, 256, id_regla="R13-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoPrincipalCIE11",
                               r.get("nomCodDiagnosticoPrincipalCIE11"), 1, 1000, id_regla="R14-0948")
            _campo_nuevo_0948(errores, ctx, "codDiagnosticoPrincipalECIE11",
                               r.get("codDiagnosticoPrincipalECIE11"), 4, 256, id_regla="R15-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoPrincipalECIE11",
                               r.get("nomCodDiagnosticoPrincipalECIE11"), 1, 1000, id_regla="R16-0948")
            for n, letra in ((1, "R17"), (2, "R19"), (3, "R21")):
                _campo_nuevo_0948(errores, ctx, f"codDiagnosticoRelacionadoE{n}CIE11",
                                   r.get(f"codDiagnosticoRelacionadoE{n}CIE11"),
                                   opcional_transicion=True, id_regla=f"{letra}-0948")
            for n, letra in ((1, "R18"), (2, "R20"), (3, "R22")):
                _campo_nuevo_0948(errores, ctx, f"nomCodDiagnosticoRelacionadoE{n}CIE11",
                                   r.get(f"nomCodDiagnosticoRelacionadoE{n}CIE11"),
                                   opcional_transicion=True, id_regla=f"{letra}-0948")
            _campo_nuevo_0948(errores, ctx, "codDiagnosticoCausaMuerteCIE11",
                               r.get("codDiagnosticoCausaMuerteCIE11"),
                               opcional_transicion=True, id_regla="R23-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoCausaMuerteCIE11",
                               r.get("nomCodDiagnosticoCausaMuerteCIE11"),
                               opcional_transicion=True, id_regla="R24-0948")
            _campo_nuevo_0948(errores, ctx, "codigoVIDA", r.get("codigoVIDA"),
                               1, 256, opcional_transicion=True, id_regla="R25-0948")

            _tamano_cambio_0948(errores, ctx, "codDiagnosticoPrincipal",
                                 r.get("codDiagnosticoPrincipal"), 4, "R04-0948", permite_null=False)
            _tamano_cambio_0948(errores, ctx, "codDiagnosticoPrincipalE",
                                 r.get("codDiagnosticoPrincipalE"), 4, "R05-0948", permite_null=False)
            for campo, id_r in (("codDiagnosticoRelacionadoE1", "R06-0948"),
                                 ("codDiagnosticoRelacionadoE2", "R07-0948"),
                                 ("codDiagnosticoRelacionadoE3", "R08-0948"),
                                 ("codDiagnosticoCausaMuerte", "R10-0948")):
                _tamano_cambio_0948(errores, ctx, campo, r.get(campo), 4, id_r)

            # RVC088/RVC089: relacionados de egreso == principal de egreso / repetidos
            diag_pe = normalizar_str(r.get("codDiagnosticoPrincipalE") or "")
            rel_e = [normalizar_str(r.get(f"codDiagnosticoRelacionadoE{n}") or "") for n in (1, 2, 3)]
            rel_e_notnull = [d for d in rel_e if d]
            for dr in rel_e_notnull:
                if dr == diag_pe:
                    _escalada_0948(errores, ctx, "RVC088", "codDiagnosticoRelacionadoE",
                                    f"Diagnóstico relacionado de egreso '{dr}' es igual al principal de "
                                    "egreso (RVC088 pasa de Notificación a Rechazo en 0948).", dr)
            if len(rel_e_notnull) != len(set(rel_e_notnull)):
                _escalada_0948(errores, ctx, "RVC089", "codDiagnosticoRelacionadoE",
                                "Existen diagnósticos relacionados de egreso repetidos entre sí "
                                "(RVC089 pasa de Notificación a Rechazo en 0948).")

            # RVC053: muerto ⇒ sin servicios posteriores a fecha_falle + 24h
            cond = r.get("condicionDestinoUsuarioEgreso")
            if fecha_falle_dt is not None and not _condicion_indica_muerto_0948(cond):
                fecha_serv_raw = normalizar_str(r.get("fechaInicioAtencion") or "")
                fecha_serv_dt = None
                try:
                    if len(fecha_serv_raw) >= 16:
                        fecha_serv_dt = datetime.strptime(fecha_serv_raw[:16], "%Y-%m-%d %H:%M")
                    elif len(fecha_serv_raw) == 10:
                        fecha_serv_dt = datetime.strptime(fecha_serv_raw, "%Y-%m-%d")
                except ValueError:
                    pass
                if fecha_serv_dt and (fecha_serv_dt - fecha_falle_dt).total_seconds() > 24 * 3600:
                    _escalada_0948(errores, ctx, "RVC053", "fechaInicioAtencion",
                                    "El usuario tiene un registro de urgencias marcado como fallecido y este "
                                    "otro servicio de urgencias tiene fecha posterior a 24h del fallecimiento "
                                    "(RVC053 pasa de Notificación a Rechazo en 0948).", fecha_serv_raw)

            # RVC062: derivado a otro servicio ⇒ deben existir esos servicios (best-effort:
            # se limita a dejar constancia informativa, ya que "esos servicios" pueden estar
            # en cualquiera de los otros bloques del mismo usuario y no hay forma fiable de
            # mapear el destino textual a un bloque concreto sin la tabla de referencia).
            if _condicion_indica_derivado_0948(cond):
                otros_bloques = servicios.get("hospitalizacion", []) or servicios.get("procedimientos", [])
                if not otros_bloques:
                    _escalada_0948(errores, ctx, "RVC062", "condicionDestinoUsuarioEgreso",
                                    "condicionDestinoUsuarioEgreso indica derivación a otro servicio, pero "
                                    "el usuario no tiene registros en hospitalización ni procedimientos "
                                    "(RVC062 pasa de Notificación a Rechazo en 0948).", cond)
    return errores


# ────────────────────────────────────────────────────────────────
# BLOQUE H: HOSPITALIZACIÓN
# ────────────────────────────────────────────────────────────────

def validar_hospitalizacion_malla_0948(data, nombre_archivo=""):
    """
    Hospitalización bajo 0948:
    - Campos nuevos H16-H30 (CIE11 x2 pares principal + relacionados +
      complicación + causa muerte + codigoVIDA).
    - Tamaño exacto H06/H07 (4), H08-H11/H13 (0,4) — informativo.
    - RVC088/RVC089 escaladas (relacionados de egreso == principal / repetidos).
    - RVC053/RVC062 escaladas, análogas a urgencias.
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        regs = servicios.get("hospitalizacion", [])
        if not isinstance(regs, list):
            continue

        fecha_falle_dt = None
        for h in regs:
            if isinstance(h, dict) and _condicion_indica_muerto_0948(h.get("condicionDestinoUsuarioEgreso")):
                feg = normalizar_str(h.get("fechaEgreso") or "")
                try:
                    if len(feg) >= 16:
                        fecha_falle_dt = datetime.strptime(feg[:16], "%Y-%m-%d %H:%M")
                    elif len(feg) == 10:
                        fecha_falle_dt = datetime.strptime(feg, "%Y-%m-%d")
                except ValueError:
                    pass
                break

        for h in regs:
            if not isinstance(h, dict):
                continue
            ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
                   "num_doc": num_doc, "consecutivo": normalizar_str(h.get("consecutivo"))}

            _campo_nuevo_0948(errores, ctx, "codDiagnosticoPrincipalECIE11",
                               h.get("codDiagnosticoPrincipalECIE11"), 4, 256, id_regla="H16-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoPrincipalECIE11",
                               h.get("nomCodDiagnosticoPrincipalECIE11"), 1, 1000, id_regla="H17-0948")
            for n, letra in ((1, "H20"), (2, "H22"), (3, "H24")):
                _campo_nuevo_0948(errores, ctx, f"codDiagnosticoRelacionadoE{n}CIE11",
                                   h.get(f"codDiagnosticoRelacionadoE{n}CIE11"),
                                   opcional_transicion=True, id_regla=f"{letra}-0948")
            for n, letra in ((1, "H21"), (2, "H23"), (3, "H25")):
                _campo_nuevo_0948(errores, ctx, f"nomCodDiagnosticoRelacionadoE{n}CIE11",
                                   h.get(f"nomCodDiagnosticoRelacionadoE{n}CIE11"),
                                   opcional_transicion=True, id_regla=f"{letra}-0948")
            _campo_nuevo_0948(errores, ctx, "codComplicacionCIE11",
                               h.get("codComplicacionCIE11"), 4, 256, id_regla="H26-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodComplicacionCIE11",
                               h.get("nomCodComplicacionCIE11"), 1, 1000, id_regla="H27-0948")
            _campo_nuevo_0948(errores, ctx, "codDiagnosticoCausaMuerteCIE11",
                               h.get("codDiagnosticoCausaMuerteCIE11"),
                               opcional_transicion=True, id_regla="H28-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoCausaMuerteCIE11",
                               h.get("nomCodDiagnosticoCausaMuerteCIE11"),
                               opcional_transicion=True, id_regla="H29-0948")
            _campo_nuevo_0948(errores, ctx, "codigoVIDA", h.get("codigoVIDA"),
                               1, 256, opcional_transicion=True, id_regla="H30-0948")

            _tamano_cambio_0948(errores, ctx, "codDiagnosticoPrincipal",
                                 h.get("codDiagnosticoPrincipal"), 4, "H06-0948", permite_null=False)
            _tamano_cambio_0948(errores, ctx, "codDiagnosticoPrincipalE",
                                 h.get("codDiagnosticoPrincipalE"), 4, "H07-0948", permite_null=False)
            for campo, id_r in (("codDiagnosticoRelacionadoE1", "H08-0948"),
                                 ("codDiagnosticoRelacionadoE2", "H09-0948"),
                                 ("codDiagnosticoRelacionadoE3", "H10-0948"),
                                 ("codComplicacion", "H11-0948"),
                                 ("codDiagnosticoCausaMuerte", "H13-0948")):
                _tamano_cambio_0948(errores, ctx, campo, h.get(campo), 4, id_r)

            diag_pe = normalizar_str(h.get("codDiagnosticoPrincipalE") or "")
            rel_e = [normalizar_str(h.get(f"codDiagnosticoRelacionadoE{n}") or "") for n in (1, 2, 3)]
            rel_e_notnull = [d for d in rel_e if d]
            for dr in rel_e_notnull:
                if dr == diag_pe:
                    _escalada_0948(errores, ctx, "RVC088", "codDiagnosticoRelacionadoE",
                                    f"Diagnóstico relacionado de egreso '{dr}' es igual al principal de "
                                    "egreso (RVC088 pasa de Notificación a Rechazo en 0948).", dr)
            if len(rel_e_notnull) != len(set(rel_e_notnull)):
                _escalada_0948(errores, ctx, "RVC089", "codDiagnosticoRelacionadoE",
                                "Existen diagnósticos relacionados de egreso repetidos entre sí "
                                "(RVC089 pasa de Notificación a Rechazo en 0948).")

            cond = h.get("condicionDestinoUsuarioEgreso")
            if fecha_falle_dt is not None and not _condicion_indica_muerto_0948(cond):
                fecha_serv_raw = normalizar_str(h.get("fechaInicioAtencion") or "")
                fecha_serv_dt = None
                try:
                    if len(fecha_serv_raw) >= 16:
                        fecha_serv_dt = datetime.strptime(fecha_serv_raw[:16], "%Y-%m-%d %H:%M")
                    elif len(fecha_serv_raw) == 10:
                        fecha_serv_dt = datetime.strptime(fecha_serv_raw, "%Y-%m-%d")
                except ValueError:
                    pass
                if fecha_serv_dt and (fecha_serv_dt - fecha_falle_dt).total_seconds() > 24 * 3600:
                    _escalada_0948(errores, ctx, "RVC053", "fechaInicioAtencion",
                                    "El usuario tiene un registro de hospitalización marcado como fallecido y "
                                    "este otro servicio tiene fecha posterior a 24h del fallecimiento "
                                    "(RVC053 pasa de Notificación a Rechazo en 0948).", fecha_serv_raw)

            if _condicion_indica_derivado_0948(cond):
                otros_bloques = servicios.get("urgencias", []) or servicios.get("procedimientos", [])
                if not otros_bloques:
                    _escalada_0948(errores, ctx, "RVC062", "condicionDestinoUsuarioEgreso",
                                    "condicionDestinoUsuarioEgreso indica derivación a otro servicio, pero "
                                    "el usuario no tiene registros en urgencias ni procedimientos "
                                    "(RVC062 pasa de Notificación a Rechazo en 0948).", cond)
    return errores


# ────────────────────────────────────────────────────────────────
# BLOQUE N: RECIÉN NACIDOS
# ────────────────────────────────────────────────────────────────

def validar_recien_nacidos_malla_0948(data, nombre_archivo=""):
    """
    Recién nacidos bajo 0948:
    - Campos nuevos N14-N18 (CIE11 principal, causa muerte, codigoVIDA).
    - RVC057 (edadGestacional 20-46) y RVC058 (peso 500-5000g) escaladas a
      Rechazo (ya existen en la malla 2275 como "media").
    - RVC053/RVC062 escaladas, análogas a urgencias/hospitalización.
    - RVC097 (nueva, N06 numConsultasCPrenatal <= 15), implementada literal.
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        rn_list = servicios.get("recienNacidos", [])
        if not isinstance(rn_list, list):
            continue

        fecha_falle_dt = None
        for n in rn_list:
            if isinstance(n, dict) and _condicion_indica_muerto_0948(n.get("condicionDestinoUsuarioEgreso")):
                feg = normalizar_str(n.get("fechaEgreso") or "")
                try:
                    if len(feg) >= 16:
                        fecha_falle_dt = datetime.strptime(feg[:16], "%Y-%m-%d %H:%M")
                    elif len(feg) == 10:
                        fecha_falle_dt = datetime.strptime(feg, "%Y-%m-%d")
                except ValueError:
                    pass
                break

        for n in rn_list:
            if not isinstance(n, dict):
                continue
            ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
                   "num_doc": num_doc, "consecutivo": normalizar_str(n.get("consecutivo"))}

            _campo_nuevo_0948(errores, ctx, "codDiagnosticoPrincipalECIE11",
                               n.get("codDiagnosticoPrincipalECIE11"), 4, 256, id_regla="N14-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoPrincipalECIE11",
                               n.get("nomCodDiagnosticoPrincipalECIE11"), 1, 1000, id_regla="N15-0948")
            _campo_nuevo_0948(errores, ctx, "codDiagnosticoCausaMuerteCIE11",
                               n.get("codDiagnosticoCausaMuerteCIE11"),
                               opcional_transicion=True, id_regla="N16-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoCausaMuerteCIE11",
                               n.get("nomCodDiagnosticoCausaMuerteCIE11"),
                               opcional_transicion=True, id_regla="N17-0948")
            _campo_nuevo_0948(errores, ctx, "codigoVIDA", n.get("codigoVIDA"),
                               1, 256, opcional_transicion=True, id_regla="N18-0948")

            # RVC057: edadGestacional 20-46
            eg_raw = n.get("edadGestacional")
            if eg_raw is not None:
                try:
                    eg = int(str(eg_raw).strip())
                    if not (20 <= eg <= 46):
                        _escalada_0948(errores, ctx, "RVC057", "edadGestacional",
                                        f"edadGestacional {eg} no está entre 20 y 46 semanas "
                                        "(RVC057 pasa de Notificación a Rechazo en 0948).", eg_raw)
                except (ValueError, TypeError):
                    pass

            # RVC058: peso 500-5000g
            peso_raw = n.get("peso")
            if peso_raw is not None:
                try:
                    peso = float(str(peso_raw).strip())
                    if not (500 <= peso <= 5000):
                        _escalada_0948(errores, ctx, "RVC058", "peso",
                                        f"Peso del recién nacido {peso}g no está entre 500 y 5000 gramos "
                                        "(RVC058 pasa de Notificación a Rechazo en 0948).", peso_raw)
                except (ValueError, TypeError):
                    pass

            # RVC097 (nueva 0948): numConsultasCPrenatal <= 15
            ncp_raw = n.get("numConsultasCPrenatal")
            if ncp_raw is not None:
                try:
                    ncp = int(str(ncp_raw).strip())
                    if ncp > 15:
                        errores.append({
                            **ctx, "id_regla": "RVC097", "severidad": "critica",
                            "campo": "numConsultasCPrenatal",
                            "mensaje": f"numConsultasCPrenatal={ncp} supera el máximo permitido de 15 "
                                       "consultas de cuidado prenatal (RVC097, nueva en 0948).",
                            "valor_actual": normalizar_str(ncp_raw),
                        })
                except (ValueError, TypeError):
                    errores.append({
                        **ctx, "id_regla": "RVC097", "severidad": "alta",
                        "campo": "numConsultasCPrenatal",
                        "mensaje": "numConsultasCPrenatal debe ser numérico (RVC097, nueva en 0948).",
                        "valor_actual": normalizar_str(ncp_raw),
                    })

            cond = n.get("condicionDestinoUsuarioEgreso")
            if fecha_falle_dt is not None and not _condicion_indica_muerto_0948(cond):
                fecha_serv_raw = normalizar_str(n.get("fechaNacimiento") or "")
                fecha_serv_dt = None
                try:
                    if len(fecha_serv_raw) >= 16:
                        fecha_serv_dt = datetime.strptime(fecha_serv_raw[:16], "%Y-%m-%d %H:%M")
                    elif len(fecha_serv_raw) == 10:
                        fecha_serv_dt = datetime.strptime(fecha_serv_raw, "%Y-%m-%d")
                except ValueError:
                    pass
                if fecha_serv_dt and (fecha_serv_dt - fecha_falle_dt).total_seconds() > 24 * 3600:
                    _escalada_0948(errores, ctx, "RVC053", "fechaNacimiento",
                                    "El usuario tiene un registro de recién nacido marcado como fallecido y "
                                    "este otro registro tiene fecha posterior a 24h del fallecimiento "
                                    "(RVC053 pasa de Notificación a Rechazo en 0948).", fecha_serv_raw)

            if _condicion_indica_derivado_0948(cond):
                otros_bloques = servicios.get("hospitalizacion", []) or servicios.get("urgencias", [])
                if not otros_bloques:
                    _escalada_0948(errores, ctx, "RVC062", "condicionDestinoUsuarioEgreso",
                                    "condicionDestinoUsuarioEgreso indica derivación a otro servicio, pero "
                                    "el usuario no tiene registros en hospitalización ni urgencias "
                                    "(RVC062 pasa de Notificación a Rechazo en 0948).", cond)
    return errores


# ────────────────────────────────────────────────────────────────
# BLOQUE M: MEDICAMENTOS
# ────────────────────────────────────────────────────────────────

def validar_medicamentos_malla_0948(data, nombre_archivo=""):
    """
    Medicamentos bajo 0948:
    - Campos nuevos M24 (opcional en transición) / M25 (obligatorio si M24
      diligenciado) / M26 / M27 / M28 vrDispensacion / M29 codigoVIDA.
    - RVC063 (M08 código IUM/CUM debe existir en SISPRO) escalada a Rechazo:
      no hay catálogo SISPRO disponible en este repo, así que se limita a
      re-emitir el chequeo de formato ya existente (M08 informado) con
      severidad escalada, dejando constancia de que el cruce contra SISPRO
      no puede verificarse aquí.
    - RVC094 (nueva 0948): vrServicio == cantidadMedicamento * vrUnitMedicamento
      cuando modalidadPago='04' y tipoMedicamento != '03'. La modalidad de
      pago no vive en el bloque de medicamentos del RIPS (es un dato de la
      FEV/contrato); se usa modalidadGrupoServicioTecSal como mejor proxy
      disponible en el propio registro y se documenta la limitación.
    - M02 (numAutorizacion) NO se valida aquí: fue eliminado en 0948.
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        meds = servicios.get("medicamentos", [])
        if not isinstance(meds, list):
            continue
        for med in meds:
            if not isinstance(med, dict):
                continue
            ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
                   "num_doc": num_doc, "consecutivo": normalizar_str(med.get("consecutivo"))}

            # M24: explícitamente "opcional durante la transición de la norma"
            m24 = normalizar_str(med.get("codDiagnosticoPrincipalECIE11") or "")
            _campo_nuevo_0948(errores, ctx, "codDiagnosticoPrincipalECIE11", m24,
                               4, 256, opcional_transicion=True, id_regla="M24-0948")

            # M25: "Campo obligatorio cuando el campo M24 es diligenciado."
            m25 = normalizar_str(med.get("nomCodDiagnosticoPrincipalECIE11") or "")
            if m24 and m24.lower() not in {"none", "null"} and not m25:
                errores.append({
                    **ctx, "id_regla": "M25-0948", "severidad": "media",
                    "campo": "nomCodDiagnosticoPrincipalECIE11",
                    "mensaje": "[0948] nomCodDiagnosticoPrincipalECIE11 (M25) es obligatorio "
                               "cuando codDiagnosticoPrincipalECIE11 (M24) está diligenciado.",
                    "valor_actual": "",
                })

            _campo_nuevo_0948(errores, ctx, "codDiagnosticoRelacionadoCIE11",
                               med.get("codDiagnosticoRelacionadoCIE11"),
                               opcional_transicion=True, id_regla="M26-0948")
            _campo_nuevo_0948(errores, ctx, "nomCodDiagnosticoRelacionadoCIE11",
                               med.get("nomCodDiagnosticoRelacionadoCIE11"),
                               opcional_transicion=True, id_regla="M27-0948")

            # M28: vrDispensacion, N 0-15; "en caso de no aplicar diligenciar en 0".
            # No se reporta la simple ausencia del campo (es normal en RIPS 2275,
            # donde M28 aún no existe); solo se valida el formato cuando sí viene informado.
            vrd = med.get("vrDispensacion")
            if vrd is not None:
                try:
                    if float(vrd) < 0:
                        errores.append({
                            **ctx, "id_regla": "M28-0948", "severidad": "media",
                            "campo": "vrDispensacion",
                            "mensaje": "[0948] vrDispensacion no puede ser negativo.",
                            "valor_actual": normalizar_str(vrd),
                        })
                except (ValueError, TypeError):
                    errores.append({
                        **ctx, "id_regla": "M28-0948", "severidad": "media",
                        "campo": "vrDispensacion",
                        "mensaje": "[0948] vrDispensacion debe ser numérico.",
                        "valor_actual": normalizar_str(vrd),
                    })

            _campo_nuevo_0948(errores, ctx, "codigoVIDA", med.get("codigoVIDA"),
                               1, 256, opcional_transicion=True, id_regla="M29-0948")

            # RVC063 escalada (best-effort: sin catálogo SISPRO disponible)
            cod_tec = normalizar_str(med.get("codTecnologiaSalud") or "")
            if not cod_tec:
                _escalada_0948(errores, ctx, "RVC063", "codTecnologiaSalud",
                                "codTecnologiaSalud (M08) no está informado; en 0948 su validación contra "
                                "el catálogo SISPRO (IUM/CUM) pasa de Notificación a Rechazo (RVC063). "
                                "Nota: este pipeline no cuenta con el catálogo SISPRO para verificar la "
                                "existencia del código; solo se controla su presencia.")

            # RVC094 (nueva 0948): vrServicio == cantidadMedicamento * vrUnitMedicamento
            # cuando modalidad de pago='04' y tipoMedicamento != '03'.
            tipo_med = normalizar_str(med.get("tipoMedicamento") or "")
            modalidad = normalizar_str(med.get("modalidadGrupoServicioTecSal") or "")
            if modalidad == "04" and tipo_med != "03":
                cant_raw = med.get("cantidadMedicamento")
                vru_raw = med.get("vrUnitMedicamento")
                vrs_raw = med.get("vrServicio")
                if cant_raw is not None and vru_raw is not None and vrs_raw is not None:
                    try:
                        cant = float(str(cant_raw).strip())
                        vru = float(str(vru_raw).strip())
                        vrs = float(str(vrs_raw).strip())
                        esperado = round(cant * vru, 2)
                        if round(vrs, 2) != esperado:
                            errores.append({
                                **ctx, "id_regla": "RVC094", "severidad": "critica",
                                "campo": "vrServicio",
                                "mensaje": f"vrServicio ({vrs}) debe ser igual a cantidadMedicamento "
                                           f"({cant}) * vrUnitMedicamento ({vru}) = {esperado} cuando "
                                           "modalidadGrupoServicioTecSal='04' y tipoMedicamento≠'03' "
                                           "(RVC094, nueva en 0948).",
                                "valor_actual": normalizar_str(vrs_raw),
                            })
                    except (ValueError, TypeError):
                        pass
    return errores


# ────────────────────────────────────────────────────────────────
# BLOQUE S: OTROS SERVICIOS
# ────────────────────────────────────────────────────────────────

def validar_otros_servicios_malla_0948(data, nombre_archivo=""):
    """
    Otros servicios bajo 0948:
    - Campos nuevos S17 codigoVIDA, S18 vrDispensacion.
    - RVC066 (S07 obligatorio si tipoOS=01) escalada a Rechazo — ya existe
      un chequeo equivalente de obligatoriedad condicional en 2275; aquí se
      re-emite con severidad escalada.
    - RVC096 (S06 codTecnologiaSalud contra catálogo CUPS): SOLO se aplica
      cuando tipoOS en {02 Traslados, 03 Estancias, 05 Honorarios}, según
      el anexo técnico ("el código del traslado, transporte o estancia
      debe corresponder al código CUPS"; "el código de honorarios debe
      corresponder al código CUPS del procedimiento"). Se omite para
      tipoOS=01 (dispositivos médicos e insumos: código IDM/UDI o propio,
      no CUPS) y tipoOS=04 (servicios complementarios: código MIPRES, no
      CUPS), para evitar falsos positivos.
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)

    for usuario in data.get("usuarios", []):
        if not isinstance(usuario, dict):
            continue
        num_doc = _num_doc_val(usuario)
        servicios = usuario.get("servicios", {})
        if not isinstance(servicios, dict):
            continue
        otros = servicios.get("otrosServicios", [])
        if not isinstance(otros, list):
            continue
        for svc in otros:
            if not isinstance(svc, dict):
                continue
            ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
                   "num_doc": num_doc, "consecutivo": normalizar_str(svc.get("consecutivo"))}
            _campo_nuevo_0948(errores, ctx, "codigoVIDA", svc.get("codigoVIDA"),
                               1, 256, opcional_transicion=True, id_regla="S17-0948")
            # No se reporta la simple ausencia del campo (es normal en RIPS 2275,
            # donde S18 aún no existe); solo se valida el formato cuando sí viene informado.
            vrd = svc.get("vrDispensacion")
            if vrd is not None:
                try:
                    if float(vrd) < 0:
                        errores.append({
                            **ctx, "id_regla": "S18-0948", "severidad": "media",
                            "campo": "vrDispensacion",
                            "mensaje": "[0948] vrDispensacion no puede ser negativo.",
                            "valor_actual": normalizar_str(vrd),
                        })
                except (ValueError, TypeError):
                    errores.append({
                        **ctx, "id_regla": "S18-0948", "severidad": "media",
                        "campo": "vrDispensacion",
                        "mensaje": "[0948] vrDispensacion debe ser numérico.",
                        "valor_actual": normalizar_str(vrd),
                    })

            tipo_os = normalizar_str(svc.get("tipoOS") or "")
            s07 = normalizar_str(svc.get("nomTecnologiaSalud") or "")
            if tipo_os == "01" and (not s07 or s07.lower() in {"none", "null"}):
                _escalada_0948(errores, ctx, "RVC066", "nomTecnologiaSalud",
                                "nomTecnologiaSalud (S07) es obligatorio cuando tipoOS='01' (Dispositivos "
                                "médicos e insumos) (RVC066 pasa de Notificación a Rechazo en 0948).")

            # RVC096 (nueva 0948): codTecnologiaSalud contra catálogo CUPS,
            # solo para tipoOS de traslados/estancias/honorarios (ver docstring).
            if tipo_os in TIPOS_OS_VALIDABLES_CONTRA_CUPS_0948:
                cups_validos = _cups_codigos_validos_0948()
                cod_tec_os = normalizar_str(svc.get("codTecnologiaSalud") or "")
                if cups_validos and cod_tec_os and cod_tec_os not in cups_validos:
                    _escalada_0948(errores, ctx, "RVC096", "codTecnologiaSalud",
                                    f"codTecnologiaSalud '{cod_tec_os}' (tipoOS='{tipo_os}') no existe en el "
                                    "catálogo oficial CUPS (RVC096, nueva en 0948).", cod_tec_os)
    return errores


# ────────────────────────────────────────────────────────────────
# BLOQUE U: USUARIOS
# ────────────────────────────────────────────────────────────────

def validar_usuarios_malla_0948(data, nombre_archivo=""):
    """
    Usuarios bajo 0948:
    - U12 registroSIRAS (RVC095): obligatorio si tipoUsuario='10', null en
      caso contrario. Notificación durante los primeros 3 meses desde la
      implementación (14-may-2026 + 3 meses ≈ 14-ago-2026), luego Rechazo.
    - RVC007 (edad vs. tipo de documento, tolerancia 1 año menos 1 día):
      ajuste en Regla/Mensaje según 3.2.3; se re-emite el chequeo existente
      en 2275 con severidad escalada y el mensaje ajustado.
    - RVG14 (nacimientos múltiples ⇒ procedimiento de parto múltiple):
      escalada de Notificación a Rechazo en 0948.
    - RVG15 (urgencias con observación ⇒ consulta de urgencia asociada):
      cambia de Rechazo(2275) a Notificación(0948), severidad fija ("media").
    - RVC098 (fecha de servicio no debe superar fecha de fallecimiento +
      24h): usa como proxy fechaEgreso del registro (urgencias/
      hospitalizacion/recienNacidos) cuyo condicionDestinoUsuarioEgreso
      indique "PACIENTE MUERTO" (código '02'), por decisión explícita del
      usuario dado que el RIPS no trae un campo dedicado de fecha de
      fallecimiento (RUAF-ND no está disponible en este pipeline).
    """
    errores = []
    if not isinstance(data, dict):
        return errores
    num_factura = _ctx_base_0948(nombre_archivo, data)
    hoy = _date.today()

    for u in data.get("usuarios", []):
        if not isinstance(u, dict):
            continue
        num_doc = _num_doc_val(u)
        ctx = {"archivo": nombre_archivo, "num_factura": num_factura,
               "num_doc": num_doc, "consecutivo": normalizar_str(u.get("consecutivo"))}
        tipo_usr = normalizar_str(u.get("tipoUsuario") or "")
        siras = normalizar_str(u.get("registroSIRAS") or "")
        es_null = (not siras) or siras.lower() in {"none", "null"}
        severidad_rvc095 = "critica" if hoy >= FECHA_RVC095_RECHAZO else "media"

        if tipo_usr == "10":
            if es_null:
                errores.append({
                    **ctx, "id_regla": "RVC095", "severidad": severidad_rvc095,
                    "campo": "registroSIRAS",
                    "mensaje": "registroSIRAS (U12) es obligatorio cuando tipoUsuario='10' "
                               "(Tomador/Amparado SOAT) (RVC095, nueva en 0948; Notificación durante los "
                               "primeros 3 meses de vigencia, luego Rechazo).",
                    "valor_actual": "",
                })
            elif not (1 <= len(siras) <= 60):
                errores.append({
                    **ctx, "id_regla": "RVC095", "severidad": "baja",
                    "campo": "registroSIRAS",
                    "mensaje": f"registroSIRAS debe tener entre 1 y 60 caracteres (actual: {len(siras)}).",
                    "valor_actual": siras,
                })
        else:
            if not es_null:
                errores.append({
                    **ctx, "id_regla": "RVC095", "severidad": severidad_rvc095,
                    "campo": "registroSIRAS",
                    "mensaje": "registroSIRAS debe informarse null cuando tipoUsuario no es '10' "
                               "(Tomador/Amparado SOAT) (RVC095, nueva en 0948).",
                    "valor_actual": siras,
                })

        # RVC007 escalada: edad según fechaNacimiento coherente con tipoDocumentoIdentificacion,
        # con tolerancia de un año menos un día (mismo criterio ya usado en la malla 2275).
        fecha_nac_raw = normalizar_str(u.get("fechaNacimiento") or "")
        tipo_doc_u = _tipo_doc(u)
        if len(fecha_nac_raw) == 10 and tipo_doc_u:
            try:
                fecha_nac_dt = datetime.strptime(fecha_nac_raw, "%Y-%m-%d")
                edad_dias = (datetime.now() - fecha_nac_dt).days
                edad_anios = edad_dias / 365.25
                permitidos_edad = None
                if tipo_doc_u == "RC":
                    permitidos_edad = (0, 7)          # hasta 6 años + tolerancia
                elif tipo_doc_u == "TI":
                    permitidos_edad = (6, 18)          # 7-17 + tolerancia
                elif tipo_doc_u == "CC":
                    permitidos_edad = (17, 200)
                if permitidos_edad and not (permitidos_edad[0] <= edad_anios <= permitidos_edad[1]):
                    _escalada_0948(errores, ctx, "RVC007", "tipoDocumentoIdentificacion",
                                    f"La edad calculada ({edad_anios:.1f} años) no es coherente con "
                                    f"tipoDocumentoIdentificacion='{tipo_doc_u}', considerando la tolerancia "
                                    "de un año menos un día (RVC007, ajuste de Regla/Mensaje en 0948).",
                                    tipo_doc_u)
            except ValueError:
                pass

        # ── RVG14 / RVG15 / RVC098: chequeos cruzados entre bloques de
        # servicios del mismo usuario ─────────────────────────────────
        servicios_u = u.get("servicios", {})
        if not isinstance(servicios_u, dict):
            servicios_u = {}
        consultas_u = servicios_u.get("consultas", []) or []
        procedimientos_u = servicios_u.get("procedimientos", []) or []
        urgencias_u = servicios_u.get("urgencias", []) or []
        hospitalizacion_u = servicios_u.get("hospitalizacion", []) or []
        recien_nacidos_u = servicios_u.get("recienNacidos", []) or []
        medicamentos_u = servicios_u.get("medicamentos", []) or []
        otros_servicios_u = servicios_u.get("otrosServicios", []) or []

        # RVG14: nacimientos múltiples ⇒ procedimiento de parto múltiple.
        if isinstance(recien_nacidos_u, list) and len(recien_nacidos_u) >= 2:
            tiene_parto_multiple = any(
                isinstance(p, dict) and normalizar_str(p.get("codProcedimiento") or "") in CUPS_PARTO_MULTIPLE_0948
                for p in procedimientos_u if isinstance(procedimientos_u, list)
            )
            if not tiene_parto_multiple:
                _escalada_0948(errores, ctx, "RVG14", "codProcedimiento",
                                f"El usuario tiene {len(recien_nacidos_u)} registros de recienNacidos "
                                "(nacimiento múltiple) pero no hay un procedimiento de parto múltiple "
                                f"({sorted(CUPS_PARTO_MULTIPLE_0948)}) asociado "
                                "(RVG14 pasa de Notificación a Rechazo en 0948).")

        # RVG15: urgencias con observación ⇒ debe existir consulta de urgencia
        # asociada con fecha/hora >= ingreso a observación.
        if isinstance(urgencias_u, list):
            for r in urgencias_u:
                if not isinstance(r, dict):
                    continue
                fecha_ingreso_urg = _parse_fecha_0948(r.get("fechaInicioAtencion"))
                if fecha_ingreso_urg is None:
                    continue
                tiene_consulta_urgencia = False
                if isinstance(consultas_u, list):
                    for c in consultas_u:
                        if not isinstance(c, dict):
                            continue
                        if normalizar_str(c.get("codConsulta") or "") not in CUPS_CONSULTA_URGENCIAS_0948:
                            continue
                        fecha_consulta = _parse_fecha_0948(c.get("fechaInicioAtencion"))
                        if fecha_consulta is not None and fecha_consulta >= fecha_ingreso_urg:
                            tiene_consulta_urgencia = True
                            break
                if not tiene_consulta_urgencia:
                    errores.append({
                        **ctx, "id_regla": "RVG15", "severidad": "media",
                        "campo": "codConsulta",
                        "mensaje": "El usuario tiene un registro de urgencias con observación pero no se "
                                   "encontró una consulta de urgencia asociada (codConsulta en "
                                   f"{sorted(CUPS_CONSULTA_URGENCIAS_0948)}) con fecha/hora posterior o "
                                   "igual al ingreso a observación (RVG15 cambia de Rechazo en 2275 a "
                                   "Notificación en 0948).",
                        "valor_actual": "",
                    })

        # RVC098: fecha de servicio no debe superar fecha de fallecimiento + 24h.
        # Proxy aprobado: fechaEgreso del registro (urgencias/hospitalizacion/
        # recienNacidos) cuyo condicionDestinoUsuarioEgreso='02' (PACIENTE MUERTO).
        fecha_muerte_dt = None
        for bloque in (urgencias_u, hospitalizacion_u, recien_nacidos_u):
            if not isinstance(bloque, list):
                continue
            for reg in bloque:
                if isinstance(reg, dict) and _condicion_indica_muerto_0948(reg.get("condicionDestinoUsuarioEgreso")):
                    dt = _parse_fecha_0948(reg.get("fechaEgreso"))
                    if dt is not None:
                        fecha_muerte_dt = dt
                        break
            if fecha_muerte_dt is not None:
                break

        if fecha_muerte_dt is not None:
            limite = fecha_muerte_dt + timedelta(hours=24)
            candidatos = []
            for c in consultas_u if isinstance(consultas_u, list) else []:
                if isinstance(c, dict):
                    candidatos.append((c, "fechaInicioAtencion", c.get("fechaInicioAtencion")))
            for p in procedimientos_u if isinstance(procedimientos_u, list) else []:
                if isinstance(p, dict):
                    candidatos.append((p, "fechaInicioAtencion", p.get("fechaInicioAtencion")))
            for m in medicamentos_u if isinstance(medicamentos_u, list) else []:
                if isinstance(m, dict):
                    candidatos.append((m, "fechaDispensAdmon", m.get("fechaDispensAdmon")))
            for s in otros_servicios_u if isinstance(otros_servicios_u, list) else []:
                if isinstance(s, dict):
                    candidatos.append((s, "fechaSuministroTecnologia", s.get("fechaSuministroTecnologia")))

            for reg, campo_fecha, valor_fecha in candidatos:
                fecha_serv_dt = _parse_fecha_0948(valor_fecha)
                if fecha_serv_dt is not None and fecha_serv_dt > limite:
                    ctx_reg = {"archivo": nombre_archivo, "num_factura": num_factura,
                               "num_doc": num_doc, "consecutivo": normalizar_str(reg.get("consecutivo"))}
                    errores.append({
                        **ctx_reg, "id_regla": "RVC098", "severidad": "critica",
                        "campo": campo_fecha,
                        "mensaje": f"'{campo_fecha}'='{normalizar_str(valor_fecha)}' es posterior a la fecha "
                                   f"de fallecimiento del usuario ('{fecha_muerte_dt.strftime('%Y-%m-%d %H:%M')}') "
                                   "más 24 horas (RVC098, nueva en 0948; proxy: fechaEgreso del registro con "
                                   "condicionDestinoUsuarioEgreso='02' PACIENTE MUERTO).",
                        "valor_actual": normalizar_str(valor_fecha),
                    })
    return errores

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)